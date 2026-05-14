"""Tool registry for the qwen agent.

Add new tools by:
  1. writing a function below
  2. adding it to DISPATCH
  3. adding its JSON schema to TOOLS

Heavy deps (crawl4ai, ddgs) are imported lazily so the agent starts fast.
"""

from __future__ import annotations

import asyncio
import json
import os
import re
import shutil
import subprocess
import sys
import threading
import urllib.request
from urllib.parse import urlparse
from pathlib import Path
from typing import Any

# Limits cribbed from Claude Code's tools.
READ_DEFAULT_LINES = 2000        # default lines returned from read_file
READ_MAX_BYTES = 256 * 1024      # hard cap on file size before truncation
GREP_DEFAULT_HEAD = 250          # default cap on grep result lines
GREP_LINE_MAX_CHARS = 500        # truncate matched lines longer than this
VCS_EXCLUDES = (".git", ".hg", ".svn", ".bzr", ".jj", ".sl")

# Shared "refused" markers. The system prompt at scripts/agent.py:71-76
# trains the model that `[REFUSED — ... cap reached]` is a HARD STOP. The
# cap-exhaustion nudge in scripts/agent.py:1003-1020 detects refusals by
# matching this exact prefix on tool-result content. Keep both REFUSED_PREFIX
# (the canonical emit prefix used by every cap site here) and
# REFUSED_DETECT_PREFIX (the broader stem used by .startswith checks so a
# future ASCII-dash variant or whitespace tweak doesn't silently slip past
# detection) in one place so a contributor adding a new refusal point can
# pattern-match through grep without missing the harness contract.
REFUSED_PREFIX = "[REFUSED — "
REFUSED_DETECT_PREFIX = "[REFUSED"


# ---------- web ------------------------------------------------------------
#
# Three primary tools live here:
#  - web_search:  DuckDuckGo, with semantic + lexical near-duplicate query
#                 detection so the model isn't rewarded for re-asking the
#                 same question with a synonym.
#  - web_fetch:   article-aware HTML extraction. Full stripped body (up to
#                 3 MB) is fed into `condense_tool_result`, which uses
#                 BM25 lexical scoring + an ONNX cross-encoder reranker
#                 to keep the chunks most relevant to the task. Memory
#                 bound stays tight because the final result returned to
#                 the model is capped at TOOL_RESULT_MAX (~60 k).
#  - web_outline: lightweight "what's on this page?" tool that returns just
#                 the heading hierarchy + a few words of body so the model
#                 can decide whether a fetch is worthwhile before paying
#                 the full content-extraction cost.

# Default per-call truncation cap. Iter 39: bumped 300 k → 3 M to keep the
# entire stripped body of any reasonable filing/paper/doc intact through
# `_smart_truncate`, since `condense_tool_result` now uses BM25 + an ONNX
# cross-encoder reranker (Xenova/ms-marco-MiniLM-L-6-v2, 22 MB int8) to
# pick the relevant chunks from the full body. The old 300 k cap forced
# `_smart_truncate` to keep head 55% + tail 45% and drop the middle —
# exactly where SEC 10-K financial-statement notes (debt schedules,
# convertible-notes tables, segment data) live, so the reranker was being
# fed the wrong document slice. Memory bound stays tight: 3 MB of stripped
# text is well within Python's comfort zone, BM25 over ~2 k chunks
# completes in <30 ms, the reranker scores only the top-32 BM25 hits
# (~50 ms), and `condense_tool_result` caps the final output to the
# agent's TOOL_RESULT_MAX (~60 k). `_smart_truncate` remains as a safety
# net for pages above this ceiling but should rarely fire in practice.
WEB_FETCH_DEFAULT_MAX = int(os.environ.get("QWEN_WEB_FETCH_MAX_CHARS", "3000000"))
WEB_FETCH_HEAD_FRACTION = 0.55  # bias toward the start (titles / abstracts)


def _normalize_url(url: str) -> str:
    """Best-effort URL canonicalization so cache lookups hit on
    http/https-and-trailing-slash variants and tracking-param noise.
    Only used as the cache key — the actual fetch keeps the original URL.
    """
    try:
        from urllib.parse import urlsplit, urlunsplit, parse_qsl, urlencode
    except ImportError:
        return url
    try:
        parts = urlsplit(url.strip())
    except Exception:  # noqa: BLE001
        return url
    if not parts.scheme:
        # Treat protocol-less input as https
        parts = urlsplit("https://" + url.strip())
    # Drop common tracking params; keep doc identifiers.
    drop = {"utm_source", "utm_medium", "utm_campaign", "utm_content",
            "utm_term", "fbclid", "gclid", "mc_cid", "mc_eid",
            "ref", "ref_src", "_hsenc", "_hsmi"}
    qs = [(k, v) for k, v in parse_qsl(parts.query, keep_blank_values=False)
          if k not in drop]
    qs.sort()
    norm_query = urlencode(qs)
    norm_path = re.sub(r"/+$", "", parts.path) or "/"
    return urlunsplit((parts.scheme.lower(), parts.netloc.lower(),
                       norm_path, norm_query, ""))


# In-process semantic-dup memo for web_search. Map: query → (qvec, hit_summary)
# so we can compare a NEW query against PRIOR queries without re-embedding
# the prior ones. Bounded at 256 entries.
_SEARCH_QUERY_MEMO: list[tuple[str, "object", str]] = []
_SEARCH_QUERY_MEMO_MAX = 256
# Bumped 0.92 → 0.97 after observing legit variant queries being blocked at
# cosine 0.92–0.95 (e.g. "<COMPANY> Q4 2024 shares repurchased common stock"
# vs "<COMPANY> Q4 2024 quarterly report repurchase of common stock shares" —
# different intent, different target documents, but high keyword overlap).
# 0.97 still catches true near-duplicates ("X shares repurchased" vs
# "shares repurchased X") while letting through reformulations that target
# different sources or different time periods.
_SEARCH_DUP_THRESHOLD = float(os.environ.get("QWEN_SEARCH_DUP_THRESHOLD", "0.97"))


def _search_query_norm(q: str) -> str:
    """Lowercase + collapse whitespace + strip surrounding quotes/punct."""
    q = q.strip().lower()
    q = re.sub(r"[\s ]+", " ", q)
    q = q.strip(" \t\n\"'!?,.")
    return q


def _embed_query_safe(q: str):
    """Return a unit-norm embedding for q, or None if the embedder is
    unavailable. We swallow load errors so web_search never fails just
    because the bge-small model isn't reachable."""
    try:
        v = _embed_texts([q])[0]
    except Exception:  # noqa: BLE001
        return None
    return v


_SEARCH_JACCARD_THRESHOLD = float(os.environ.get(
    "QWEN_SEARCH_JACCARD_THRESHOLD", "0.7"))


def _check_search_duplicate(query: str):
    """Look up the new query against recent searches by either cosine OR
    Jaccard token overlap.

    Returns the matching (prior_query, prior_summary, score) tuple if the
    new query is a near-duplicate; else None. Two-tier detection:

    1. Semantic cosine (BGE embedding, threshold 0.97). Catches "X shares
       repurchased" ↔ "shares repurchased X" — different word order, same
       meaning — that wouldn't share enough tokens for Jaccard.

    2. Lexical Jaccard (significant-token overlap, threshold 0.7). Catches
       the iter 28 search-spam pattern where the model reformulates with
       slightly different words ("<COMPANY> Q3 GAAP gross margin guidance"
       → "<COMPANY> Q3 earnings GAAP gross margin guidance plan") that
       cosine missed. The Jaccard-only branch produces the same near-dup
       refusal so the spam-throttle counter ticks correctly.

    Without (2), cosine 0.97 lets through 80%+ of the same-topic
    reformulations that fueled the high-call-count sessions in iter 28
    (e.g. 17- and 19-call sessions on multi-period metric questions).
    """
    if not _SEARCH_QUERY_MEMO:
        return None
    qnorm = _search_query_norm(query)
    qtokens = set(_tokenize_for_search(qnorm))
    if not qtokens:
        return None
    qvec = _embed_query_safe(query)
    import numpy as np
    qvec_np = (np.asarray(qvec, dtype=np.float32)
               if qvec is not None else None)
    best = None
    best_score = 0.0
    best_kind = ""
    for prior_q, prior_vec, prior_summary in _SEARCH_QUERY_MEMO:
        prior_tokens = set(_tokenize_for_search(_search_query_norm(prior_q)))
        if not (qtokens & prior_tokens):
            continue
        # Cosine path (when embedder is available).
        if qvec_np is not None:
            cos = float((qvec_np * prior_vec).sum())
            if cos > best_score:
                best_score = cos
                best = (prior_q, prior_summary)
                best_kind = "cos"
        # Jaccard path: fraction of unique significant tokens shared.
        if prior_tokens:
            jacc = len(qtokens & prior_tokens) / len(qtokens | prior_tokens)
            if jacc >= _SEARCH_JACCARD_THRESHOLD and jacc > best_score:
                best_score = jacc
                best = (prior_q, prior_summary)
                best_kind = "jacc"
    if best is None:
        return None
    threshold = (_SEARCH_DUP_THRESHOLD if best_kind == "cos"
                 else _SEARCH_JACCARD_THRESHOLD)
    if best_score >= threshold:
        return (*best, best_score)
    return None


def _record_search(query: str, summary: str) -> None:
    """Add a query + summary to the dedup memo; bounded LRU eviction."""
    qvec = _embed_query_safe(query)
    if qvec is None:
        return
    import numpy as np
    qvec_np = np.asarray(qvec, dtype=np.float32)
    _SEARCH_QUERY_MEMO.append((query, qvec_np, summary))
    while len(_SEARCH_QUERY_MEMO) > _SEARCH_QUERY_MEMO_MAX:
        _SEARCH_QUERY_MEMO.pop(0)


# Module-level state for rate-limit-aware backend selection. DDG returns
# HTTP 202 (or empty results) when it rate-limits us; instead of falling
# straight through to the multi-engine cascade EVERY time after a 202,
# we count consecutive failures and once we cross a small threshold we
# enter a 5-min cooldown where we use a different SINGLE backend (brave,
# then mojeek) before grudgingly falling back to the cascade.
_DDG_FAIL_COUNT = 0
_DDG_COOLDOWN_UNTIL = 0.0  # epoch seconds; 0 = no cooldown active
_DDG_FAIL_THRESHOLD = 2     # 2 consecutive failures triggers cooldown
_DDG_COOLDOWN_SECS = 300    # 5 minutes
# Ordered list of single-backend fallbacks tried during cooldown.
_FALLBACK_BACKENDS = ("brave", "mojeek", "yandex")


def _ddg_in_cooldown() -> bool:
    import time
    return _DDG_COOLDOWN_UNTIL > time.time()


def _record_ddg_outcome(success: bool) -> None:
    """Track consecutive DDG failures. Trip a 5-min cooldown after N in a row."""
    global _DDG_FAIL_COUNT, _DDG_COOLDOWN_UNTIL
    import time
    if success:
        _DDG_FAIL_COUNT = 0
        return
    _DDG_FAIL_COUNT += 1
    if _DDG_FAIL_COUNT >= _DDG_FAIL_THRESHOLD:
        _DDG_COOLDOWN_UNTIL = time.time() + _DDG_COOLDOWN_SECS


def web_search(query: str, max_results: int = 5,
               site: str = "", filetype: str = "", **_unused) -> str:
    """DuckDuckGo search (with brave/mojeek fallback) and semantic dedup.

    `site` and `filetype` are sugar for the standard search operators —
    `site=arxiv.org` becomes `site:arxiv.org` appended to the query,
    `filetype=pdf` becomes `filetype:pdf`. Use them instead of writing the
    operator yourself so the dedup memo treats variants as the same intent.

    Tolerates and discards unknown kwargs (`**_unused`) — the model
    sometimes confuses web_search and web_fetch and passes args meant for
    the other (e.g. `max_chars`); silently dropping them is safer than a
    TypeError that kills the calling node.
    """
    from ddgs import DDGS
    full_query = query.strip()
    if site:
        full_query += f" site:{site.strip()}"
    if filetype:
        full_query += f" filetype:{filetype.strip()}"

    # Semantic-dup check before we hit the network.
    dup = _check_search_duplicate(full_query)
    if dup is not None:
        prior_q, prior_summary, score = dup
        return (f"[near-duplicate of earlier search: {prior_q!r} "
                f"(cosine={score:.2f}). Result reused below — if you need "
                f"different angles, change the keywords or add site:/filetype:.]\n\n"
                + prior_summary)

    # Fetch 2x candidates so BGE reranking has something to choose from.
    rerank_enabled = os.environ.get("QWEN_WEB_SEARCH_RERANK", "1") == "1"
    fetch_count = max_results * 2 if rerank_enabled else max_results
    # Backend strategy:
    #   - Default tries DDG first (fast, single backend).
    #   - If DDG has been failing recently (rate-limited), use cooldown:
    #     skip DDG, try `brave` then `mojeek` (single backends), only fall
    #     back to the full cascade if those also fail.
    #   - Set QWEN_WEB_SEARCH_BACKEND=auto to bypass cooldown logic and
    #     always cascade.
    backend = os.environ.get("QWEN_WEB_SEARCH_BACKEND", "duckduckgo")

    def _try_backend(name: str) -> list[dict]:
        try:
            with DDGS() as ddgs:
                return list(ddgs.text(full_query, max_results=fetch_count,
                                       backend=name))
        except Exception:  # noqa: BLE001
            return []

    results: list[dict] = []
    backends_used: list[str] = []
    try:
        if backend and backend != "auto":
            if backend == "duckduckgo" and _ddg_in_cooldown():
                # Skip DDG during cooldown; try each fallback single backend.
                for fb in _FALLBACK_BACKENDS:
                    backends_used.append(fb)
                    results = _try_backend(fb)
                    if results:
                        break
            else:
                backends_used.append(backend)
                results = _try_backend(backend)
                # Track DDG outcome to drive cooldown.
                if backend == "duckduckgo":
                    _record_ddg_outcome(bool(results))
                # If primary failed and it was DDG, try fallbacks BEFORE the
                # full cascade — saves the 4-backend fan-out.
                if not results and backend == "duckduckgo":
                    for fb in _FALLBACK_BACKENDS:
                        backends_used.append(fb)
                        results = _try_backend(fb)
                        if results:
                            break
        if not results:
            # Last resort: full cascade.
            backends_used.append("auto-cascade")
            with DDGS() as ddgs:
                results = list(ddgs.text(full_query, max_results=fetch_count))
    except Exception as e:  # noqa: BLE001
        return f"[search error] {type(e).__name__}: {e}"
    # Domain blocklist: drop results whose host matches QWEN_WEB_SEARCH_BLOCK
    # (comma-separated). Used to prevent the model from finding leaked
    # answers in public benchmark/leaderboard pages while we're evaluating.
    # General mechanism — works for any benchmark host the user wants to
    # exclude, not specific to vals.ai. Off by default; enable per-session.
    block_hosts = {
        h.strip().lower()
        for h in os.environ.get("QWEN_WEB_SEARCH_BLOCK", "").split(",")
        if h.strip()
    }
    if block_hosts:
        from urllib.parse import urlparse
        def _blocked(r: dict) -> bool:
            url = r.get("href") or r.get("url", "")
            try:
                host = (urlparse(url).hostname or "").lower()
            except Exception:  # noqa: BLE001
                return False
            return any(host == h or host.endswith("." + h) for h in block_hosts)
        results = [r for r in results if not _blocked(r)]
    if not results:
        out = "(no results)"
        _record_search(full_query, out)
        return out

    if rerank_enabled and len(results) > max_results:
        try:
            import numpy as np
            snippets = [
                f"{r.get('title', '')}. {r.get('body') or r.get('snippet', '')}"
                for r in results
            ]
            vecs = _embed_texts([full_query] + snippets)
            qvec = vecs[0]
            scored = [
                (float((qvec * vecs[i + 1]).sum()), r)
                for i, r in enumerate(results)
            ]
            scored.sort(key=lambda x: x[0], reverse=True)
            results = [r for _s, r in scored[:max_results]]
        except Exception:  # noqa: BLE001
            # Embedder unavailable / OOM — fall back to ddgs ordering.
            results = results[:max_results]

    rendered = []
    for i, r in enumerate(results, 1):
        title = r.get("title", "")
        url = r.get("href") or r.get("url", "")
        snippet = r.get("body") or r.get("snippet", "")
        rendered.append(f"{i}. {title}\n   {url}\n   {snippet}")
    out = "\n\n".join(rendered)
    _record_search(full_query, out)
    return out


def _detect_machine_readable_format(text: str, content_type: str = "") -> str:
    """Return a short label if `text` looks like a machine-readable data format
    that's hostile for a model to read (XBRL / SGML / SOAP / RSS / sitemap /
    raw XML / etc.). Empty string if it looks like normal prose / markdown / json.

    Detection by leading sentinel only — fast O(1), no full-document parse.
    Avoids false positives on JSON (which is tractable) and prose that
    happens to mention an XML tag.
    """
    if not text:
        return ""
    head = text.lstrip()[:200]
    # Order matters: more-specific sentinels first.
    if head.startswith("<SEC-DOCUMENT>") or head.startswith("<sec-document>"):
        return "SEC submission SGML"
    if head.startswith(("<XBRL>", "<xbrl>", "<xbrl:", "<XBRL:")):
        return "XBRL"
    if head.startswith(("<rss", "<RSS")):
        return "RSS feed"
    if head.startswith(("<feed", "<atom:feed")):
        return "Atom feed"
    if head.startswith(("<urlset", "<sitemapindex")):
        return "sitemap"
    if head.startswith(("<soap:", "<SOAP:", "<env:Envelope", "<s:Envelope")):
        return "SOAP envelope"
    if head.startswith("<?xml"):
        # Generic XML — distinguish further by the first non-prologue tag.
        # Read until the first non-prologue tag to label.
        m = re.search(r"<\?xml[^>]*\?>\s*<([A-Za-z][\w:.-]*)", text[:2000])
        first = m.group(1) if m else "xml"
        return f"XML ({first})"
    # Content-type fallback for binary/encoded formats reaching this branch.
    if any(t in content_type for t in ("application/octet-stream",
                                        "application/x-binary")):
        return "binary"
    return ""


def _suggest_readable_alternative(url: str) -> str:
    """Heuristic: given a URL pointing at a machine-readable file, suggest
    the likely-corresponding human-readable URL on the same host. Returns
    empty string when no clear alternative exists.

    Rules (general, not site-specific):
      - `.txt` / `.xml` / `.xbrl` ending → swap to `.htm` then `.html`
        (the model can pick whichever the host actually serves)
      - URL ending with `/` or no extension → no alternative
    """
    try:
        from urllib.parse import urlparse, urlunparse
    except Exception:  # noqa: BLE001
        return ""
    try:
        p = urlparse(url)
    except Exception:  # noqa: BLE001
        return ""
    path = p.path
    for ext in (".txt", ".xml", ".xbrl", ".sgml", ".rss"):
        if path.lower().endswith(ext):
            for replacement in (".htm", ".html"):
                new_path = path[: -len(ext)] + replacement
                return urlunparse((p.scheme, p.netloc, new_path,
                                    p.params, p.query, p.fragment))
    return ""


def _extract_outbound_links(html: str, base_url: str, max_links: int = 25) -> list[tuple[str, str]]:
    """Return [(absolute_href, anchor_text), ...] for links in the HTML.

    Used by web_fetch to surface navigation options to the model so it can
    follow real links instead of fabricating URLs. The most common URL-
    fabrication failure mode (`sec.gov/Archives/edgar/.../some-fake-slug`)
    happens because the model sees a page mentioning related documents but
    has no way to navigate there without a new web_search; surfacing the
    actual hrefs cuts that whole class of error.

    Filtering:
      - External http(s) only (skip mailto:, tel:, javascript:)
      - Non-empty anchor text (skip image-only links)
      - Deduplicated by normalized URL
      - Anchor text trimmed to 100 chars
      - Up to max_links (default 25) — the most-linked-to URLs typically
        appear earlier in the DOM (nav + main content)
    """
    if not html:
        return []
    try:
        from bs4 import BeautifulSoup
        from urllib.parse import urljoin
    except ImportError:
        return []
    try:
        soup = BeautifulSoup(html, "html.parser")
    except Exception:  # noqa: BLE001
        return []
    seen: set[str] = set()
    out: list[tuple[str, str]] = []
    for a in soup.find_all("a", href=True):
        try:
            href = urljoin(base_url, a["href"])
        except Exception:  # noqa: BLE001
            continue
        if not href.startswith(("http://", "https://")):
            continue
        # Skip in-page anchors (same URL with fragment-only difference).
        if "#" in href:
            href = href.split("#", 1)[0]
        text = (a.get_text(strip=True) or "")
        if not text:
            continue
        if len(text) > 100:
            text = text[:100] + "…"
        norm = _normalize_url(href)
        if not norm or norm in seen:
            continue
        seen.add(norm)
        out.append((href, text))
        if len(out) >= max_links:
            break
    return out


def _format_links_section(links: list[tuple[str, str]]) -> str:
    """Render a links section appended to web_fetch output. Markdown-format
    so it's compact and parseable by the model."""
    if not links:
        return ""
    lines = ["", "## Links on this page",
             "(navigate via these without burning a web_search call)"]
    for href, text in links:
        lines.append(f"- [{text}]({href})")
    return "\n".join(lines)


def _smart_truncate(text: str, max_chars: int) -> str:
    """Keep both ends of a long body so abstract+conclusion both survive.

    Standard slice-from-start truncation is brutal on academic papers and
    long blog posts where the most useful content is at the start (title +
    abstract) AND the end (conclusion / refs). Here we keep ~55% from the
    head and ~45% from the tail, separated by a clear marker.
    """
    if len(text) <= max_chars:
        return text
    head_chars = int(max_chars * WEB_FETCH_HEAD_FRACTION)
    tail_chars = max_chars - head_chars - 64  # 64 chars for the marker
    if tail_chars < 1024:
        # Pathologically small budget — degrade to single-end slice.
        return text[:max_chars] + f"\n\n…[truncated {len(text) - max_chars} chars]"
    head = text[:head_chars]
    tail = text[-tail_chars:]
    skipped = len(text) - head_chars - tail_chars
    return (f"{head}\n\n"
            f"…[middle {skipped} chars omitted; head/tail kept]…\n\n"
            f"{tail}")


def web_fetch(url: str, max_chars: int = WEB_FETCH_DEFAULT_MAX,
              force_browser: bool = False, head_only: bool = False,
              mode: str = "semantic") -> str:
    """Fetch a URL and return its readable text.

    Pipeline:
      1. Static httpx GET. If the page is HTML, extract <article>/<main>
         body via BeautifulSoup. ~1 s for typical pages.
      2. Fall back to a JS-rendering Chromium browser via crawl4ai for
         pages where the static path can't recover content (≈ SPAs).

    `mode` selects the retrieval profile that `condense_tool_result` will
    apply downstream:
      - "numerical": answer is stated verbatim (specific value, count,
        name, date, or list item). Lexical + structural ranking only —
        cross-encoder is bypassed because it demotes numeric/tabular
        chunks in favor of prose that merely describes them.
      - "semantic" (default): answer is a synthesis/judgment built from
        prose. Engages the cross-encoder + sentence-extract pipeline so
        the kept chunks are the ones that actually answer the question,
        not just mention the topic.

    Truncation keeps both ends of the body so abstract + conclusion both
    survive on long pages. Set `head_only=True` to skip the tail and only
    keep the start — useful when you only want a heading-and-abstract.
    """
    _set_fetch_mode(mode)
    if not force_browser:
        text = _static_fetch(url, max_chars, head_only=head_only)
        if text is not None:
            return text
    return _browser_fetch(url, max_chars, head_only=head_only)


def _static_fetch(url: str, max_chars: int, head_only: bool = False) -> str | None:
    """Fast HTTP fetch + content extraction. Returns None to fall back to
    the JS-rendering browser path.

    Extraction order:
      1. trafilatura — paragraph-level scoring, designed for content
         extraction across diverse page layouts. Strips boilerplate
         (nav, ads, footer, "related articles") far better than BS4.
      2. BS4 fallback — strips obvious nav/footer/aside, prefers
         <article>/<main>. Used when trafilatura returns nothing on
         pages it doesn't understand.
    """
    try:
        import httpx
    except ImportError:
        return None
    try:
        r = httpx.get(
            url,
            timeout=10,
            follow_redirects=True,
            headers={"User-Agent": "Mozilla/5.0 (qwen-agent)"},
        )
        r.raise_for_status()
    except Exception:  # noqa: BLE001
        return None

    # Block-page redirect detection. If we followed redirects and the FINAL
    # path matches a known bot-block / captcha pattern, return a short
    # `[blocked: ...]` notice instead of the chrome HTML. Walmart, Amazon,
    # and Cloudflare-protected sites self-identify this way — we'd otherwise
    # feed the model 5 KB of "Robot detected, please verify" noise that
    # poisons context and triggers fabricated retry URLs.
    try:
        from urllib.parse import urlparse
        final_url = str(r.url)
        final_parsed = urlparse(final_url)
        final_path = (final_parsed.path or "").lower()
        original_host = (urlparse(url).hostname or "").lower()
        final_host = (final_parsed.hostname or "").lower()
        if any(p in final_path for p in _BLOCK_REDIRECT_PATH_PATTERNS):
            cross_host = (final_host and original_host and
                          final_host != original_host)
            return (
                f"[blocked: {original_host or url} "
                f"{'redirected cross-host to ' + final_host + ' and ' if cross_host else ''}"
                f"served a bot-block / captcha page at {final_url}. "
                f"This site blocks automated access — try a different source "
                f"(e.g. the manufacturer's official site, USDA FoodData Central "
                f"for nutrition data, or a public API).]"
            )
    except Exception:  # noqa: BLE001
        pass

    ct = r.headers.get("content-type", "").lower()
    # PDF handling: detect by content-type OR URL suffix (some servers
    # send `application/octet-stream` for PDFs, or no content-type at
    # all). pypdf extracts page text; we concatenate with form-feed
    # separators so chunkers can detect page boundaries downstream.
    # Without this, web_fetch on a PDF would return raw binary as
    # `r.text` — useless to the model and noise in context. Common
    # case in finance: SEC press releases, IR site decks, FOMC
    # statements, BIS / IMF / OECD reports — all PDFs that the model
    # otherwise can't read.
    is_pdf = ("application/pdf" in ct
              or url.lower().split("?")[0].split("#")[0].endswith(".pdf"))
    if is_pdf:
        try:
            import pypdf  # type: ignore
            import io as _io
            reader = pypdf.PdfReader(_io.BytesIO(r.content))
            page_texts = []
            for p in reader.pages:
                try:
                    t = p.extract_text() or ""
                except Exception:  # noqa: BLE001
                    t = ""
                if t.strip():
                    page_texts.append(t)
            if not page_texts:
                return (f"[empty: PDF at {url} extracted to 0 chars of "
                        f"text — likely scanned/image-only or encrypted. "
                        f"Try a text-format alternative (HTML/HTM, .htm "
                        f"version on the same site).]")
            full = "\n\n\f\n\n".join(page_texts)  # \f = page break
            n_pages = len(reader.pages)
            n_extracted = len(page_texts)
            header = (f"[pdf: {n_extracted}/{n_pages} pages extracted, "
                      f"{len(full)} chars; \\f marks page breaks]\n\n")
            body = head_only and full[:max_chars] or _smart_truncate(full, max_chars)
            tail = ""
            if head_only and len(full) > max_chars:
                tail = f"\n\n…[truncated {len(full) - max_chars} chars]"
            return header + body + tail
        except ImportError:
            return (f"[error: pypdf not installed in venv — can't extract "
                    f"text from PDF at {url}. Try the HTML version on the "
                    f"same site.]")
        except Exception as e:  # noqa: BLE001
            return (f"[error reading PDF at {url}: {type(e).__name__}: {e}. "
                    f"The file may be scanned/encrypted; try the HTML "
                    f"version or a text-format alternative.]")
    if "text/html" not in ct:
        text = r.text
        # Machine-readable-format detection. Pages that wrap data in XML /
        # SGML / XBRL / SOAP / RSS / Atom / sitemap formats are unreadable
        # to a model: they're ~30% angle-brackets, namespace declarations,
        # and encoded payloads. Returning the raw bytes wastes a turn and
        # poisons context with markup. Detect by leading sentinel and
        # suggest a likely-readable alternative URL when one is derivable.
        fmt_label = _detect_machine_readable_format(text, ct)
        if fmt_label:
            suggested = _suggest_readable_alternative(str(r.url))
            head = text[:600].replace("\n", " ")
            tail = (f" Try: web_fetch({suggested!r})" if suggested else
                    " Look for an .htm/.html link to the same document on "
                    "the parent index page, or run web_search for the "
                    "human-readable version.")
            return (
                f"[machine-readable format detected ({fmt_label}) — this is "
                f"raw data/metadata, not a readable page. Refetch the "
                f"human-readable version instead.{tail} "
                f"First 600 chars for confirmation: {head!r}]"
            )
        # plain text / json / markdown — return raw, truncated
        if head_only:
            return text[:max_chars] + (f"\n\n…[truncated {len(text) - max_chars} chars]"
                                       if len(text) > max_chars else "")
        return _smart_truncate(text, max_chars)

    extracted: str | None = None
    page_title = ""

    # Pre-strip XBRL inline tags (iXBRL — `<ix:nonFraction>`, `<ix:nonNumeric>`,
    # `<ix:continuation>`, etc.). These wrap numeric values in SEC filings;
    # downstream parsers (trafilatura, BS4) treat them as foreign-namespace
    # elements and drop their text content along with the tags, which silently
    # eats the values in financial-statement tables (Snap 10-K convertible-
    # notes-if-converted, dilutive share counts, revenue line items, etc.).
    # Strip just the opening/closing tags, preserve the inner text — this is
    # the same content that human readers see in the document. Regex is bounded
    # to the `ix:` namespace so non-XBRL namespaces (svg:, math:) are unaffected.
    raw_html = r.text
    if "<ix:" in raw_html:
        raw_html = re.sub(r"</?ix:[a-zA-Z0-9_]+(?:\s[^>]*)?>", "", raw_html)

    # Try trafilatura first — much better at separating content from chrome.
    try:
        import trafilatura
        # NOTE: dropped favor_recall=True after diagnosing the Snap 10-K
        # table-cell-value loss bug. With favor_recall=True, trafilatura
        # was aggressively merging/deduping short numeric paragraphs and
        # losing the actual numbers in financial-statement tables — the
        # opposite of what the flag's name suggests. Default (precision-
        # oriented) extraction keeps the cell values intact.
        # include_comments=False since we're never grading comment threads.
        tx = trafilatura.extract(
            raw_html,
            output_format="markdown",
            include_comments=False,
            include_tables=True,
            with_metadata=False,
        )
        if tx and len(tx) > 300:
            extracted = tx
        # Fish out the title separately — trafilatura.extract drops it
        # in markdown mode; metadata gives us a clean version.
        try:
            md = trafilatura.extract_metadata(raw_html)
            if md and getattr(md, "title", None):
                page_title = md.title.strip()
        except Exception:  # noqa: BLE001
            pass
    except ImportError:
        pass
    except Exception:  # noqa: BLE001
        # Trafilatura can choke on malformed HTML — fall through to BS4.
        pass

    # BS4 fallback for pages trafilatura doesn't understand.
    if extracted is None:
        try:
            from bs4 import BeautifulSoup
            # Use the ix-stripped html so XBRL inline values survive here too.
            soup = BeautifulSoup(raw_html, "html.parser")
        except Exception:  # noqa: BLE001
            return None

        for tag in soup(["script", "style", "nav", "header", "footer", "aside", "noscript"]):
            tag.decompose()

        if not page_title and soup.title and soup.title.string:
            page_title = soup.title.string.strip()

        for tag_name in ("article", "main"):
            el = soup.find(tag_name)
            if el:
                text = el.get_text(separator="\n", strip=True)
                if len(text) > 300:
                    extracted = text
                    break
        if extracted is None and soup.body:
            text = soup.body.get_text(separator="\n", strip=True)
            if len(text) > 300:
                extracted = text

    if extracted is None:
        # Probably JS-heavy — let the browser path handle it.
        return None

    # Body-keyword block detection. Some sites return 200 OK with a body
    # that's actually a "verify you're human" / "robot detected" page —
    # no redirect to detect. Catch them by scanning the extracted text
    # for hallmark phrases AND a low signal-to-noise ratio (short, mostly
    # boilerplate). Returning a `[blocked: ...]` notice keeps the captcha
    # text out of context AND lets the dispatcher mark this URL dead.
    if _looks_like_block_body(extracted):
        return (
            f"[blocked: {urlparse(url).hostname} returned a bot-detection / "
            f"captcha page (HTTP 200 with verify-human content). Try a "
            f"different source.]"
        )

    extracted = re.sub(r"\n{3,}", "\n\n", extracted)
    # Empty-content guard (matches the one in _browser_fetch). Pre-fix,
    # a page that extracted to <200 chars of body would still return that
    # tiny string, looking indistinguishable from a real but short page.
    # Force a clear `[empty: ...]` so the model pivots rather than
    # retrying same-host URLs.
    if len(extracted.strip()) < 200:
        try:
            from urllib.parse import urlparse
            host = (urlparse(url).hostname or url).lower()
        except Exception:  # noqa: BLE001
            host = url
        return (
            f"[empty: {host} returned only {len(extracted.strip())} chars of "
            f"readable content — likely paywalled, login-walled, or behind "
            f"a 'subscribe to continue' wall. Stop retrying this host; try a "
            f"different source (the SEC filing if numeric, the official "
            f"investor-relations site, or a different aggregator).]"
        )
    if page_title and page_title not in extracted[:200]:
        extracted = f"# {page_title}\n\n{extracted}"
    # Link extraction — append outbound links so the model can navigate
    # without burning a web_search call. Reserve a small char budget
    # (`_LINKS_CHAR_BUDGET`) so the truncated body still fits in max_chars.
    links_section = _format_links_section(
        _extract_outbound_links(r.text, str(r.url))
    )
    body_budget = max(max_chars - len(links_section), max_chars // 2)
    if head_only:
        if len(extracted) > body_budget:
            extracted = extracted[:body_budget] + f"\n\n…[truncated {len(extracted) - body_budget} chars]"
        return extracted + links_section
    return _smart_truncate(extracted, body_budget) + links_section


# Phrases that overwhelmingly indicate a captcha / bot-block landing page.
# Match case-insensitively — these have very low collision with real
# article content. Pair with a length check (bot pages are <2k chars
# of meaningful content) to avoid false positives on articles that
# legitimately discuss CAPTCHA.
_BLOCK_BODY_PHRASES = (
    "verify you are human",
    "verify you're human",
    "verify that you are human",
    "are you a robot",
    "robot check",
    "bot detection",
    "checking your browser",
    "please complete the captcha",
    "complete the security check",
    "ddos protection by cloudflare",
    "access denied",
    "you have been blocked",
    "request unsuccessful",
    "press and hold the button",
    "additional security check is required",
    "as part of our zero-shot fraud detection",
)


def _looks_like_block_body(text: str) -> bool:
    """True if the extracted page body looks like a bot-block landing page.
    Length ceiling avoids matching real articles that mention captchas in
    their content."""
    if not text:
        return False
    if len(text) > 4000:
        # Real content + a stray "verify" mention is unlikely to be a block.
        return False
    low = text.lower()
    return any(p in low for p in _BLOCK_BODY_PHRASES)


def _browser_fetch(url: str, max_chars: int, head_only: bool = False) -> str:
    """Heavy fetch via crawl4ai (real Chromium, full JS)."""
    from crawl4ai import AsyncWebCrawler, BrowserConfig, CrawlerRunConfig

    async def _run() -> tuple[str, str]:
        """Returns (extracted_body_text, raw_html). raw_html may be empty."""
        browser_cfg = BrowserConfig(headless=True, verbose=False)
        run_cfg = CrawlerRunConfig(verbose=False, page_timeout=30000)
        async with AsyncWebCrawler(config=browser_cfg) as crawler:
            result = await crawler.arun(url=url, config=run_cfg)
            raw_html = getattr(result, "html", "") or ""
            if not isinstance(raw_html, str):
                raw_html = ""
            # Prefer trafilatura over crawl4ai's markdown when raw HTML is
            # available — gives content-extraction parity with _static_fetch
            # and tends to strip more chrome (newsletter banners, related-
            # articles boxes) than crawl4ai's default markdown converter.
            if raw_html and len(raw_html) > 500:
                try:
                    import trafilatura
                    tx = trafilatura.extract(
                        raw_html, output_format="markdown",
                        favor_recall=True, include_comments=False,
                        include_tables=True, with_metadata=False,
                    )
                    if tx and len(tx) > 300:
                        return tx, raw_html
                except Exception:  # noqa: BLE001
                    pass
            md = (
                getattr(result, "markdown", None)
                or getattr(result, "cleaned_html", None)
                or raw_html
                or ""
            )
            if hasattr(md, "raw_markdown"):
                md = md.raw_markdown
            return str(md), raw_html

    text, raw_html = asyncio.run(_run())
    # Empty-content guard: if the page rendered to ~nothing (paywall,
    # auth-wall, JS gated), surface an explicit `[empty: ...]` notice so
    # the model knows to PIVOT rather than fetch a sibling URL on the
    # same dead host. Pre-fix: in iter 23 a session fetched seekingalpha
    # and streetinsider, got "" and "\n", didn't realize the pages were dead
    # → re-tried sibling URLs and burned tool budget. The threshold (200
    # chars of body, ignoring whitespace and the links section) is well
    # below any real article and high above empty/whitespace responses.
    body_signal = (text or "").strip()
    if len(body_signal) < 200:
        try:
            from urllib.parse import urlparse
            host = (urlparse(url).hostname or url).lower()
        except Exception:  # noqa: BLE001
            host = url
        return (
            f"[empty: {host} returned only {len(body_signal)} chars of "
            f"readable content — likely paywalled, login-walled, or "
            f"JS-rendered without a usable shell. Stop retrying this host; "
            f"try a different source (the SEC 10-K/10-Q filing if numeric, "
            f"or another aggregator).]"
        )
    # Same link-extraction logic as _static_fetch — surface outbound links so
    # the model can navigate without burning a web_search call.
    links_section = _format_links_section(_extract_outbound_links(raw_html, url))
    body_budget = max(max_chars - len(links_section), max_chars // 2)
    if head_only:
        if len(text) > body_budget:
            text = text[:body_budget] + f"\n\n…[truncated {len(text) - body_budget} chars]"
        return text + links_section
    return _smart_truncate(text, body_budget) + links_section


def find_in_url(url: str, pattern: str,
                context_lines: int = 3, max_matches: int = 20,
                ignore_case: bool = True) -> str:
    """Fetch a URL and return only the lines matching `pattern`, each with
    a few surrounding context lines.

    Targeted retrieval for "needle in haystack" extraction: when
    `web_fetch` returns a condensed result that dropped the specific
    table cell or footnote you need, use this to grep the raw page for
    the exact phrase. Works on the FULL page contents, not the
    condensed view, so you can locate values regardless of where they
    sit in the document.

    Args:
      url: page URL (must be a previously-seen URL or on a preapproved
           host — same URL guard as web_fetch).
      pattern: regex (Python `re` syntax) to match. For literal text,
               escape with `\\` or just pass plain words.
      context_lines: lines of context before/after each match.
      max_matches: cap on returned matches.
      ignore_case: case-insensitive matching (default True).

    Examples:
      find_in_url(<10-K url>, r"Long-term debt")
      find_in_url(<10-K url>, r"\\$\\d+,\\d{3}")  # find dollar amounts
      find_in_url(<10-K url>, r"convertible notes?", context_lines=8)
    """
    if not url:
        return "[error] find_in_url: url is required"
    if not pattern:
        return "[error] find_in_url: pattern is required"
    text = _static_fetch(url, max_chars=2_000_000, head_only=False)
    if text is None:
        # Fall back to browser path for JS-heavy pages.
        text = _browser_fetch(url, max_chars=2_000_000, head_only=False)
    if not text:
        return f"[find_in_url] empty fetch from {url!r}"
    if isinstance(text, str) and text.lstrip().startswith("[blocked"):
        return text
    try:
        flags = re.IGNORECASE if ignore_case else 0
        rx = re.compile(pattern, flags)
    except re.error as e:
        return f"[find_in_url] bad regex {pattern!r}: {e}"
    lines = text.splitlines()
    out: list[str] = []
    n_matches = 0
    last_emitted_line = -1
    for i, line in enumerate(lines):
        if rx.search(line):
            n_matches += 1
            if n_matches > max_matches:
                out.append(f"…[truncated; more than {max_matches} matches; "
                           f"narrow the pattern]")
                break
            start = max(0, i - context_lines)
            end = min(len(lines), i + context_lines + 1)
            # Avoid re-emitting context that overlaps the previous block.
            if start <= last_emitted_line:
                start = last_emitted_line + 1
                if start >= end:
                    continue
            else:
                if out:
                    out.append("---")
            for j in range(start, end):
                marker = ">" if j == i else " "
                out.append(f"{marker} {lines[j]}")
            last_emitted_line = end - 1
    if n_matches == 0:
        return (f"[find_in_url] 0 matches for {pattern!r} in {url!r} "
                f"({len(lines)} lines, {len(text)} chars). Try a "
                f"different pattern or check the URL is correct.")
    header = (f"[find_in_url: {n_matches} match(es) for {pattern!r} in "
              f"{url!r} — {len(lines)} total lines]")
    return header + "\n\n" + "\n".join(out)


def web_outline(url: str, max_headings: int = 80) -> str:
    """Return only the heading hierarchy of a page — a quick "what's here?".

    Use this BEFORE web_fetch when you're not sure if the page is worth a
    full read. Output: indented list of h1/h2/h3 headings (and link anchors
    for major sections), each with a few words of following body so the
    skim is meaningful. ~30× cheaper than web_fetch on a typical paper.
    """
    try:
        import httpx
        from bs4 import BeautifulSoup
    except ImportError:
        return "[error] httpx/bs4 not installed"
    try:
        r = httpx.get(url, timeout=10, follow_redirects=True,
                      headers={"User-Agent": "Mozilla/5.0 (qwen-agent)"})
        r.raise_for_status()
    except Exception as e:  # noqa: BLE001
        return f"[fetch failed] {type(e).__name__}: {e}"
    if "html" not in (r.headers.get("content-type") or "").lower():
        return f"[not html: {r.headers.get('content-type', '?')}]"
    try:
        soup = BeautifulSoup(r.text, "html.parser")
    except Exception as e:  # noqa: BLE001
        return f"[parse failed] {e}"
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    title = soup.title.string.strip() if soup.title and soup.title.string else ""
    out: list[str] = []
    if title:
        out.append(f"# {title}")
    seen = 0
    for h in soup.find_all(["h1", "h2", "h3", "h4"]):
        if seen >= max_headings:
            out.append(f"…[+{len(soup.find_all(['h1','h2','h3','h4'])) - seen} more headings]")
            break
        level = int(h.name[1])
        indent = "  " * (level - 1)
        text = " ".join(h.get_text(strip=True).split())[:120]
        if not text:
            continue
        # Take the next non-empty sibling text as a snippet.
        snippet = ""
        sib = h.find_next_sibling()
        for _ in range(3):
            if sib is None:
                break
            t = " ".join(sib.get_text(strip=True).split())
            if t:
                snippet = t[:120]
                break
            sib = sib.find_next_sibling()
        line = f"{indent}- {text}"
        if snippet:
            line += f"\n{indent}  → {snippet}"
        out.append(line)
        seen += 1
    if seen == 0:
        return "(no headings — try web_fetch for the full body)"
    return "\n".join(out)


# ---------- files ----------------------------------------------------------

def _too_broad_scope_refusal(resolved: Path, tool_name: str) -> str | None:
    """Refuse recursive walks that would scan the home dir, the filesystem
    root, or system dirs holding credentials / app state. The list is
    intentionally tight — `$HOME/Documents`, `$HOME/Desktop`, `~/code`,
    and other legitimate working folders are NOT refused. We only block:

      - the actual root / parent-of-users dirs (`/`, `/Users`, `/home`)
      - the home dir itself (greppable subdirs are fine, the root is not)
      - macOS / Unix system roots (`/System`, `/Library`, `/etc`, `/var`,
        `/usr`, `/opt`, `/Applications`, `/Volumes`, `/private`, `/bin`,
        `/sbin`)
      - credential / app-state dirs inside `$HOME` (`.ssh`, `.gnupg`,
        `.aws`, `.kube`, `.config`, `Library`)

    Single-file targets are always allowed — only directory walks fire."""
    try:
        if not resolved.is_dir():
            return None
    except OSError:
        return None
    p = str(resolved).rstrip("/") or "/"
    home = str(Path.home()).rstrip("/")
    # macOS resolves /etc, /var, /tmp through /private. The resolved forms
    # are included explicitly so the exact-match check still fires; /tmp
    # itself (resolves to /private/tmp) is deliberately NOT blocked
    # because it's the scratch working dir.
    blocked = {
        "/", home, "/Users", "/home",
        "/Library", "/System", "/private",
        "/etc", "/var", "/usr", "/opt",
        "/Applications", "/Volumes", "/bin", "/sbin",
        "/private/etc", "/private/var",  # macOS-resolved siblings
        f"{home}/Library", f"{home}/.ssh", f"{home}/.gnupg",
        f"{home}/.aws", f"{home}/.kube", f"{home}/.config",
    }
    if p in blocked:
        return (
            f"[refused] {tool_name} scope '{resolved}' is too broad — that "
            f"path holds ssh keys, browser data, app state, or personal "
            f"documents. Pass a specific subdirectory (e.g. the active "
            f"project root) or a single file path instead."
        )
    return None


def read_file(path: str, offset: int = 0, limit: int = READ_DEFAULT_LINES) -> str:
    """Read a slice of a file. offset is 0-based; limit caps lines returned."""
    p = Path(path).expanduser().resolve()
    if not p.is_file():
        return f"[error] no such file: {p}"
    if p.stat().st_size > READ_MAX_BYTES and offset == 0 and limit >= READ_DEFAULT_LINES:
        return (
            f"[error] {p} is {p.stat().st_size} bytes (>256KB). Read a slice with "
            f"offset/limit, or grep first to find the relevant section."
        )
    lines = p.read_text(encoding="utf-8", errors="replace").splitlines()
    total = len(lines)
    end = min(offset + limit, total)
    chunk = lines[offset:end]
    body = "\n".join(f"{i + 1:>6}\t{l}" for i, l in enumerate(chunk, start=offset))
    suffix = ""
    if end < total:
        suffix = f"\n\n…[showing lines {offset + 1}-{end} of {total}; use offset/limit for more]"
    elif offset > 0:
        suffix = f"\n\n…[showing lines {offset + 1}-{end} of {total}]"
    return body + suffix


def list_files(path: str = ".", pattern: str = "**/*", max_results: int = 100) -> str:
    """List files under a path, optionally filtered by a glob pattern."""
    base = Path(path).expanduser().resolve()
    if not base.exists():
        return f"[error] no such path: {base}"
    refusal = _too_broad_scope_refusal(base, "list_files")
    if refusal:
        return refusal
    matches = []
    for m in base.glob(pattern):
        if m.is_file():
            try:
                rel = m.relative_to(base)
            except ValueError:
                rel = m
            matches.append(str(rel))
        if len(matches) >= max_results:
            break
    if not matches:
        return "(no matches)"
    return "\n".join(matches)


def grep(
    pattern: str,
    path: str = ".",
    glob: str | None = None,
    output_mode: str = "content",
    head_limit: int = GREP_DEFAULT_HEAD,
) -> str:
    """Regex-search file contents under a path.

    output_mode:
      - "files_with_matches": just file paths (cheapest, use first)
      - "content": matching lines with file:line: prefix (default)
      - "count": match counts per file
    """
    base = Path(path).expanduser().resolve()
    if not base.exists():
        return f"[error] no such path: {base}"
    refusal = _too_broad_scope_refusal(base, "grep")
    if refusal:
        return refusal
    use_rg = shutil.which("rg") is not None

    if use_rg:
        cmd = ["rg", "--color=never"]
        if output_mode == "files_with_matches":
            cmd += ["-l"]
        elif output_mode == "count":
            cmd += ["-c"]
        else:  # content
            cmd += ["--line-number", "--no-heading"]
        if glob:
            cmd += ["-g", glob]
        for ex in VCS_EXCLUDES:
            cmd += ["-g", f"!{ex}/"]
        cmd += ["-e", pattern, str(base)]
    else:
        cmd = ["grep", "-r", "-E", "--color=never"]
        if output_mode == "files_with_matches":
            cmd += ["-l"]
        elif output_mode == "count":
            cmd += ["-c"]
        else:
            cmd += ["-n"]
        for ex in VCS_EXCLUDES:
            cmd += ["--exclude-dir", ex]
        if glob:
            cmd += ["--include", glob]
        cmd += ["-e", pattern, str(base)]

    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
    except subprocess.TimeoutExpired:
        return "[timeout after 30s]"

    out = result.stdout.strip()
    if not out:
        return "(no matches)"

    lines = out.splitlines()
    if output_mode == "content":
        # Cap each line at GREP_LINE_MAX_CHARS to dodge minified-JS / base64 noise.
        lines = [
            (l if len(l) <= GREP_LINE_MAX_CHARS else l[:GREP_LINE_MAX_CHARS] + "…[line truncated]")
            for l in lines
        ]

    if head_limit > 0 and len(lines) > head_limit:
        extra = len(lines) - head_limit
        lines = lines[:head_limit] + [f"…[{extra} more results — narrow your search or raise head_limit]"]
    return "\n".join(lines)


def write_file(path: str, content: str) -> str:
    """Create or overwrite a file with the given content."""
    p = Path(path).expanduser().resolve()
    p.parent.mkdir(parents=True, exist_ok=True)
    existed = p.exists()
    p.write_text(content, encoding="utf-8")
    _track_write(str(p))
    verb = "overwrote" if existed else "created"
    return f"{verb} {p} ({len(content)} chars)"


def edit_file(path: str, old_string: str, new_string: str) -> str:
    """Replace one occurrence of old_string with new_string. old_string must appear exactly once."""
    p = Path(path).expanduser().resolve()
    if not p.is_file():
        return f"[error] no such file: {p}"
    text = p.read_text(encoding="utf-8")
    count = text.count(old_string)
    if count == 0:
        return f"[error] old_string not found in {p}"
    if count > 1:
        return f"[error] old_string appears {count} times in {p} — make it more specific (include surrounding context)"
    p.write_text(text.replace(old_string, new_string, 1), encoding="utf-8")
    _track_write(str(p))
    return f"edited {p}"


# ---------- memory (persistent across sessions, vector-search) -------------
# Cross-session knowledge store. Stores text + a vector embedding per entry.
# Search is cosine-similarity over the embedding space (semantic, not just
# keyword). Backed by SQLite (atomic writes, file-based, crash-safe). Linear
# scan is fast enough for personal-scale (<50k entries); the search function
# is structured so swapping in an ANN index later is a one-function change.

_MEMORY_DB_ENV = "QWEN_MEMORY_DB"
_MEMORY_DEFAULT_NAME = ".qwen_memory.db"
_MEMORY_EMBED_MODEL_ENV = "QWEN_EMBED_MODEL"
_MEMORY_EMBED_MODEL_DEFAULT = "mlx-community/bge-small-en-v1.5-bf16"


def _memory_db_path() -> str:
    """Default ~/.qwen_memory.db (user-global so memories survive cd-ing
    between projects). Override with $QWEN_MEMORY_DB."""
    p = os.environ.get(_MEMORY_DB_ENV)
    if p:
        return os.path.expanduser(p)
    return os.path.expanduser(f"~/{_MEMORY_DEFAULT_NAME}")


def _memory_embed_model_name() -> str:
    return os.environ.get(_MEMORY_EMBED_MODEL_ENV, _MEMORY_EMBED_MODEL_DEFAULT)


# Lazy-loaded embedding model + tokenizer. Loaded once, cached for the
# process lifetime. Loading is ~0.1-2s for bge-small; first call to a memory
# tool pays this cost, subsequent calls reuse the loaded model.
_embed_state: dict[str, Any] = {"model": None, "tokenizer": None, "name": None, "dim": None}


def _embed_load() -> tuple[Any, Any, str, int]:
    """Lazy-load the embedding model. Returns (model, tokenizer, name, dim)."""
    name = _memory_embed_model_name()
    if _embed_state["model"] is not None and _embed_state["name"] == name:
        return _embed_state["model"], _embed_state["tokenizer"], name, _embed_state["dim"]
    # Heavy import is intentionally lazy — don't load mlx_embeddings on
    # agent startup, only when the model actually calls a memory tool.
    from mlx_embeddings.utils import load as _load_embed
    import mlx.core as mx_local
    # Cap MLX's Metal allocator BEFORE the model loads. Default cache pool
    # on Apple Silicon is multi-GB (sized for huge models like dflash's 35B);
    # bge-small only needs ~50 MB of weights and ≤512-token forward passes,
    # so 256 MB is plenty. Without this cap the qwen_ui process shows ~3.7 GB
    # in Activity Monitor (MLX claims buffer pool space it never uses).
    # Measured perf cost: <1 ms per embed call — within run-to-run jitter.
    # Disable with QWEN_EMBED_CACHE_MB=0; tune up via the same env var.
    try:
        cap_mb = int(os.environ.get("QWEN_EMBED_CACHE_MB", "256"))
        if cap_mb > 0:
            mx_local.set_cache_limit(cap_mb * 1024 * 1024)
    except Exception:  # noqa: BLE001
        pass
    model, tok = _load_embed(name)
    # Probe dim with a single tiny forward pass.
    enc = tok._tokenizer(["probe"], padding=True, truncation=True,
                         max_length=8, return_tensors="np")
    out = model(input_ids=mx_local.array(enc["input_ids"]),
                attention_mask=mx_local.array(enc["attention_mask"]))
    last_hidden = out.last_hidden_state if hasattr(out, "last_hidden_state") else out["last_hidden_state"]
    dim = int(last_hidden.shape[-1])
    _embed_state.update(model=model, tokenizer=tok, name=name, dim=dim)
    # Drop probe-pass transient buffers so the post-load resident set is
    # close to the actual weight footprint (~50 MB for bge-small).
    try:
        mx_local.clear_cache()
    except Exception:  # noqa: BLE001
        pass
    return model, tok, name, dim


def _embed_texts(texts: list[str]) -> "np.ndarray":  # type: ignore[name-defined]
    """Embed a list of strings → unit-norm L2 vectors. Uses CLS pooling
    (BGE-style); for mean-pooling models you'd want a different reduction
    but BGE/most retrieval models use CLS."""
    import numpy as np
    import mlx.core as mx_local
    model, tok, _name, _dim = _embed_load()
    enc = tok._tokenizer(
        texts, padding=True, truncation=True, max_length=512, return_tensors="np"
    )
    input_ids = mx_local.array(enc["input_ids"])
    attention_mask = mx_local.array(enc["attention_mask"])
    out = model(input_ids=input_ids, attention_mask=attention_mask)
    last_hidden = out.last_hidden_state if hasattr(out, "last_hidden_state") else out["last_hidden_state"]
    # CLS-token pool then L2-normalize so cosine similarity = dot product.
    cls = last_hidden[:, 0, :]
    norms = mx_local.sqrt((cls * cls).sum(axis=1, keepdims=True))
    cls = cls / norms
    mx_local.eval(cls)
    return np.asarray(cls.tolist(), dtype=np.float32)


# ---------- tool-result triage --------------------------------------------
# Cheap classification of "is this tool result useful for the current task?"
# using the SAME bge-small embedder already loaded for memory search. No new
# weights, no new RAM. Bi-encoder relevance is less accurate than a cross-
# encoder reranker, but free in our setup and good enough for the headline
# case: detecting that a long web/file result is mostly off-topic chrome.
#
# Result condensing is enabled by default for very large web/paper results.
# It is deliberately chunked and conservative: keep the lead, keep top
# relevance-scored chunks, preserve headings/source-ish lines, and never touch
# exec/edit/error outputs. Disable with QWEN_RESULT_CONDENSE=0.
#
# The older embedding triage remains available behind QWEN_TRIAGE_ENABLE=1,
# but it now runs only after condensing has had the first chance to reduce
# obviously large research payloads.

_TRIAGE_NEVER_PRUNE_TOOLS = frozenset({
    # These return short, structured data — never noisy enough to triage.
    "now", "make_table", "memory_save", "memory_get", "memory_list",
    "memory_delete", "todo_write", "done", "mcp_list",
    "agent_graph_list", "enter_worktree", "exit_worktree",
    # Exec/edit tools — keep verbatim so the model can see what actually
    # happened on disk.
    "write_file", "edit_file", "apply_patch", "write_file_verified",
    "python_run", "python_reset", "test_run", "notebook_edit", "notebook_run",
})


def triage_tool_result(task: str, tool_name: str, result: str) -> tuple[str, dict]:
    """Score relevance of a tool result against the current task and either
    keep / shorten / stub it.

    Returns (kept_or_replaced_text, info) where info has fields:
        score:    cosine sim 0..1, or None if no scoring happened
        verdict:  "kept" | "low_relevance" | "skipped"
        chars_in: original result length
        chars_out: returned-text length

    Pure no-op when QWEN_TRIAGE_ENABLE != "1" — emits the result unchanged
    with verdict="skipped". Safe to call on every tool result; cheap when
    the result is short or the tool is in the never-prune list.
    """
    info = {"score": None, "verdict": "skipped",
            "chars_in": len(result), "chars_out": len(result)}
    condensed, cinfo = condense_tool_result(task, tool_name, result)
    if cinfo["verdict"] == "condensed":
        return condensed, cinfo
    if os.environ.get("QWEN_TRIAGE_ENABLE", "0") != "1":
        return result, info
    if tool_name in _TRIAGE_NEVER_PRUNE_TOOLS:
        return result, info
    # Don't bother scoring tiny results.
    min_chars = int(os.environ.get("QWEN_TRIAGE_MIN_CHARS", "2000"))
    if len(result) < min_chars:
        return result, info
    # Tool errors and refusals carry signal — don't drop them.
    if (result.startswith("[tool error]") or result.startswith("[error]")
            or result.startswith("[refused]") or result.startswith("[REFUSED")):
        return result, info
    if not task.strip():
        return result, info
    try:
        # Use the head of the result for embedding (bge-small caps at 512
        # tokens regardless). Tail of a wrong-page result is usually MORE
        # off-topic than the head, so the head is a fair proxy.
        result_head = result[:1500]
        vecs = _embed_texts([task, result_head])
        score = float((vecs[0] * vecs[1]).sum())  # both unit-norm
    except Exception:  # noqa: BLE001
        # Embedder unavailable — fail open.
        return result, info
    info["score"] = round(score, 3)
    # Default 0.55 fits bge-small's score distribution: it puts most English
    # text in 0.3-0.9 cosine range, so 0.55 is the natural inflection. Drop
    # to 0.50 to be more aggressive about pruning, raise to 0.60 to be more
    # conservative. Re-tune with empirical results from your own runs.
    threshold = float(os.environ.get("QWEN_TRIAGE_THRESHOLD", "0.55"))
    if score >= threshold:
        info["verdict"] = "kept"
        return result, info
    # Low relevance: keep a small head as evidence + replace the rest with
    # a stub. The head lets the model judge for itself if the triage was
    # wrong; the stub keeps tokens out of context.
    keep_chars = int(os.environ.get("QWEN_TRIAGE_KEEP_CHARS", "600"))
    head = result[:keep_chars]
    pruned = (
        f"{head}\n\n"
        f"…[triage: pruned {len(result) - keep_chars} chars — "
        f"relevance score {score:.2f} below threshold {threshold:.2f} "
        f"vs the current task. If you need the full result, re-run the "
        f"tool with a more specific query.]"
    )
    info["verdict"] = "low_relevance"
    info["chars_out"] = len(pruned)
    return pruned, info


_CONDENSE_TOOLS = frozenset({
    "web_fetch", "pdf_extract", "arxiv_fetch", "web_outline",
})


# Fetch mode plumbing. The model picks a mode in its tool call
# (`web_fetch(..., mode="numerical"|"semantic")`); web_fetch sets a
# module-level variable; condense_tool_result reads it and applies the
# corresponding retrieval profile. Default is "semantic" — the safe
# choice for prose-derived answers. Models pick "numerical" when the
# answer is stated verbatim in the source (a value, count, name, date,
# or list item) — bypassing the cross-encoder which can demote tabular /
# numeric chunks in favor of prose that merely describes them.
_DEFAULT_FETCH_MODE = "semantic"
_LAST_FETCH_MODE: str = _DEFAULT_FETCH_MODE

_MODE_PROFILES: dict[str, dict] = {
    # Verbatim lookup: lexical + structural only. Tables and numeric
    # content survive intact at top_k=10. CE rerank and sentence-extract
    # are bypassed.
    "numerical": {
        "top_k": 10,
        "continuity": 2,
        "use_rerank": False,
        "use_sentence_extract": False,
    },
    # Synthesis from prose: BM25 picks the candidate pool, cross-encoder
    # picks the answer-bearing passages within it, and sentence-extract
    # tightens prose chunks.
    "semantic": {
        "top_k": 10,
        "continuity": 2,
        "use_rerank": True,
        "use_sentence_extract": True,
    },
}


def _set_fetch_mode(mode: str | None) -> None:
    global _LAST_FETCH_MODE
    if mode in _MODE_PROFILES:
        _LAST_FETCH_MODE = mode
    else:
        _LAST_FETCH_MODE = _DEFAULT_FETCH_MODE


def _get_fetch_mode_profile() -> dict:
    return _MODE_PROFILES.get(_LAST_FETCH_MODE, _MODE_PROFILES[_DEFAULT_FETCH_MODE])


_RELEVANCE_STOP = frozenset({
    "about", "after", "again", "also", "because", "before", "between",
    "could", "from", "have", "into", "more", "must", "need", "only",
    "over", "should", "than", "that", "their", "there", "these",
    "this", "those", "through", "using", "what", "when", "where",
    "which", "while", "with", "would", "your",
})


def _tokenize_relevance_list(text: str) -> list[str]:
    """Tokenize preserving term frequency. BM25 needs the list (TF matters)
    even though the older set-based scorer didn't."""
    return [
        w for w in re.findall(r"[a-zA-Z][a-zA-Z0-9_+-]{2,}", text.lower())
        if w not in _RELEVANCE_STOP
    ]


def _tokenize_relevance(text: str) -> set[str]:
    return set(_tokenize_relevance_list(text))


# Hybrid retrieval state (BM25 lexical + cross-encoder semantic rerank).
# Both stages are lazy: rank_bm25 imports happen on first condense call that
# clears min_chars; the ONNX cross-encoder loads on first call where
# QWEN_CONDENSE_RERANK is enabled. Loading is guarded by a lock so concurrent
# tool dispatches don't race on the HF snapshot_download.
_RERANK_LOCK = threading.Lock()
_RERANK_SESSION = None
_RERANK_TOKENIZER = None
_RERANK_LOAD_FAILED = False


def _rerank_model_id() -> str:
    return os.environ.get(
        "QWEN_RERANK_MODEL", "Xenova/ms-marco-MiniLM-L-6-v2",
    )


def _rerank_load() -> bool:
    """Lazy-load the ONNX cross-encoder. Returns True if ready, False on any
    failure (network down, model missing, onnxruntime not installed). Failure
    is sticky for the process so we don't retry the download on every fetch."""
    global _RERANK_SESSION, _RERANK_TOKENIZER, _RERANK_LOAD_FAILED
    if _RERANK_SESSION is not None:
        return True
    if _RERANK_LOAD_FAILED:
        return False
    with _RERANK_LOCK:
        if _RERANK_SESSION is not None:
            return True
        if _RERANK_LOAD_FAILED:
            return False
        try:
            from huggingface_hub import snapshot_download
            from transformers import AutoTokenizer
            import onnxruntime as ort  # noqa: WPS433
            local = snapshot_download(
                _rerank_model_id(),
                allow_patterns=[
                    "tokenizer*", "vocab.txt", "special_tokens_map.json",
                    "config.json", "onnx/model_quantized.onnx",
                ],
            )
            tok = AutoTokenizer.from_pretrained(local)
            # CPU EP is the right pick: 22M-param MiniLM on Apple Silicon is
            # ~1.4ms/pair on CPU; CoreML EP adds ~200ms warmup and contends
            # with dflash-serve on the ANE for nothing.
            sess = ort.InferenceSession(
                f"{local}/onnx/model_quantized.onnx",
                providers=["CPUExecutionProvider"],
            )
            _RERANK_TOKENIZER = tok
            _RERANK_SESSION = sess
            return True
        except Exception:  # noqa: BLE001
            _RERANK_LOAD_FAILED = True
            return False


def _cross_encoder_rerank(query: str, chunks: list[str]) -> list[float] | None:
    """Score (query, chunk) pairs with the local cross-encoder. Returns a
    list of floats parallel to `chunks`, or None if the reranker isn't
    available (lets caller fall back to BM25-only ranking).

    Implementation: each chunk gets scored at BOTH its head and tail and
    we keep the MAX score. The cross-encoder has a hard 512-token cap, so
    a chunk where the answer-bearing value sits at the END (common pattern
    in SEC tables — preamble + header rows precede the row that has the
    actual number) would otherwise score low because the CE only saw the
    irrelevant preamble. Snap 10-K example: the "Convertible Notes
    (if-converted)" value (85,945) lives at char 1344 of a 1391-char
    chunk; the chunk's head is "Net loss" income-statement rows that
    don't match the dilution query. Scoring head+tail catches it."""
    if not _rerank_load():
        return None
    try:
        import numpy as np  # noqa: WPS433
    except Exception:  # noqa: BLE001
        return None
    # Cap fed to the cross-encoder per window: ~1.0 KB is comfortably
    # inside the 512-token cap even for dense numeric text (each digit,
    # $, comma, newline gets its own token, so 1000 chars ≈ 300-450
    # tokens).
    WINDOW = 1000

    def _score(c_list: list[str]) -> list[float] | None:
        try:
            enc = _RERANK_TOKENIZER(
                [query] * len(c_list), c_list,
                padding=True, truncation=True, max_length=512, return_tensors="np",
            )
            feed = {k: v.astype(np.int64) for k, v in enc.items()}
            out = _RERANK_SESSION.run(None, feed)
            arr = out[0]
            if arr.ndim == 2 and arr.shape[1] == 1:
                arr = arr.squeeze(-1)
            return arr.astype(float).tolist()
        except Exception:  # noqa: BLE001
            return None

    # Always score the head; for chunks long enough that the tail window
    # is different content, score the tail too and keep max. "Long enough"
    # = chunk longer than the head window (otherwise tail == head verbatim
    # and we'd just be paying for the same score twice).
    head_scores = _score([c[:WINDOW] for c in chunks])
    if head_scores is None:
        return None
    long_idx = [i for i, c in enumerate(chunks) if len(c) > WINDOW]
    if long_idx:
        tail_pairs = [chunks[i][-WINDOW:] for i in long_idx]
        tail_scores = _score(tail_pairs)
        if tail_scores is None:
            return head_scores  # graceful degrade
        merged = list(head_scores)
        for j, i in enumerate(long_idx):
            merged[i] = max(merged[i], tail_scores[j])
        return merged
    return head_scores


def _bm25_scores(query_tokens: list[str], chunk_tokens: list[list[str]]) -> list[float] | None:
    """BM25Okapi scores for chunks against query. Returns None if rank_bm25
    is unavailable so the caller can fall back to set-intersection scoring."""
    if not query_tokens or not any(chunk_tokens):
        return None
    try:
        from rank_bm25 import BM25Okapi  # noqa: WPS433
    except Exception:  # noqa: BLE001
        return None
    try:
        bm25 = BM25Okapi(chunk_tokens)
        return list(bm25.get_scores(query_tokens))
    except Exception:  # noqa: BLE001
        return None


# Sentence splitter for the optional second-pass cross-encoder. We need to be
# conservative around finance prose where periods appear in numbers ($10.50),
# abbreviations (Inc., Corp., Mr.), and section labels (Item 7.). The regex
# matches period/!/? followed by whitespace and an uppercase / quote /
# bracket / digit / dollar — a reasonable approximation of a sentence start
# that avoids splitting "$10.50" or "Item 7. Financial Statements" in the
# middle. The post-split merge keeps very short fragments attached to the
# preceding sentence so trailing parenthetical clauses don't get fragmented.
_SENT_SPLIT_RE = re.compile(r"(?<=[.!?])\s+(?=[A-Z(\"'\[#$\d])")


def _split_sentences(text: str) -> list[str]:
    """Split prose text into sentence-like chunks. Tuned to finance/SEC
    text where periods are heavily used inside numbers and abbreviations."""
    if not text:
        return []
    parts = _SENT_SPLIT_RE.split(text.strip())
    out: list[str] = []
    buf = ""
    for p in parts:
        p = p.strip()
        if not p:
            continue
        buf = p if not buf else f"{buf} {p}"
        # Don't emit micro-fragments — headers like "Item 7." are kept
        # attached to the following sentence instead of becoming a one-
        # word chunk that scores poorly on its own.
        if len(buf) >= 40:
            out.append(buf)
            buf = ""
    if buf:
        out.append(buf)
    return out


def _extract_top_sentences(query: str, chunk_text: str, top_n: int) -> str:
    """Run the cross-encoder over sentences within a chunk; return the top
    `top_n` sentences joined in original order. Falls back to the full
    chunk text if scoring fails or the chunk has too few sentences to
    benefit from filtering. Skips chunks that look like tables — they
    need cross-row context that sentence splitting would destroy."""
    if _looks_like_table(chunk_text):
        return chunk_text
    sentences = _split_sentences(chunk_text)
    if len(sentences) <= top_n:
        return chunk_text
    scores = _cross_encoder_rerank(query, sentences)
    if scores is None:
        return chunk_text
    # Pick the top_n by score (highest first), then re-sort by original
    # position so the output reads in the same order as the source.
    ranked = sorted(range(len(sentences)), key=lambda i: -scores[i])[:top_n]
    keep_set = set(ranked)
    return " ".join(sentences[i] for i in range(len(sentences)) if i in keep_set)


def _result_chunks(text: str, target_chars: int) -> list[str]:
    """Split text into paragraph-ish chunks near target_chars."""
    paras = [p.strip() for p in re.split(r"\n\s*\n", text) if p.strip()]
    if not paras:
        paras = [text[i:i + target_chars] for i in range(0, len(text), target_chars)]
    chunks: list[str] = []
    cur: list[str] = []
    cur_len = 0
    for para in paras:
        if cur and cur_len + len(para) > target_chars:
            chunks.append("\n\n".join(cur))
            cur = []
            cur_len = 0
        if len(para) > target_chars * 2:
            if cur:
                chunks.append("\n\n".join(cur))
                cur = []
                cur_len = 0
            chunks.extend(para[i:i + target_chars] for i in range(0, len(para), target_chars))
            continue
        cur.append(para)
        cur_len += len(para) + 2
    if cur:
        chunks.append("\n\n".join(cur))
    return chunks


_TABLE_ROW_RE = re.compile(r"^\s*\S.*?[|\t].*?[|\t]", re.MULTILINE)
_NUM_HEAVY_LINE_RE = re.compile(r"^\s*\S.*?\d[\d,.\s$%-]{4,}\s*\S*\s*$", re.MULTILINE)
_SECTION_HEADER_RE = re.compile(
    r"(^|\n)\s*("
    r"#+\s|"                                              # markdown
    r"(item|part|section|note)\s+\d+[A-Z.\-]*\s*[.:]\s*|"  # SEC-style
    r"title|abstract|summary|conclusion|results|discussion"
    r")", re.IGNORECASE,
)


def _looks_like_table(chunk: str) -> bool:
    """True if the chunk looks like tabular data: at least 3 lines that are
    pipe/tab-delimited rows, OR at least 4 lines dominated by numbers (e.g.
    SEC monthly-purchase tables that lose pipes after HTML-to-text).
    Both cheap regex; keeps tables from getting dropped by paragraph
    keyword-overlap ranking, which scores numeric content near zero."""
    if len(_TABLE_ROW_RE.findall(chunk)) >= 3:
        return True
    if len(_NUM_HEAVY_LINE_RE.findall(chunk)) >= 4:
        return True
    return False


def _looks_like_header(chunk: str) -> bool:
    """True if the chunk's leading lines look like a section heading.
    Used to drive section-continuity (keep the next chunks after a header).
    Header-only chunks tend to score well on keyword overlap with the
    task — but the data behind them lives in the FOLLOWING chunks, which
    score poorly in isolation. Without this signal we keep the header but
    drop its data."""
    head = chunk[:400]
    return bool(_SECTION_HEADER_RE.search(head))


def _structural_boost(chunk: str, idx: int) -> float:
    """Structural signals that complement lexical scoring: section headers
    score well on keyword overlap but their data lives in following chunks
    (handled by continuity), tables encode the actual numbers the user
    asks about but lose to bag-of-words because cells are mostly digits,
    and the lead chunk is the page's topical preamble."""
    score = 0.0
    if _looks_like_header(chunk):
        score += 1.5
    if _looks_like_table(chunk):
        score += 2.0
    if idx == 0:
        score += 1.0
    return score


def _chunk_score(task_terms: set[str], chunk: str, idx: int) -> float:
    """Legacy set-intersection scorer. Kept as a fallback path for when
    rank_bm25 is unavailable (and used by the existing test which exercises
    it directly via condense_tool_result's fallback)."""
    words = _tokenize_relevance(chunk)
    overlap = len(task_terms & words)
    return float(overlap) + _structural_boost(chunk, idx)


def condense_tool_result(task: str, tool_name: str, result: str) -> tuple[str, dict]:
    """Chunk long research outputs and keep the most task-relevant chunks.

    This is safer than the old whole-result triage: it never drops a large
    result solely because the head embeds poorly, and the model still sees a
    compact evidence bundle with chunk numbers and a re-fetch hint.
    """
    info = {
        "score": None,
        "verdict": "skipped",
        "chars_in": len(result),
        "chars_out": len(result),
        "chunks_in": 0,
        "chunks_kept": 0,
    }
    if os.environ.get("QWEN_RESULT_CONDENSE", "1") in ("0", "false", "False"):
        return result, info
    if tool_name not in _CONDENSE_TOOLS:
        return result, info
    if (result.startswith("[tool error]") or result.startswith("[error]")
            or result.startswith("[refused]") or result.startswith("[REFUSED")):
        return result, info
    min_chars = int(os.environ.get("QWEN_CONDENSE_MIN_CHARS", "12000"))
    if len(result) < min_chars:
        return result, info
    task_terms = _tokenize_relevance(task)
    if not task_terms:
        return result, info
    chunk_chars = int(os.environ.get("QWEN_CONDENSE_CHUNK_CHARS", "1400"))
    # Iter 43: retrieval profile is selected by the fetch mode that
    # web_fetch (or pdf_extract/arxiv_fetch) recorded for this call. The
    # model decides per tool call: "numerical" when the answer is stated
    # verbatim in the source (lexical+structural ranking, CE bypassed),
    # "semantic" when the answer is synthesized from prose (BM25 + CE +
    # sentence-extract). Env vars still override for A/B / debug.
    profile = _get_fetch_mode_profile()
    top_k = max(1, int(os.environ.get("QWEN_CONDENSE_TOP_K", str(profile["top_k"]))))
    # Iter 38: lead 1200 → 1800. The first chunk is the page intro
    # (title, abstract, first paragraph) — for SEC 10-Ks and earnings
    # press releases the high-level period / entity context lives there.
    # Cutting it at 1200 chars sometimes truncated mid-sentence in
    # critical preambles.
    lead_chars = int(os.environ.get("QWEN_CONDENSE_LEAD_CHARS", "1800"))
    chunks = _result_chunks(result, max(500, chunk_chars))
    # Iter 39: hybrid retrieval. BM25 (rank_bm25) replaces the old set-
    # intersection scorer for first-pass lexical relevance, and an ONNX
    # cross-encoder (Xenova/ms-marco-MiniLM-L-6-v2, 22 MB int8) re-scores
    # the top candidates from BM25 when QWEN_CONDENSE_RERANK=1. The
    # cross-encoder distinguishes "topic mentioned" from "answer present"
    # — the gap BM25 alone can't close (e.g. "ARPU is a key metric…"
    # vs "ARPU was $11.50"). Both stages degrade gracefully: BM25
    # falls back to set-intersection if rank_bm25 import fails, and
    # rerank silently skips if the ONNX model can't load.
    query_tokens = _tokenize_relevance_list(task)
    chunk_tokens = [_tokenize_relevance_list(c) for c in chunks]
    bm25 = _bm25_scores(query_tokens, chunk_tokens)
    if bm25 is None:
        # Fallback: legacy set-intersection scorer (test-mode + safety net).
        scored = [
            (_chunk_score(task_terms, chunk, i), i, chunk)
            for i, chunk in enumerate(chunks)
        ]
    else:
        scored = [
            (bm25[i] + _structural_boost(chunk, i), i, chunk)
            for i, chunk in enumerate(chunks)
        ]
    # Snapshot the per-chunk-INDEX score BEFORE any rerank pool
    # replacement. The continuity-span logic and the lead-chunk auto-keep
    # both look up scores by chunk index, but post-rerank `scored` gets
    # replaced with the pool (keyed by position-in-pool, not chunk
    # index). Indexing `scored[j]` for chunks outside the pool then
    # fires IndexError on long docs.
    score_by_idx: dict[int, float] = {i: s for s, i, _ in scored}
    if bm25 is not None:
        # Optional second stage: cross-encoder rerank. Only re-rank the
        # top candidates from BM25 (4x top_k by default) — the
        # cross-encoder is ~1.4ms/pair so 32 pairs ≈ 50 ms, cheap during
        # tool dispatch when the proxy is idle. Gated by the fetch-mode
        # profile (numerical → off, semantic → on), with env override.
        rerank_default = "1" if profile["use_rerank"] else "0"
        if os.environ.get("QWEN_CONDENSE_RERANK", rerank_default) in ("1", "true", "True"):
            rerank_pool = max(top_k * 4, int(os.environ.get("QWEN_CONDENSE_RERANK_POOL", "32")))
            top_idxs = [
                idx for _, idx, _ in sorted(scored, key=lambda x: (-x[0], x[1]))[:rerank_pool]
            ]
            top_chunks = [chunks[i] for i in top_idxs]
            ce = _cross_encoder_rerank(task, top_chunks)
            if ce is not None:
                # Iter 41: the kept set must come from the rerank pool. Earlier
                # we rescaled CE scores into the BM25 range and then re-ran the
                # top-K selection across all chunks — but with raw CE values
                # often -10..+5 vs BM25 8..16, rescaling could push CE
                # candidates *below* non-reranked BM25 chunks and lose the
                # actual answer-bearing chunk to its neighbors. Snap 10-K
                # caught this: target chunk at CE rank 4 lost to chunks at
                # hybrid rank 6-10 that weren't in the rerank pool.
                #
                # New behavior: the BM25 stage selects the pool of
                # candidates that "could plausibly be relevant." The CE
                # stage picks the top_k *within* that pool. Non-pool
                # chunks are out of contention (except for the auto-kept
                # lead chunk handled separately below). Structural boosts
                # still tip ties between candidates.
                pool_with_scores = [
                    (ce[j] + _structural_boost(chunks[top_idxs[j]], top_idxs[j]),
                     top_idxs[j],
                     chunks[top_idxs[j]])
                    for j in range(len(top_idxs))
                ]
                # Replace `scored` so the downstream top_k loop picks
                # from the rerank pool only. Non-pool chunks are gone
                # from contention (the lead chunk index 0 is auto-kept
                # below regardless of whether it survived rerank).
                scored = pool_with_scores
                info["rerank_used"] = True
                info["rerank_pool"] = len(top_idxs)
    keep: dict[int, tuple[float, str]] = {}
    if chunks:
        # Lead chunk auto-kept. Use its own pre-rerank score (not
        # scored[0] which after rerank-pool replacement is the
        # highest-scoring pool entry, not the lead chunk).
        keep[0] = (score_by_idx.get(0, 0.0), chunks[0][:lead_chars])
    for score, i, chunk in sorted(scored, key=lambda x: (-x[0], x[1]))[:top_k]:
        keep[i] = (score, chunk[:chunk_chars])
    # Section-continuity: when we kept a header-like chunk, also keep the
    # next 1–2 chunks after it. Headers score well on keyword overlap, but
    # the actual content (data tables, numeric breakdowns, follow-up prose)
    # lives in the following chunks which score poorly individually. Without
    # this, we keep "Item 5. Issuer Purchases of Equity Securities" but drop
    # the table that immediately follows.
    # Iter 39: pulled back 3 → 1 alongside the top_k trim. Continuity
    # exists for the case where a section header scores well on keyword
    # overlap but the data lives in the chunks immediately after. With
    # BM25 + rerank now scoring the body chunks correctly on their own
    # merit (TF-IDF doesn't tie header and follow-on chunks the way set-
    # intersection did), 1 trailing chunk is enough to bridge a table
    # split across a chunk boundary; more just dilutes the kept set with
    # near-context the model doesn't need.
    continuity_span = max(0, int(os.environ.get("QWEN_CONDENSE_HEADER_CONTINUITY", str(profile["continuity"]))))
    if continuity_span > 0:
        for i in list(keep.keys()):
            chunk_text = chunks[i] if i < len(chunks) else ""
            if not _looks_like_header(chunk_text):
                continue
            for j in range(i + 1, min(i + 1 + continuity_span, len(chunks))):
                if j not in keep:
                    keep[j] = (score_by_idx.get(j, 0.0), chunks[j][:chunk_chars])
    ordered = sorted(keep.items())
    # Concrete recovery hint: the previous "narrower query / higher max_chars"
    # text was misleading — web_fetch has no `query` arg, and the model would
    # waste turns trying to find one. Point at args that actually exist
    # (`max_chars`, `head_only=False`) plus the structural escape hatch
    # (web_outline → fetch the right sub-URL).
    # Iter 40: optional sentence-level second pass. After BM25 + cross-
    # encoder pick the top chunks, score sentences WITHIN each prose
    # chunk and keep only the top N. Cuts the bytes the model has to
    # decode (typically 3-5× on prose-heavy chunks) while preserving the
    # exact answer-bearing line. Skips tables (handled by
    # `_looks_like_table` — they need cross-row context). Off by default
    # so it can be A/B'd against the baseline; enable with
    # QWEN_CONDENSE_SENTENCE_EXTRACT=1.
    sent_extract_default = "1" if profile["use_sentence_extract"] else "0"
    sent_extract = os.environ.get("QWEN_CONDENSE_SENTENCE_EXTRACT", sent_extract_default) in ("1", "true", "True")
    sent_top_n = max(1, int(os.environ.get("QWEN_CONDENSE_SENTENCE_TOP_N", "3")))
    out: list[str] = [
        f"[condensed {tool_name}: kept {len(ordered)}/{len(chunks)} chunks, "
        f"{len(result)}→{{PENDING_LEN}} chars. If the section you need was "
        f"dropped: (1) call web_outline on the same URL to see the heading "
        f"structure and find a more targeted sub-URL, (2) or re-fetch with "
        f"max_chars=300000 to get more raw text, (3) or fetch a more specific "
        f"page (e.g. an exhibit, appendix, or section anchor URL). Do not "
        f"pass a `query` arg — web_fetch has none.]"
    ]
    n_sent_extracted = 0
    for i, (score, chunk) in ordered:
        body = chunk.strip()
        if sent_extract and i != 0:  # never sentence-extract the lead chunk
            extracted = _extract_top_sentences(task, body, sent_top_n)
            if len(extracted) < len(body):
                n_sent_extracted += 1
                body = extracted
        out.append(f"\n--- chunk {i + 1}/{len(chunks)} score={score:.1f} ---\n{body}")
    condensed = "\n".join(out)
    condensed = condensed.replace("{PENDING_LEN}", str(len(condensed)))
    # If condensing somehow fails to save at least 25%, fail open.
    if len(condensed) > len(result) * 0.75:
        return result, info
    info.update({
        "verdict": "condensed",
        "chars_out": len(condensed),
        "chunks_in": len(chunks),
        "chunks_kept": len(ordered),
        "mode": _LAST_FETCH_MODE,
    })
    return condensed, info


def _memory_connect():
    import sqlite3
    path = _memory_db_path()
    conn = sqlite3.connect(path)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS memories (
            key          TEXT PRIMARY KEY,
            content      TEXT NOT NULL,
            tags         TEXT NOT NULL DEFAULT '',
            created_at   REAL NOT NULL,
            updated_at   REAL NOT NULL,
            embed_model  TEXT NOT NULL,
            embed_dim    INTEGER NOT NULL,
            embedding    BLOB NOT NULL
        )
    """)
    return conn


def _ensure_compatible_model(conn) -> None:
    """If the DB has entries embedded with a different model than we're
    currently loading, warn loudly. Mixing models in the same vector space
    gives nonsense distances. The user can re-embed by deleting + re-saving."""
    cur = conn.execute("SELECT DISTINCT embed_model FROM memories LIMIT 5")
    rows = [r[0] for r in cur.fetchall()]
    cur_name = _memory_embed_model_name()
    foreign = [r for r in rows if r and r != cur_name]
    if foreign:
        # Print once per process. Not a hard error so users can read existing
        # entries by `memory_get(key)` (exact-key lookup doesn't use vectors).
        if not _embed_state.get("_mixed_warned"):
            print(
                f"[memory warning] DB contains entries from model(s) {foreign}; "
                f"current model is {cur_name!r}. Semantic search across mixed "
                "models is unreliable — re-embed (delete + memory_save again) "
                "or set QWEN_EMBED_MODEL back to the original."
            )
            _embed_state["_mixed_warned"] = True


def memory_save(key: str, content: str, tags: str = "") -> str:
    """Upsert a memory entry. The content is embedded once at save time;
    subsequent searches are cheap dot products against the stored vector."""
    import time, numpy as np
    if not key or not key.strip():
        return "[error] key required"
    if not content or not content.strip():
        return "[error] content required"
    # Embed key + tags + content together. Including tags is critical for
    # short / generic queries: "my arxiv paper" has very little overlap
    # with a memory whose CONTENT is full of math-jargon, but its TAGS
    # contain the word "paper" so the embedding gains a useful anchor.
    parts = [key]
    if tags and tags.strip():
        parts.append(f"Tags: {tags}")
    parts.append(content)
    embed_text = "\n\n".join(parts)
    vec = _embed_texts([embed_text])[0].astype(np.float32)
    _, _, name, dim = _embed_load()
    now = time.time()
    blob = vec.tobytes()
    with _memory_connect() as conn:
        _ensure_compatible_model(conn)
        cur = conn.execute("SELECT 1 FROM memories WHERE key = ?", (key,))
        existed = cur.fetchone() is not None
        if existed:
            conn.execute(
                "UPDATE memories SET content=?, tags=?, updated_at=?, "
                "embed_model=?, embed_dim=?, embedding=? WHERE key=?",
                (content, tags, now, name, dim, blob, key),
            )
        else:
            conn.execute(
                "INSERT INTO memories(key, content, tags, created_at, updated_at, "
                "embed_model, embed_dim, embedding) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (key, content, tags, now, now, name, dim, blob),
            )
    return f"{'updated' if existed else 'saved'} memory {key!r} "\
           f"({len(content)} chars, {dim}-d via {name.split('/')[-1]})"


def memory_get(key: str) -> str:
    """Fetch a single memory by exact key (no embedding needed)."""
    if not key:
        return "[error] key required"
    with _memory_connect() as conn:
        cur = conn.execute(
            "SELECT content, tags, created_at, updated_at, embed_model "
            "FROM memories WHERE key = ?",
            (key,),
        )
        row = cur.fetchone()
    if row is None:
        return f"(no memory with key {key!r})"
    content, tags, created_at, updated_at, model = row
    import datetime as _dt
    out = [
        f"key: {key}",
        f"created: {_dt.datetime.fromtimestamp(created_at).isoformat(timespec='seconds')}",
        f"updated: {_dt.datetime.fromtimestamp(updated_at).isoformat(timespec='seconds')}",
        f"embed_model: {model}",
    ]
    if tags:
        out.append(f"tags: {tags}")
    out.append("")
    out.append(content)
    return "\n".join(out)


# Common English stopwords — filtered out of query/document tokens for the
# lexical scoring component. Kept short and conservative; we want
# discriminative keywords ("paper", "arxiv", "trader", model names) to count,
# but not "the", "is", "what".
_MEMORY_STOPWORDS = frozenset("""
a about above after again against all am an and any are as at be because been
before being below between both but by can did do does doing don down during
each few for from further had has have having he her here hers him his how i
if in into is it its itself just like make me more most my myself need no nor
not now of off on once only or other our ours out over own please same she
should so some such tell than that the their them then there these they this
those through to too under until up use used very was we were what when where
which while who whom why will with would you your yours
""".split())

# Words that look like time/relative-day chatter; ignored for matching but
# kept as low signal so we don't accidentally drop names or terms.
_MEMORY_GENERIC_PRONOUNS = frozenset({"my", "mine", "yours", "ours"})


def _tokenize_for_search(text: str) -> list[str]:
    if not text:
        return []
    # Lowercase + split on word boundaries; keep alphanumerics and dashes.
    raw = re.findall(r"[A-Za-z0-9][A-Za-z0-9_\-]+", text.lower())
    return [t for t in raw if t not in _MEMORY_STOPWORDS and len(t) >= 2]


def _keyword_score(query_tokens: list[str], doc_text: str) -> float:
    """Bounded lexical relevance — fraction of query tokens that appear in
    the doc, with a small bonus for repeated hits. Returns a value in [0, 1].
    Cheap to compute and dramatically improves ranking when bge-small can't
    distinguish abstract queries (e.g. "my arxiv paper")."""
    if not query_tokens:
        return 0.0
    doc_tokens = _tokenize_for_search(doc_text)
    if not doc_tokens:
        return 0.0
    doc_freq: dict[str, int] = {}
    for t in doc_tokens:
        doc_freq[t] = doc_freq.get(t, 0) + 1
    matches = 0
    bonus = 0.0
    for qt in set(query_tokens):
        if qt in doc_freq:
            matches += 1
            # Tiny bonus when the term appears multiple times.
            bonus += min(doc_freq[qt] - 1, 3) * 0.02
    base = matches / len(set(query_tokens))
    return min(1.0, base + bonus)


def memory_search(query: str, max_results: int = 5, tag: str | None = None,
                   min_score: float | None = None,
                   alpha: float | None = None) -> str:
    """Hybrid semantic + lexical search over stored memories.

    The final score per memory is:
        score = alpha * cosine(qvec, memory_vec) + (1 - alpha) * keyword_overlap

    where `keyword_overlap` is the fraction of query tokens that appear in the
    memory's (key + tags + content), with a small bonus for repeated hits.
    Pure cosine alone is unreliable on small embedding models (bge-small):
    queries like "my arxiv paper" can rank a trader-related memory ABOVE
    the actual paper memory. The lexical term anchors generic queries.

    Only entries with final score above `min_score` are returned (default
    0.55). When NO memory clears the threshold, the closest few are still
    shown but tagged BELOW THRESHOLD so the model doesn't treat them as
    authoritative recall.

    `alpha` defaults to 0.6 (semantic-leaning hybrid). Set 1.0 for pure
    cosine, 0.0 for pure keyword overlap.

    Linear scan: fine up to ~50k entries. Replace with an ANN index later.
    """
    import numpy as np
    if not query or not query.strip():
        return "[error] empty query"
    if min_score is None:
        try:
            min_score = float(os.environ.get("QWEN_MEMORY_MIN_SCORE", "0.55"))
        except ValueError:
            min_score = 0.55
    if alpha is None:
        try:
            alpha = float(os.environ.get("QWEN_MEMORY_ALPHA", "0.6"))
        except ValueError:
            alpha = 0.6
    alpha = max(0.0, min(1.0, alpha))

    query_tokens = _tokenize_for_search(query)
    qvec = _embed_texts([query])[0].astype(np.float32)
    sql = "SELECT key, content, tags, updated_at, embed_dim, embedding FROM memories"
    args: list = []
    if tag:
        sql += " WHERE tags LIKE ?"
        args.append(f"%{tag}%")
    with _memory_connect() as conn:
        _ensure_compatible_model(conn)
        rows = conn.execute(sql, args).fetchall()
    if not rows:
        return "(no memories stored)"

    # Each scored entry: (final, cos, kw, key, content, tags, updated_at)
    scored: list[tuple[float, float, float, str, str, str, float]] = []
    for key, content, tags_, updated_at, dim, blob in rows:
        try:
            v = np.frombuffer(blob, dtype=np.float32)
            if v.size != dim or v.size != qvec.size:
                continue
            cos = float((qvec * v).sum())  # both pre-normalized → cosine
        except Exception:  # noqa: BLE001
            continue
        kw = _keyword_score(query_tokens, f"{key}\n{tags_ or ''}\n{content}")
        final = alpha * cos + (1.0 - alpha) * kw
        scored.append((final, cos, kw, key, content, tags_, updated_at))
    if not scored:
        return "(no compatible memories)"
    scored.sort(key=lambda x: x[0], reverse=True)

    above = [s for s in scored if s[0] >= min_score]
    below = len(scored) - len(above)
    top = above[: max(1, int(max_results))]

    import datetime as _dt
    if not top:
        nearest = scored[:3]
        out = [f"(no memories scored ≥ {min_score:.2f} for query — "
               f"closest 3 shown below threshold; lower QWEN_MEMORY_MIN_SCORE or "
               f"refine the query)"]
        for (final, cos, kw, key, content, tags_, updated_at) in nearest:
            when = _dt.datetime.fromtimestamp(updated_at).strftime("%Y-%m-%d")
            tag_str = f" [{tags_}]" if tags_ else ""
            snippet = " ".join(content.split())
            if len(snippet) > 160:
                snippet = snippet[:160] + "…"
            out.append(f"  ~ {key}  (score={final:+.3f} cos={cos:+.3f} kw={kw:+.3f}, "
                       f"{when}, BELOW THRESHOLD){tag_str}")
            out.append(f"      {snippet}")
        return "\n".join(out)

    out = []
    for i, (final, cos, kw, key, content, tags_, updated_at) in enumerate(top, 1):
        when = _dt.datetime.fromtimestamp(updated_at).strftime("%Y-%m-%d")
        tag_str = f" [{tags_}]" if tags_ else ""
        snippet = " ".join(content.split())
        if len(snippet) > 240:
            snippet = snippet[:240] + "…"
        out.append(f"{i}. {key}  (score={final:+.3f} cos={cos:+.3f} kw={kw:+.3f}, {when}){tag_str}")
        out.append(f"   {snippet}")
    if below:
        out.append(f"   ({below} additional below score<{min_score:.2f} omitted)")
    return "\n".join(out)


def memory_reembed(rebuild_all: bool = False) -> str:
    """Re-embed every memory entry with the current embedding scheme
    (key + tags + content) using the currently-loaded embedding model.
    Use this after upgrading the embedder OR when older entries were
    embedded without the tag-anchor (the user-facing "my arxiv paper"
    fix). Pass rebuild_all=False (default) to only re-embed entries that
    are missing tag content in their stored vector."""
    import time, numpy as np
    _, _, name, dim = _embed_load()
    n_seen = n_redone = 0
    with _memory_connect() as conn:
        _ensure_compatible_model(conn)
        rows = conn.execute(
            "SELECT key, content, tags, embed_model, embed_dim FROM memories"
        ).fetchall()
        for key, content, tags_, model_, dim_ in rows:
            n_seen += 1
            if not rebuild_all and model_ == name and dim_ == dim:
                # Heuristic: skip if same model + dim AND tags is empty (the
                # only way the new key+tags+content embed differs from the
                # old key+content embed is when tags are non-empty).
                if not (tags_ or "").strip():
                    continue
            parts = [key]
            if (tags_ or "").strip():
                parts.append(f"Tags: {tags_}")
            parts.append(content)
            embed_text = "\n\n".join(parts)
            vec = _embed_texts([embed_text])[0].astype(np.float32)
            blob = vec.tobytes()
            conn.execute(
                "UPDATE memories SET embed_model=?, embed_dim=?, embedding=?, "
                "updated_at=updated_at WHERE key=?",
                (name, dim, blob, key),
            )
            n_redone += 1
        conn.commit()
    return f"re-embedded {n_redone}/{n_seen} memories with {name.split('/')[-1]}"


def memory_list(limit: int = 20, tag: str | None = None) -> str:
    """List recent memory keys (most recently updated first). Optional tag
    substring filter. No embedding involved — fast even for large stores."""
    sql = "SELECT key, tags, updated_at FROM memories"
    args: list = []
    if tag:
        sql += " WHERE tags LIKE ?"
        args.append(f"%{tag}%")
    sql += " ORDER BY updated_at DESC LIMIT ?"
    args.append(int(limit))
    with _memory_connect() as conn:
        rows = conn.execute(sql, args).fetchall()
    if not rows:
        return "(no memories stored)"
    import datetime as _dt
    out = []
    for key, tags_, updated_at in rows:
        when = _dt.datetime.fromtimestamp(updated_at).strftime("%Y-%m-%d %H:%M")
        tag_str = f"  [{tags_}]" if tags_ else ""
        out.append(f"{when}  {key}{tag_str}")
    return "\n".join(out)


def memory_delete(key: str) -> str:
    """Delete a memory entry by exact key."""
    if not key:
        return "[error] key required"
    with _memory_connect() as conn:
        cur = conn.execute("DELETE FROM memories WHERE key = ?", (key,))
        n = cur.rowcount
    return f"deleted {n} memory entr{'y' if n == 1 else 'ies'} ({key!r})"


# ---------- todo_write (session task list, ported from Claude Code) -------

def _todo_path() -> str:
    p = os.environ.get("QWEN_TODO_FILE")
    if p:
        return p
    return f"/tmp/qwen_todos_{os.getpid()}.json"


def todo_write(todos: list) -> str:
    """Update the agent's structured task list for the current session.

    Each todo: {"content": str (imperative), "activeForm": str (present-continuous),
                "status": "pending" | "in_progress" | "completed"}.
    Exactly one task should be in_progress at a time.
    Stored in /tmp/qwen_todos_<pid>.json (override via QWEN_TODO_FILE env)
    so the agent can read it back across turns. Renders the list back as a
    human-readable checklist.
    """
    if not isinstance(todos, list):
        return "[error] todos must be a list"
    valid = []
    for i, t in enumerate(todos):
        if not isinstance(t, dict):
            return f"[error] todo {i} not an object"
        content = t.get("content")
        active = t.get("activeForm") or content
        status = t.get("status", "pending")
        if not content:
            return f"[error] todo {i} missing 'content'"
        if status not in ("pending", "in_progress", "completed"):
            return f"[error] todo {i} invalid status '{status}' (must be pending|in_progress|completed)"
        valid.append({"content": content, "activeForm": active, "status": status})
    in_progress = [t for t in valid if t["status"] == "in_progress"]
    if len(in_progress) > 1:
        return f"[error] only ONE todo can be in_progress; you have {len(in_progress)}. Mark others pending or completed."
    path = _todo_path()
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(valid, f, indent=2)
    except OSError as e:
        return f"[error] writing todo file: {e}"
    if not valid:
        return "(empty todo list)"
    glyph = {"completed": "[x]", "in_progress": "[>]", "pending": "[ ]"}
    lines = [f"  {glyph[t['status']]} {t['content']}" for t in valid]
    n_done = sum(1 for t in valid if t["status"] == "completed")
    return f"updated todo list ({n_done}/{len(valid)} done):\n" + "\n".join(lines)


# ---------- worktree (git worktree or sandbox dir, ported from Claude Code)

_WORKTREE_STATE = "/tmp/qwen_worktree_state.json"


def enter_worktree(name: str = "") -> str:
    """Create a git worktree (or sandbox copy if not in a git repo) and chdir into it.

    Subsequent edit_file/write_file/bash will operate in the worktree, leaving the
    original tree untouched. When done, call exit_worktree to restore cwd and
    remove (or keep) the worktree.
    """
    if os.path.exists(_WORKTREE_STATE):
        with open(_WORKTREE_STATE) as f:
            st = json.load(f)
        return f"[error] already in a worktree session at {st.get('worktree_path')}; call exit_worktree first"
    cwd = os.getcwd()
    try:
        root = subprocess.check_output(
            ["git", "rev-parse", "--show-toplevel"],
            cwd=cwd, stderr=subprocess.DEVNULL, timeout=5
        ).decode().strip()
        is_git = bool(root)
    except (subprocess.CalledProcessError, FileNotFoundError, subprocess.TimeoutExpired):
        is_git = False
        root = cwd
    import time as _t
    slug = name or f"qwen-{int(_t.time())}"
    # sanitize slug
    slug = re.sub(r"[^a-zA-Z0-9._-]", "-", slug)[:64]
    base = "/tmp/qwen_worktree"
    os.makedirs(base, exist_ok=True)
    wt_path = os.path.join(base, slug)
    if is_git:
        if os.path.exists(wt_path):
            shutil.rmtree(wt_path, ignore_errors=True)
        try:
            subprocess.check_output(
                ["git", "worktree", "add", "-b", f"qwen-wt-{slug}", wt_path, "HEAD"],
                cwd=root, stderr=subprocess.STDOUT, timeout=20
            )
        except subprocess.CalledProcessError as e:
            return f"[error] git worktree add failed: {e.output.decode(errors='replace')}"
        os.chdir(wt_path)
        with open(_WORKTREE_STATE, "w") as f:
            json.dump({"original_cwd": cwd, "worktree_path": wt_path,
                       "git_root": root, "branch": f"qwen-wt-{slug}", "is_git": True}, f)
        return (f"Created git worktree at {wt_path} (branch qwen-wt-{slug}).\n"
                f"Working dir is now the worktree; the original tree at {root} is untouched.\n"
                f"Use exit_worktree to leave.")
    else:
        if os.path.exists(wt_path):
            shutil.rmtree(wt_path, ignore_errors=True)
        ignore = shutil.ignore_patterns(
            ".git", "__pycache__", "*.pyc", "node_modules", ".venv", "venv",
            ".pytest_cache", ".mypy_cache", ".tox", "dist", "build"
        )
        try:
            shutil.copytree(cwd, wt_path, ignore=ignore)
        except OSError as e:
            return f"[error] copytree failed: {e}"
        os.chdir(wt_path)
        with open(_WORKTREE_STATE, "w") as f:
            json.dump({"original_cwd": cwd, "worktree_path": wt_path, "is_git": False}, f)
        return (f"Created sandbox copy at {wt_path} (no git repo detected; copied {cwd}).\n"
                f"Working dir switched. Use exit_worktree to leave.")


def exit_worktree(keep: bool = False) -> str:
    """Leave the worktree, restore original cwd. keep=True preserves files; default cleans them."""
    if not os.path.exists(_WORKTREE_STATE):
        return "[error] not in a worktree session"
    with open(_WORKTREE_STATE) as f:
        state = json.load(f)
    try:
        os.chdir(state["original_cwd"])
    except OSError as e:
        return f"[error] chdir back failed: {e}"
    parts = [f"Restored cwd to {state['original_cwd']}."]
    if not keep:
        try:
            if state.get("is_git"):
                try:
                    subprocess.check_output(
                        ["git", "worktree", "remove", "--force", state["worktree_path"]],
                        cwd=state["git_root"], stderr=subprocess.STDOUT, timeout=20
                    )
                except subprocess.CalledProcessError as e:
                    parts.append(f"[warn] git worktree remove failed: {e.output.decode(errors='replace')}")
                try:
                    subprocess.check_output(
                        ["git", "branch", "-D", state["branch"]],
                        cwd=state["git_root"], stderr=subprocess.DEVNULL, timeout=5
                    )
                except subprocess.CalledProcessError:
                    pass
                parts.append(f"Removed worktree {state['worktree_path']} and branch {state['branch']}.")
            else:
                shutil.rmtree(state["worktree_path"], ignore_errors=True)
                parts.append(f"Removed sandbox {state['worktree_path']}.")
        except Exception as e:  # noqa: BLE001
            parts.append(f"[warn] cleanup error: {e}")
    else:
        parts.append(f"Preserved {state['worktree_path']}.")
    try:
        os.unlink(_WORKTREE_STATE)
    except OSError:
        pass
    return " ".join(parts)


# ---------- notebook (Jupyter) edit + run ---------------------------------

def notebook_edit(path: str, source: str = "", cell_index: int = -1,
                  cell_id: str = "", action: str = "replace",
                  cell_type: str = "code") -> str:
    """Edit a Jupyter notebook .ipynb file.

    action: 'replace' | 'insert_after' | 'insert_before' | 'delete' | 'append'.
    Specify the target cell by cell_index (0-based) or cell_id. If neither given
    and action='append', a new cell is added at the end.
    cell_type: 'code' (default) or 'markdown'.
    Auto-creates the notebook if it doesn't exist (only for append/insert).
    """
    try:
        import nbformat
    except ImportError:
        return "[error] nbformat not installed; pip install nbformat"
    if not path.endswith(".ipynb"):
        return "[error] path must end with .ipynb"
    if os.path.exists(path):
        with open(path) as f:
            nb = nbformat.read(f, as_version=4)
    else:
        if action not in ("append", "insert_after", "insert_before"):
            return f"[error] notebook not found at {path}; use action='append' to create it"
        nb = nbformat.v4.new_notebook()
    target = None
    if cell_id:
        for i, c in enumerate(nb.cells):
            if c.get("id") == cell_id:
                target = i
                break
        if target is None:
            return f"[error] cell_id {cell_id!r} not found"
    elif cell_index >= 0:
        if cell_index >= len(nb.cells):
            return f"[error] cell_index {cell_index} out of range (have {len(nb.cells)} cells)"
        target = cell_index

    def _new_cell(src):
        return nbformat.v4.new_markdown_cell(src) if cell_type == "markdown" else nbformat.v4.new_code_cell(src)

    if action == "replace":
        if target is None:
            return "[error] specify cell_index or cell_id for action=replace"
        nb.cells[target].source = source
    elif action == "delete":
        if target is None:
            return "[error] specify cell_index or cell_id for action=delete"
        del nb.cells[target]
    elif action == "insert_after":
        new = _new_cell(source)
        if target is None:
            nb.cells.append(new)
        else:
            nb.cells.insert(target + 1, new)
    elif action == "insert_before":
        new = _new_cell(source)
        if target is None:
            nb.cells.insert(0, new)
        else:
            nb.cells.insert(target, new)
    elif action == "append":
        nb.cells.append(_new_cell(source))
    else:
        return f"[error] unknown action {action!r}; use replace|insert_after|insert_before|delete|append"

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w") as f:
        nbformat.write(nb, f)
    _track_write(path)
    return f"updated {path}: action={action}, total cells={len(nb.cells)}"


def notebook_run(path: str, timeout: int = 120) -> str:
    """Execute every cell in a notebook in a fresh kernel, save the outputs back, and return them."""
    try:
        import nbformat
        from nbclient import NotebookClient
    except ImportError:
        return "[error] need nbformat + nbclient; pip install nbformat nbclient"
    if not os.path.exists(path):
        return f"[error] notebook not found: {path}"
    with open(path) as f:
        nb = nbformat.read(f, as_version=4)
    client = NotebookClient(nb, timeout=timeout, kernel_name="python3",
                            allow_errors=True, resources={"metadata": {"path": os.path.dirname(path) or "."}})
    try:
        client.execute()
    except Exception as e:  # noqa: BLE001
        with open(path, "w") as f:
            nbformat.write(nb, f)
        return f"[execution failed] {type(e).__name__}: {e}"
    with open(path, "w") as f:
        nbformat.write(nb, f)
    out_lines = []
    n_err = 0
    for i, c in enumerate(nb.cells):
        if c.cell_type != "code":
            continue
        outs = c.get("outputs", [])
        if not outs:
            continue
        out_lines.append(f"--- cell {i} ---")
        for o in outs:
            t = o.get("output_type")
            if t == "stream":
                out_lines.append(o.get("text", "").rstrip())
            elif t == "execute_result":
                d = o.get("data", {}).get("text/plain", "")
                out_lines.append(d.rstrip() if isinstance(d, str) else str(d).rstrip())
            elif t == "display_data":
                # for matplotlib etc. — note the mime types present
                mt = ", ".join(o.get("data", {}).keys())
                out_lines.append(f"[display: {mt}]")
            elif t == "error":
                n_err += 1
                ename = o.get("ename", "?")
                evalue = o.get("evalue", "")
                out_lines.append(f"[error] {ename}: {evalue}")
    summary = f"executed {sum(1 for c in nb.cells if c.cell_type=='code')} code cells"
    if n_err:
        summary += f" ({n_err} errored)"
    out = summary + "\n" + "\n".join(out_lines)
    if len(out) > 50000:
        out = out[:50000] + f"\n…[truncated {len(out) - 50000} chars]"
    return out


# ---------- persistent Python kernel (REPL with cross-call state) ---------
# A jupyter kernel that lives for the agent process lifetime. Variables,
# imports, and open file handles persist across python_run calls — so the
# model can iteratively explore data without restarting Python each time.
# Generalizable: helps with data analysis (s5/h5), algorithm dev (h4/h11),
# reverse engineering (h6), bug hunting (h12), self-improvement (h10).

_kernel_state: dict[str, Any] = {"manager": None, "client": None, "started_at": None}


def _kernel_ensure():
    if _kernel_state["client"] is not None:
        return _kernel_state["client"]
    try:
        from jupyter_client import KernelManager
    except ImportError:
        raise RuntimeError("jupyter_client not installed; pip install jupyter_client ipykernel")
    km = KernelManager(kernel_name="python3")
    km.start_kernel()
    kc = km.client()
    kc.start_channels()
    try:
        kc.wait_for_ready(timeout=15)
    except Exception as e:
        kc.stop_channels()
        km.shutdown_kernel(now=True)
        raise RuntimeError(f"kernel failed to start: {e}")
    import time as _t
    _kernel_state["manager"] = km
    _kernel_state["client"] = kc
    _kernel_state["started_at"] = _t.time()
    return kc


def python_run(code: str, timeout: int = 60) -> str:
    """Execute Python in a persistent kernel that lives for the agent process.

    State persists across calls. Use to iteratively explore data, build up an
    analysis without re-running setup, or debug interactively. Outputs (stdout,
    return values, errors with tracebacks) are streamed back. For matplotlib
    figures, use plt.savefig() — bare display data is captured as a marker only.
    """
    if not code or not code.strip():
        return "[error] empty code"
    try:
        kc = _kernel_ensure()
    except RuntimeError as e:
        return f"[error] {e}"
    msg_id = kc.execute(code)
    outputs: list[str] = []
    n_errors = 0
    import time as _t
    deadline = _t.time() + timeout
    while True:
        remaining = deadline - _t.time()
        if remaining <= 0:
            outputs.append(f"[timeout after {timeout}s; kernel still running, state preserved]")
            try:
                _kernel_state["manager"].interrupt_kernel()
            except Exception:  # noqa: BLE001
                pass
            break
        try:
            msg = kc.get_iopub_msg(timeout=min(remaining, 5))
        except Exception:
            continue
        if msg.get("parent_header", {}).get("msg_id") != msg_id:
            continue
        msg_type = msg["msg_type"]
        content = msg["content"]
        if msg_type == "stream":
            outputs.append(content.get("text", ""))
        elif msg_type == "execute_result":
            outputs.append(str(content.get("data", {}).get("text/plain", "")))
        elif msg_type == "display_data":
            mts = ", ".join(content.get("data", {}).keys())
            outputs.append(f"[display: {mts}]")
        elif msg_type == "error":
            n_errors += 1
            ename = content.get("ename", "?")
            evalue = content.get("evalue", "")
            tb = content.get("traceback", [])
            # strip ANSI escape codes from tracebacks for readability
            tb_clean = [re.sub(r"\x1b\[[0-9;]*m", "", line) for line in tb]
            outputs.append(f"{ename}: {evalue}\n" + "\n".join(tb_clean))
        elif msg_type == "status" and content.get("execution_state") == "idle":
            break
    text = "".join(outputs).rstrip()
    if not text:
        text = "(no output)"
    if len(text) > 50000:
        text = text[:50000] + f"\n…[truncated {len(text) - 50000} chars]"
    if n_errors:
        text = f"[{n_errors} error(s)]\n{text}"
    return text


def python_reset() -> str:
    """Restart the persistent Python kernel (clears all state)."""
    km = _kernel_state.get("manager")
    kc = _kernel_state.get("client")
    if kc is not None:
        try:
            kc.stop_channels()
        except Exception:  # noqa: BLE001
            pass
    if km is not None:
        try:
            km.shutdown_kernel(now=True)
        except Exception:  # noqa: BLE001
            pass
    _kernel_state["manager"] = None
    _kernel_state["client"] = None
    _kernel_state["started_at"] = None
    return "kernel reset; next python_run will start a fresh kernel"


# ---------- structured test runner --------------------------------------
# Wraps pytest with parsed pass/fail/error counts and short tracebacks for
# failing tests. Generalizable: useful any time tests are run.

def test_run(path: str = ".", k: str = "", timeout: int = 90) -> str:
    """Run pytest with structured output: per-test pass/fail/error + short tracebacks.

    `path` can be a file or directory (default: cwd).
    `k` is a pytest -k expression to select a subset.
    Returns: passed/failed/errored counts plus short tracebacks for failing tests.
    """
    import sys as _sys
    cmd = [_sys.executable, "-m", "pytest", path, "-v", "--tb=short", "--no-header"]
    if k:
        cmd.extend(["-k", k])
    try:
        rc = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout, cwd=os.getcwd())
    except subprocess.TimeoutExpired:
        return f"[timeout after {timeout}s]"
    out = rc.stdout + ("\n" + rc.stderr if rc.stderr else "")
    # Parse summary numbers anywhere in the output ("4 passed in 0.00s" format).
    n_pass = n_fail = n_err = 0
    m = re.search(r"(\d+)\s+passed", out); n_pass = int(m.group(1)) if m else 0
    m = re.search(r"(\d+)\s+failed", out); n_fail = int(m.group(1)) if m else 0
    m = re.search(r"(\d+)\s+error", out);  n_err = int(m.group(1)) if m else 0
    # Extract short summary if present
    summary_match = re.search(r"=+ (?:short test summary info|FAILURES) =+\n(.*?)(?:\n=+|\Z)", out, re.S)
    summary = summary_match.group(1).strip() if summary_match else ""
    last_lines = "\n".join(out.splitlines()[-10:])
    parts = [f"pytest exit={rc.returncode} | passed={n_pass} failed={n_fail} errors={n_err}"]
    if summary:
        parts.append("--- summary ---\n" + summary[:30000])
    if rc.returncode != 0 and not summary:
        parts.append("--- last lines ---\n" + last_lines)
    out_text = "\n\n".join(parts)
    if len(out_text) > 50000:
        out_text = out_text[:50000] + f"\n…[truncated]"
    return out_text


# ---------- append_finding (incremental artifact writer) ----------------
# Generalizable cure for the "read 12 files, write 0" failure mode. Lets the
# agent commit findings as it discovers them — by the time investigation
# ends, the artifact is already ~80% written. Cheap to call, low friction,
# naturally fits any research/review workflow.

def append_finding(path: str, heading: str, content: str,
                   create_with_title: str = "") -> str:
    """Append a section (## heading + content) to a markdown artifact.

    On first call, creates the file. If `create_with_title` is given, emits
    `# {create_with_title}\\n\\n` as line 1. Each call appends:

        \\n## {heading}\\n\\n{content}\\n

    Idempotent on duplicate (heading, content) pairs — won't double-add.

    Use freely while investigating. Don't batch all your findings into one
    big write_file at the end — append them as you find them. By the time
    you're done reading, your artifact will already be most of the way done.
    """
    if not path or not heading:
        return "[error] path and heading required"
    p = path
    parent = os.path.dirname(p) or "."
    os.makedirs(parent, exist_ok=True)
    is_new = not os.path.exists(p)
    existing = ""
    if not is_new:
        try:
            existing = open(p, encoding="utf-8").read()
        except OSError as e:
            return f"[error] reading existing artifact: {e}"
    # idempotent dedup
    section_marker = f"\n## {heading.strip()}\n\n{content.rstrip()}\n"
    if section_marker.strip() in existing:
        return f"already-present: section {heading!r} already in {p}"
    parts = []
    if is_new and create_with_title:
        parts.append(f"# {create_with_title.strip()}\n")
    elif is_new:
        # Default title from filename
        base = os.path.splitext(os.path.basename(p))[0].replace("_", " ").title()
        parts.append(f"# {base}\n")
    if existing and not existing.endswith("\n"):
        existing += "\n"
    parts.append(section_marker)
    new_content = (existing + "".join(parts)) if not is_new else "".join(parts)
    try:
        with open(p, "w", encoding="utf-8") as f:
            f.write(new_content)
    except OSError as e:
        return f"[error] writing: {e}"
    _track_write(p)
    n_sections = new_content.count("\n## ")
    n_words = len(new_content.split())
    action = "created" if is_new else "appended"
    return f"{action} section {heading!r} → {p} (now {n_sections} sections, {n_words} words)"


# ---------- write_file_verified (atomic write + Python verifier) --------
# Generalizable cure for "wrote a file but the values are wrong" failures
# (m2-style sequence math, any artifact with computable claims). The model
# provides verification code that re-derives or re-checks the artifact's
# claims. If the verifier raises, the write is reverted with diagnostics.
# Works for any scenario where the claim is expressible as Python.

def write_file_verified(path: str, content: str, verifier_code: str,
                        timeout: int = 30) -> str:
    """Write content to path, then run verifier_code in the persistent kernel.

    If the verifier raises (AssertionError, NameError, anything), the file is
    REVERTED to its prior content (or deleted if it was new) and the error
    is returned to the caller. Use this for any artifact where you've made a
    numerical, structural, or behavioral claim that you can re-verify in
    Python — sequence values, algebraic formulas, math identities, file
    structure invariants, etc.

    Example:
      write_file_verified(
        path='/tmp/sequence.md',
        content='a_1=1, a_2=2, a_3=0, ...',
        verifier_code='''
            def a(n):
                if n == 1: return 1
                return a(n-1) + (n-1) * (-1)**n
            assert [a(n) for n in range(1,4)] == [1, 2, 0]
        '''
      )

    The verifier should be self-contained — imports, helper functions,
    assertions. If it executes without error, the write is committed.
    """
    p = Path(path).expanduser().resolve()
    p.parent.mkdir(parents=True, exist_ok=True)
    backup = p.read_text(encoding="utf-8") if p.exists() else None
    p.write_text(content, encoding="utf-8")
    _track_write(str(p))
    # Run verifier in the persistent kernel
    try:
        verifier_result = python_run(verifier_code, timeout=timeout)
    except Exception as e:  # noqa: BLE001
        verifier_result = f"[error] {type(e).__name__}: {e}"
    # Detect failure: any error marker in the kernel output
    failure_markers = ["[error]", "Traceback", "AssertionError", "NameError",
                       "TypeError", "ValueError", "ZeroDivisionError",
                       "IndexError", "KeyError", "AttributeError",
                       "[1 error", "[2 error", "[3 error"]
    failed = any(marker in verifier_result for marker in failure_markers)
    if failed:
        # Revert
        if backup is None:
            try:
                p.unlink()
            except OSError:
                pass
        else:
            p.write_text(backup, encoding="utf-8")
        return (f"[VERIFICATION FAILED] write to {p} REVERTED. "
                f"Your verifier output:\n{verifier_result[:1500]}\n"
                f"Adjust your claim to match what the recurrence/spec actually produces, "
                f"OR adjust the verifier code if it's wrong, then call write_file_verified again.")
    return (f"[VERIFIED] wrote {p} ({len(content)} chars). "
            f"Verifier output:\n{verifier_result[:600]}")


# ---------- apply_patch (unified-diff editing) ---------------------------
# When modifying an existing file, regenerating the whole file is wasteful:
# a 600-line file = ~5K tokens, but a 30-line patch = ~300 tokens. Emitting
# a unified diff and applying it via `git apply` is 5-20× faster on edits.
#
# Format: standard unified diff. Use `--- /dev/null` for new files.
# All-or-nothing: any hunk failure aborts the whole patch.

def apply_patch(patch: str) -> str:
    """Apply a unified diff to one or more files.

    Use this INSTEAD of write_file when modifying an EXISTING file: emit
    only the changed lines as a diff, not the whole file. For new files,
    use write_file (or use `--- /dev/null` here).

    Format (standard unified diff):

        --- a/path/to/file.py
        +++ b/path/to/file.py
        @@ -10,5 +10,7 @@
         context line
        -removed
        +added 1
        +added 2
         context line

    Multiple files in one patch are supported (concatenate `--- ... +++ ...`
    blocks). All-or-nothing: any hunk failure aborts the whole patch.

    Returns success/error message; on success lists touched files + line counts.
    """
    import tempfile, subprocess, re as _re
    if not patch or not patch.strip():
        return "[error] empty patch"

    # Parse out the +++ paths so we can find the real file location and
    # rewrite the patch with paths that git apply can actually find.
    header_pat = _re.compile(r"^(\+\+\+|---) (.+?)(?:\t|$)", _re.MULTILINE)

    def _norm(p: str) -> str:
        p = p.strip()
        if p == "/dev/null":
            return p
        # strip a/ b/ prefix
        if p.startswith("a/") or p.startswith("b/"):
            p = p[2:]
        return p

    def _resolve_target_path(raw: str) -> str | None:
        """Find the actual filesystem path for a +++ header value."""
        if raw == "/dev/null":
            return None
        cwd = os.getcwd()
        candidates = [raw]
        # try with a/ b/ stripped
        if raw.startswith(("a/", "b/")):
            candidates.append(raw[2:])
        # Each candidate: try as absolute, else as cwd-relative.
        for c in candidates:
            if c.startswith("/") and os.path.exists(c):
                return c
            absp = os.path.join(cwd, c)
            if os.path.exists(absp):
                return absp
            # try with leading slash (model-emitted "tmp/foo" → "/tmp/foo")
            if not c.startswith("/"):
                absp2 = "/" + c
                if os.path.exists(absp2):
                    return absp2
        # For new-file diffs, the +++ path won't exist yet — return abspath
        # under cwd as the place to create.
        first = candidates[0]
        return first if first.startswith("/") else os.path.join(cwd, first)

    # Rewrite +++ and --- paths to absolute, detected paths.
    def rewrite(m: _re.Match) -> str:
        marker, raw = m.group(1), m.group(2).strip()
        if raw == "/dev/null":
            return f"{marker} /dev/null"
        resolved = _resolve_target_path(raw)
        if resolved is None:
            return f"{marker} /dev/null"
        # Use a/ b/ prefix per convention so -p1 strips it cleanly.
        prefix = "a" if marker == "---" else "b"
        return f"{marker} {prefix}{resolved}"

    rewritten_patch = header_pat.sub(rewrite, patch)
    text = rewritten_patch if rewritten_patch.endswith("\n") else rewritten_patch + "\n"

    with tempfile.NamedTemporaryFile("w", suffix=".patch", delete=False, encoding="utf-8") as f:
        f.write(text)
        patch_path = f.name
    try:
        # With our absolute-path rewrite, -p1 strips the `a/`/`b/` prefix and
        # gets us back to the real absolute path. Try several flag combos to
        # tolerate the most common LLM diff-construction mistakes:
        #
        #   1) plain                — strict (matches if everything is right)
        #   2) --recount            — fixes wrong line counts in @@ headers
        #                             (very common LLM error)
        #   3) --recount --3way     — also resolves small context drift via
        #                             3-way merge against the source file
        #   4) GNU `patch -F 99`    — most permissive: allows large fuzz
        last_err = ""
        attempts = [
            ["git", "apply", "--unsafe-paths", "--allow-empty",
             "--whitespace=nowarn", "-p1", patch_path],
            ["git", "apply", "--unsafe-paths", "--allow-empty",
             "--whitespace=nowarn", "--recount", "-p1", patch_path],
            ["git", "apply", "--unsafe-paths", "--allow-empty",
             "--whitespace=nowarn", "--recount", "--3way", "-p1", patch_path],
            ["patch", "-p1", "-F", "99", "-f", "-i", patch_path],
        ]
        applied = False
        for cmd in attempts:
            r = subprocess.run(cmd, cwd="/", capture_output=True, text=True, timeout=30)
            if r.returncode == 0:
                applied = True
                break
            last_err = (r.stderr or r.stdout or "")[:600]
        if not applied:
            return (f"[error] patch application failed after 4 attempts (plain, "
                    f"--recount, --recount+--3way, GNU patch -F99).\n"
                    f"Last stderr: {last_err}\n"
                    f"Hint: the @@ header counts may be wrong. Common LLM mistake. "
                    f"Either (a) read_file the affected lines first and re-construct "
                    f"with correct counts, or (b) use edit_file with exact "
                    f"old_string/new_string pairs (less drift-prone).")
        # Track touched files (extract from rewritten +++ headers).
        # The rewrite emits `b<abs_path>` so stripping just the "b" prefix
        # leaves the absolute path with its leading slash.
        touched: list[str] = []
        for line in rewritten_patch.splitlines():
            if line.startswith("+++ "):
                p = line[4:].split("\t", 1)[0].strip()
                if p == "/dev/null":
                    continue
                if p.startswith(("a", "b")) and len(p) > 1 and p[1] == "/":
                    p = p[1:]  # strip just the leading 'a' or 'b', keep '/'
                full = Path(p).resolve() if os.path.isabs(p) else (Path(os.getcwd()) / p).resolve()
                touched.append(str(full))
                _track_write(str(full))
        if not touched:
            # No `+++ <path>` headers in the rewritten patch means we never
            # found a real target. The git/patch invocations may have
            # "succeeded" on garbage input (a no-op), but since nothing was
            # actually changed, surface the issue rather than reporting OK.
            return ("[error] patch had no usable +++ headers — wrong format. "
                    "Use unified diff with `--- a/<path>` and `+++ b/<path>` "
                    "lines (NOT the `*** Begin Patch / *** Update File` style). "
                    "If creating a new file, use `--- /dev/null`. "
                    "For small edits prefer edit_file with old_string/new_string.")
        summary = []
        for fp in touched:
            try:
                lc = sum(1 for _ in open(fp, "rb"))
            except OSError:
                lc = -1
            summary.append(f"{fp}: {lc} lines")
        return "[ok] patch applied. " + " | ".join(summary)
    finally:
        try:
            os.unlink(patch_path)
        except OSError:
            pass


# ---------- scratchpad / ask_user / graph_compose -----------------------
# Three small affordances added so the system prompt's anti-loop /
# anti-guess / graphs-first guidance has actual tool support.
#
# scratchpad: in-task working notes that don't pollute long-term memory.
#   The system prompt steers the model away from using memory_save for
#   loose intermediate thoughts; without scratchpad it had to choose
#   between leaking transient state into permanent memory or carrying
#   the state in context (which then survives compaction noisily).
#   Per-pid file keeps it isolated per session, auto-cleaned on /clear.
#
# ask_user: clarifying question affordance. In an interactive chat the UI
#   layer can intercept this and round-trip a question to the user; in
#   the headless eval harness there's no user to ask, so we return a
#   structured "[no user — proceed]" marker that the model treats the
#   same as a REFUSED cap. This means a confused model can't loop on
#   asking; it either asks ONCE and proceeds, or skips it.
#
# graph_compose: a single call that takes a NL description, runs the
#   graph designer, and optionally runs the resulting graph. Closes the
#   gap where the model knew agent_graph_run existed but couldn't build
#   a new graph mid-task.

_scratchpad_path = f"/tmp/qwen_scratchpad_{os.getpid()}.json"


def scratchpad(action: str = "append", content: str = "",
                key: str = "default") -> str:
    """In-session working notes — write down what you're thinking BEFORE
    a long search, what you found AFTER a complex fetch, etc.

    Use this instead of memory_save when the note is for *this* task only.
    Notes persist for the session but are cleared on /clear (chat) or new
    headless invocation. Multiple named scratchpads via the `key` arg.

    Actions:
      - append (default): add content as a timestamped line to the keyed pad
      - read: return the full contents of the keyed pad (or all pads)
      - clear: erase the keyed pad (or all pads if key='*')
      - list: list known keys with size + line counts
    """
    import datetime as _dt
    try:
        if os.path.exists(_scratchpad_path):
            data = json.loads(open(_scratchpad_path).read() or "{}")
        else:
            data = {}
    except (OSError, ValueError):
        data = {}
    if not isinstance(data, dict):
        data = {}
    action = (action or "append").lower().strip()
    key = (key or "default").strip() or "default"

    if action == "append":
        if not content:
            return "[error] content required for append"
        ts = _dt.datetime.now().strftime("%H:%M:%S")
        entries = data.setdefault(key, [])
        if not isinstance(entries, list):
            entries = []
            data[key] = entries
        entries.append({"ts": ts, "text": content})
        try:
            with open(_scratchpad_path, "w") as f:
                json.dump(data, f)
        except OSError as e:
            return f"[error] could not persist scratchpad: {e}"
        return (f"[scratchpad:{key}] line {len(entries)} added "
                f"({len(content)} chars).")

    if action == "read":
        if key == "*":
            out = []
            for k, lines in data.items():
                if not isinstance(lines, list):
                    continue
                out.append(f"# scratchpad:{k} ({len(lines)} lines)")
                for ln in lines:
                    out.append(f"  [{ln.get('ts','?')}] {ln.get('text','')}")
            return "\n".join(out) or "[scratchpad empty]"
        entries = data.get(key) or []
        if not entries:
            return f"[scratchpad:{key}] empty"
        out = [f"# scratchpad:{key} ({len(entries)} lines)"]
        for ln in entries:
            out.append(f"  [{ln.get('ts','?')}] {ln.get('text','')}")
        return "\n".join(out)

    if action == "clear":
        if key == "*":
            data = {}
        else:
            data.pop(key, None)
        try:
            with open(_scratchpad_path, "w") as f:
                json.dump(data, f)
        except OSError as e:
            return f"[error] could not persist scratchpad: {e}"
        return f"[scratchpad:{key}] cleared."

    if action == "list":
        if not data:
            return "[scratchpad: no keys]"
        out = ["scratchpad keys:"]
        for k, lines in data.items():
            n = len(lines) if isinstance(lines, list) else 0
            out.append(f"  - {k}: {n} line(s)")
        return "\n".join(out)

    return (f"[error] unknown action {action!r}. "
            f"Use append|read|clear|list.")


# Sentinel file the UI / chat layer may set to point at a queued reply.
# Headless mode never sets this; ask_user always falls through to the
# "no user — proceed" branch in headless mode, which is the correct
# behavior in evals where there's no human to ask.
_ask_user_inbox = f"/tmp/qwen_ask_user_inbox_{os.getpid()}.json"


def ask_user(question: str, options: str = "") -> str:
    """Ask the user a clarifying question — use ONLY when the request is
    genuinely ambiguous in ways that change the artifact you produce.

    In headless/eval mode there is no user available; this returns
    "[no-user] proceeding with best inference" and the model is expected
    to make its best guess and continue. The system prompt frames this
    as the equivalent of a REFUSED marker for ambiguity loops.

    In chat mode the UI may intercept this call and round-trip a reply
    via the inbox sentinel; the chat-side wrapper handles that.

    Args:
        question: a short, specific question (1-2 sentences).
        options: optional comma-separated list of suggested choices that
            the UI can render as buttons.
    """
    if not question or len(question.strip()) < 5:
        return "[error] question required (1-2 sentences, ≥5 chars)"
    # Chat-mode hand-off: the UI writes the reply to _ask_user_inbox.
    try:
        if os.path.exists(_ask_user_inbox):
            with open(_ask_user_inbox) as f:
                obj = json.loads(f.read() or "{}")
            if isinstance(obj, dict) and obj.get("question") == question.strip():
                reply = obj.get("reply") or ""
                if reply:
                    try:
                        os.unlink(_ask_user_inbox)
                    except OSError:
                        pass
                    return f"[user] {reply}"
    except (OSError, ValueError):
        pass
    # Default: no user reachable. Tell the model to proceed with best
    # inference. Single-fire per question (same prose every time, so the
    # cache dedup won't suppress it as a near-duplicate question).
    opts_clause = f" (suggested options were: {options})" if options else ""
    return ("[no-user] proceeding with best inference — there is no human "
            "to ask in this session. Your next call must use your best "
            "guess at the user's intent and continue the task." + opts_clause)


def graph_compose(description: str, run: bool = False,
                   inputs: str = "{}") -> str:
    """Design + save a fresh agent_graph from a natural-language description.
    Optionally run it immediately with the provided inputs.

    Use when none of the graphs from `agent_graph_list()` fit but the task
    decomposes cleanly into research → analyze → produce (or similar).
    Saves a new file under examples/<auto-name>_graph.py.

    Args:
        description: one paragraph describing the pipeline's purpose and
            shape (5-15 sentences). Vague descriptions yield generic graphs.
        run: if True, immediately invoke the saved graph with `inputs`.
        inputs: JSON object string for the graph's initial inputs.

    Returns: human-readable status with the new graph's name and (if run)
    a compact run summary.
    """
    if not description or len(description.strip()) < 8:
        return "[error] description required (≥8 chars, ideally a paragraph)"
    proj = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    sys.path.insert(0, os.path.join(proj, "scripts"))
    try:
        from graph_designer import design_and_save  # type: ignore
    except Exception as e:  # noqa: BLE001
        return f"[error] graph_designer unavailable: {type(e).__name__}: {e}"
    try:
        result = design_and_save(description)
    except Exception as e:  # noqa: BLE001
        return f"[error] graph_compose failed: {type(e).__name__}: {e}"
    if not isinstance(result, dict) or not result.get("ok"):
        err = (result or {}).get("error") if isinstance(result, dict) else "?"
        return f"[error] graph design rejected: {err}"
    name = result.get("name") or "(unnamed)"
    path = result.get("path") or "?"
    out = [f"[graph_compose] saved graph {name!r} at {path}."]
    if run:
        run_out = agent_graph_run(graph=name, inputs=inputs)
        out.append(f"\n--- run output ---\n{run_out}")
    else:
        out.append("Call agent_graph_run(graph=%r, inputs=...) to invoke."
                   % name)
    return "\n".join(out)


# ---------- done (explicit completion signal + anti-give-up guard) -------
# Generalizable cure for two failure modes at once:
#  - Agent writes artifact but doesn't /exit → runner times out (s2-style).
#  - Agent gives up after 1 step on a hard prompt (h7-style).
# The tool refuses to mark complete if no artifact was written this session,
# which is enforceable at the tool level (not just system-prompt nudging).
# When accepted, drops a sentinel file that agent.py's main loop watches.

_session_writes_path = f"/tmp/qwen_session_writes_{os.getpid()}.json"
_done_sentinel_path = f"/tmp/qwen_session_done_{os.getpid()}.txt"


def _track_write(path: str) -> None:
    """Bump the write counter for the current session. Called from
    write_file / edit_file / append_finding / notebook_edit on success."""
    try:
        if os.path.exists(_session_writes_path):
            data = json.loads(open(_session_writes_path).read() or "{}")
        else:
            data = {}
        data[path] = data.get(path, 0) + 1
        with open(_session_writes_path, "w") as f:
            json.dump(data, f)
    except (OSError, ValueError):
        pass


def _session_writes() -> dict[str, int]:
    if not os.path.exists(_session_writes_path):
        return {}
    try:
        return json.loads(open(_session_writes_path).read() or "{}")
    except (OSError, ValueError):
        return {}


def done(summary: str) -> str:
    """Signal the task is complete. Use as your FINAL tool call.

    The harness reads this signal and gracefully closes the session — no
    more LLM rounds, no waiting for /exit. Pass a 1-2 sentence summary
    naming the deliverable (artifact path + what's in it).

    SAFETY: this tool refuses to accept completion if no artifact files
    have been written this session. If you see "no writes recorded" you
    haven't actually done the task yet — write the requested artifact
    first, then call done.
    """
    if not summary or len(summary.strip()) < 5:
        return "[error] summary required (1-2 sentences)"
    writes = _session_writes()
    if not writes:
        return ("[refused] no writes recorded this session. The task is not "
                "complete until you've written the requested artifact. Use "
                "write_file / edit_file / append_finding / notebook_edit / "
                "python_run (with file output) first, then call done() again.")
    try:
        with open(_done_sentinel_path, "w") as f:
            f.write(summary)
    except OSError as e:
        return f"[error] could not write done sentinel: {e}"
    n_files = len(writes)
    n_total = sum(writes.values())
    return (f"DONE accepted. Session wrote {n_files} file(s), {n_total} write op(s). "
            f"Summary: {summary!r}. Harness will close this session shortly.")


# ---------- shell ----------------------------------------------------------

def bash(command: str, timeout: int = 60) -> str:
    """Run a shell command in the current working directory."""
    try:
        result = subprocess.run(
            command,
            shell=True,
            capture_output=True,
            text=True,
            timeout=timeout,
            cwd=os.getcwd(),
        )
    except subprocess.TimeoutExpired:
        return f"[timeout after {timeout}s]"
    parts = []
    if result.stdout:
        parts.append(result.stdout.rstrip())
    if result.stderr:
        parts.append("[stderr]\n" + result.stderr.rstrip())
    if result.returncode != 0:
        parts.append(f"[exit {result.returncode}]")
    out = "\n".join(parts) or "(no output)"
    if len(out) > 80000:
        out = out[:80000] + f"\n…[truncated {len(out) - 80000} chars]"
    return out


# ---------- explore (subagent) ---------------------------------------------

EXPLORE_SYSTEM_PROMPT = """\
Read-only research subagent. Answer ONE question from filesystem + web, \
return a concise evidence-backed report. The parent dispatched you so the \
search noise stays out of its context — return signal only.

# Budget
5-10 steps typical, draft by step 12. Past that wastes the parent's time.

# Principles
1. Hypothesis first — guess the answer in one line before tools, then \
confirm or refute. (Skip for trivial lookups.)
2. Breadth before depth — list_files / structural grep / READMEs first; \
read only the 1-3 most promising files.
3. Parallelize independent calls (greps for different terms, reads of \
known-relevant files) in one turn. Sequential only when each depends on \
the prior result.
4. Triangulate non-trivial claims with two independent pieces of evidence \
(definition + caller, two corroborating docs). Flag single-source claims.
5. Cite file:line, function name, or URL on every factual claim. No "I \
think the code does X" — show the line or drop the claim.
6. Tag uncertainty: Confirmed / Likely / Unclear. Don't fake certainty.
7. Stop when you can answer — the parent is blocked. Concise + correct \
beats exhaustive.
8. `(no matches)` is a confirmed negative; `[cached…]` means you already \
asked. Don't retry case variants or synonyms — revise the hypothesis.

# Tools
list_files, grep (start `output_mode="files_with_matches"`, then "content" \
on narrowed set), read_file (slice with offset/limit), web_search, \
web_fetch. Parallel where independent.

# Cannot use
write_file, edit_file, bash, explore. Read-only — no shell, no side \
effects, no further subagents. If the question genuinely needs that, say \
so and let the parent decide.

# Output (≤250 words; no preamble, no meta-commentary, no further tool calls)
ANSWER: 1-3 sentences. Direct.
EVIDENCE:
  - <file:line or URL> — what it shows / what you concluded
  - ...
CAVEATS:
  - Single-source claims, ambiguities, things you couldn't verify.

Out of budget? Return what you have with Confirmed/Likely/Unverified \
tags. Partial > silence."""

EXPLORE_TOOL_NAMES = {"list_files", "grep", "read_file", "web_search", "web_fetch"}


def _llm_endpoint() -> str:
    host = os.environ.get("QWEN_HOST", "127.0.0.1")
    if host in ("0.0.0.0", ""):
        host = "127.0.0.1"
    port = os.environ.get("QWEN_PORT", "8000")
    return f"http://{host}:{port}/v1/chat/completions"


# Cached model id — the proxy rejects requests with the alias "qwen3.6"
# now that it pins to the loaded model path. We resolve the real id from
# /v1/models on first successful fetch and cache forever; under load the
# /v1/models endpoint can queue behind in-flight chats (single-stream
# upstream), so we keep retrying with backoff rather than locking in
# the alias on the first failure.
_RESOLVED_MODEL_ID: str | None = None
_RESOLVE_LOCK = threading.Lock()


def _resolve_model_id() -> str:
    """Return the actual model id the upstream is serving.

    The dflash proxy enforces an exact match against the loaded model path
    (e.g. './models/Qwen3.6-35B-A3B-OptiQ-4bit'); the friendly alias
    'qwen3.6' returns 400. Tries up to 30 seconds with a backoff so a
    cold-load proxy still resolves cleanly. Falls back to the env alias
    if upstream is genuinely unreachable.
    """
    global _RESOLVED_MODEL_ID
    if _RESOLVED_MODEL_ID:
        return _RESOLVED_MODEL_ID
    # If QWEN_MODEL_NAME already looks like a real model id (path-like or
    # contains "models/"), trust it without round-tripping to upstream.
    # qwen_ui sets this at startup once it has resolved against /v1/models.
    env_id = os.environ.get("QWEN_MODEL_NAME", "")
    if env_id and (env_id.startswith(("./", "/")) or "models/" in env_id):
        _RESOLVED_MODEL_ID = env_id
        return env_id
    with _RESOLVE_LOCK:
        if _RESOLVED_MODEL_ID:  # double-check under lock
            return _RESOLVED_MODEL_ID
        fallback = os.environ.get("QWEN_MODEL_NAME", "qwen3.6")
        host = os.environ.get("QWEN_HOST", "127.0.0.1")
        if host in ("0.0.0.0", ""):
            host = "127.0.0.1"
        port = os.environ.get("QWEN_PORT", "8000")
        url = f"http://{host}:{port}/v1/models"
        # Retries: 4s, 8s — total ~12s of patience before falling back.
        # /v1/models can queue behind a chat that's mid-stream because
        # dflash is single-stream; one retry past the typical chat wall
        # time usually gets through.
        for timeout_s in (4, 8, 16):
            try:
                with urllib.request.urlopen(url, timeout=timeout_s) as r:
                    data = json.loads(r.read())
            except Exception:  # noqa: BLE001
                continue
            items = data.get("data") or []
            if items and isinstance(items[0], dict):
                mid = items[0].get("id")
                if isinstance(mid, str) and mid:
                    _RESOLVED_MODEL_ID = mid
                    return mid
            break  # got a response but malformed — fall through to fallback
        return fallback


def _post_chat_once(messages: list[dict], tools: list[dict] | None = None) -> dict:
    """Single POST attempt to the local LLM endpoint."""
    payload: dict[str, Any] = {
        "model": _resolve_model_id(),
        "messages": messages,
        "stream": False,
    }
    if tools:
        payload["tools"] = tools
        payload["tool_choice"] = "auto"
    req = urllib.request.Request(
        _llm_endpoint(),
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=600) as resp:
        return json.loads(resp.read())


def _post_chat(messages: list[dict], tools: list[dict] | None = None,
               retries: int = 3) -> dict:
    """POST with retry-with-backoff for transient 5xx + URL errors. Used by
    the explore subagent and maybe_compact's summarization. Same resilience
    pattern as agent.py's post_chat — keeps internal LLM calls robust against
    server hiccups."""
    import time
    import urllib.error as _urlerr
    last_err: Exception | None = None
    for attempt in range(retries + 1):
        try:
            return _post_chat_once(messages, tools)
        except _urlerr.HTTPError as e:
            last_err = e
            transient = e.code in (500, 502, 503, 504)
            if not transient or attempt == retries:
                raise
            time.sleep(2 ** attempt)
        except _urlerr.URLError as e:
            last_err = e
            if attempt == retries:
                raise
            time.sleep(2 ** attempt)
    if last_err is not None:
        raise last_err
    raise RuntimeError("_post_chat: exhausted retries with no specific error")


# ---------- context compaction ---------------------------------------------

# Main agent: compact at 60k. The local mlx-openai-server has a hardcoded
# 300s internal asyncio timeout on prompt processing — at very large prompts
# the model handler can stall and the server returns 500. 60k keeps prompts
# comfortably below that danger zone. Override with QWEN_AGENT_COMPACT_AT
# if you have a beefier server / different runtime.
COMPACT_AT_TOKENS = int(os.environ.get("QWEN_AGENT_COMPACT_AT", "60000"))
# Explore subagent: should be tight — its job is to return a summary, not
# accumulate context. Compact at 25k so it self-summarizes much earlier.
EXPLORE_COMPACT_AT_TOKENS = int(os.environ.get("QWEN_EXPLORE_COMPACT_AT", "25000"))
# Hard step ceiling for explore. 25 covers list_files + ~5 greps + ~5 reads
# + answer with room to spare. Investigations longer than this should be
# decomposed into multiple narrower explore() calls.
EXPLORE_MAX_STEPS = int(os.environ.get("QWEN_EXPLORE_MAX_STEPS", "25"))


def approx_tokens(messages: list[dict]) -> int:
    """Cheap char/4 token estimate. Used as a fallback when the real
    tokenizer isn't available. Prefer `real_tokens()` for any decision
    that affects user-visible state (compaction triggers, telemetry,
    status bars) — char/4 undercounts JSON / code-heavy chats by ~30%.
    """
    total = 0
    for m in messages:
        c = m.get("content")
        if isinstance(c, str):
            total += len(c) // 4
        elif isinstance(c, list):
            total += sum(len(str(x)) for x in c) // 4
        for tc in m.get("tool_calls") or []:
            total += len(json.dumps(tc, ensure_ascii=False)) // 4
    return total


# ---------- real Qwen tokenizer (shared across CLI / UI / graph) ----------
# Lifted from qwen_ui's _count_chat_tokens so every caller (compaction
# decisions in agent.py, agent_graph.py, status bars in qwen_ui) uses the
# SAME count and they all match. Char/4 underestimates JSON/code-heavy
# chats by ~30%, which is why `_maybe_compact_chat` was failing to trip
# at the UI's reported 60k threshold.

_TOKENIZER_STATE: dict = {"obj": None, "loaded": False, "name": None}
_TOKENIZER_LOCK = threading.Lock()


def _load_tokenizer():
    """Lazy-load the qwen tokenizer. Returns None on failure (callers must
    handle that — `real_tokens` falls back to `approx_tokens`)."""
    if _TOKENIZER_STATE["loaded"]:
        return _TOKENIZER_STATE["obj"]
    with _TOKENIZER_LOCK:
        if _TOKENIZER_STATE["loaded"]:
            return _TOKENIZER_STATE["obj"]
        _TOKENIZER_STATE["loaded"] = True
        # Heuristic: env override → loaded model dir → typical default.
        candidates = [
            os.environ.get("QWEN_TOKENIZER_DIR"),
            os.environ.get("QWEN_MODEL_DIR"),
            os.environ.get("QWEN_MODEL_NAME"),
            "./models/Qwen3.6-35B-A3B-OptiQ-4bit",
        ]
        seen: set[str] = set()
        for name in candidates:
            if not name or name in seen:
                continue
            seen.add(name)
            try:
                from transformers import AutoTokenizer
                tok = AutoTokenizer.from_pretrained(name, trust_remote_code=True)
                if tok is not None:
                    _TOKENIZER_STATE["obj"] = tok
                    _TOKENIZER_STATE["name"] = name
                    return tok
            except Exception:  # noqa: BLE001
                continue
        return None


def _normalize_messages_for_tokenizer(messages: list[dict]) -> list[dict]:
    """Qwen3.6's chat template is strict:
      - tool_calls[i].function.arguments must be a dict (NOT a JSON string).
      - assistant `content` must be a non-None string.
    Normalize both before passing to the template."""
    out: list[dict] = []
    for m in messages:
        m2 = dict(m)
        if m2.get("role") == "assistant" and m2.get("content") is None:
            m2["content"] = ""
        tcs = m2.get("tool_calls")
        if isinstance(tcs, list):
            new_tcs = []
            for tc in tcs:
                if not isinstance(tc, dict):
                    continue
                fn = (tc.get("function") or {}).copy()
                args = fn.get("arguments")
                if isinstance(args, str):
                    try:
                        fn["arguments"] = json.loads(args) if args.strip() else {}
                    except json.JSONDecodeError:
                        fn["arguments"] = {"_raw": args}
                tc2 = dict(tc)
                tc2["function"] = fn
                new_tcs.append(tc2)
            m2["tool_calls"] = new_tcs
        out.append(m2)
    return out


def real_tokens(messages: list[dict], tools: list[dict] | None = None) -> int:
    """Real Qwen token count via chat template. Same algorithm dflash uses.
    Falls back to `approx_tokens` if the tokenizer isn't available so this
    function is always safe to call. ~10ms per call on typical chats —
    cheap enough to use anywhere `approx_tokens` was previously used."""
    tok = _load_tokenizer()
    if tok is None:
        return approx_tokens(messages)
    try:
        msgs = _normalize_messages_for_tokenizer(messages)
        kwargs: dict = {"tokenize": False, "add_generation_prompt": True}
        if tools:
            try:
                kwargs["tools"] = tools
                text = tok.apply_chat_template(msgs, **kwargs)
            except (TypeError, ValueError):
                kwargs.pop("tools", None)
                text = tok.apply_chat_template(msgs, **kwargs)
        else:
            text = tok.apply_chat_template(msgs, **kwargs)
        return len(tok(text)["input_ids"])
    except Exception:  # noqa: BLE001
        return approx_tokens(messages)


def message_content_tokens(message: dict) -> int:
    """Token count for a SINGLE message's content, ignoring chat-template
    overhead (role markers, im_start/im_end, etc.). Use this for size
    triage decisions ("is this user message a file upload?") where the
    ~10-20 token template overhead is irrelevant compared to the content.

    `real_tokens([single_message])` doesn't work for this purpose — Qwen's
    template apparently produces near-empty output for one-message
    conversations and returns ~1 token regardless of content size. This
    helper bypasses the template and tokenizes the content string alone.
    Falls back to char/4 if the tokenizer isn't available.
    """
    content = message.get("content")
    if content is None:
        return 0
    # Multimodal content: sum text parts, ignore image parts (their token
    # cost is computed by the model from image dimensions, not text).
    if isinstance(content, list):
        text_parts: list[str] = []
        for part in content:
            if isinstance(part, dict) and part.get("type") == "text":
                t = part.get("text", "")
                if isinstance(t, str):
                    text_parts.append(t)
            elif isinstance(part, str):
                text_parts.append(part)
        content_str = "\n".join(text_parts)
    elif isinstance(content, str):
        content_str = content
    else:
        content_str = str(content)
    if not content_str:
        return 0
    tok = _load_tokenizer()
    if tok is None:
        return len(content_str) // 4
    try:
        return len(tok(content_str)["input_ids"])
    except Exception:  # noqa: BLE001
        return len(content_str) // 4


_COMPACT_REQUEST = (
    "Compact the conversation above into AGFMT — the next turn reads these "
    "sections, not prose:\n\n"
    "@question:t\n<original question, 1 sentence>\n"
    "@findings:l\n<one per line — file:line, URL, number, name>\n"
    "@decisions:l\n<one per line — what tried, what kept/rejected, why>\n"
    "@open:l\n<unresolved sub-questions>\n"
    "@next:t\n<the very next concrete step>\n"
    "@END\n\n"
    "Rules: be specific (citations beat adjectives), <500 words total, no "
    "tools, skip empty sections — don't pad."
)


def maybe_compact(messages: list[dict], threshold: int | None = None) -> list[dict] | None:
    """Return a compacted message list if over threshold, else None.

    Strategy: keep the system prompt + first user message; replace everything
    in between with a single user message containing a summary. Tool-call /
    tool-result pairings disappear cleanly because the entire middle is
    replaced — no dangling references.

    threshold defaults to COMPACT_AT_TOKENS (main agent). Explore passes
    EXPLORE_COMPACT_AT_TOKENS so subagents self-compact much earlier.
    """
    limit = threshold if threshold is not None else COMPACT_AT_TOKENS
    # Use the real tokenizer so the trigger matches what the UI status bar
    # reports — char/4 underestimates JSON-heavy chats by ~30% and was
    # missing the threshold by a wide margin.
    tokens = real_tokens(messages)
    if tokens < limit or len(messages) < 4:
        return None

    head = messages[:2]
    summary_request = list(messages) + [{"role": "user", "content": _COMPACT_REQUEST}]
    try:
        resp = _post_chat(summary_request, tools=None)
        summary = (resp["choices"][0]["message"].get("content") or "").strip()
    except Exception as e:  # noqa: BLE001
        return None  # leave messages alone if summarization fails
    if not summary:
        return None

    compacted = list(head) + [
        {
            "role": "user",
            "content": (
                "[Earlier in this conversation a long investigation took place. "
                "Here is the compacted state of the world:]\n\n"
                f"{summary}\n\n"
                "[Now continue from this state. Take the next step needed to "
                "answer the original question.]"
            ),
        }
    ]
    return compacted


def explore(question: str, max_steps: int = EXPLORE_MAX_STEPS) -> str:
    """Run a read-only subagent in an isolated context and return its summary.

    Use this for broad codebase or research questions ("how does X work?",
    "where is Y handled?") so the search noise stays out of the main context.
    Returns only the subagent's final message.

    Tight by design: default max_steps=25, self-compacts at 40k tokens, and
    runs a per-invocation tool-call cache so identical greps/reads inside one
    explore are deduped. Past the halfway mark we inject a "commit now" nudge.
    On hard ceiling we ask the model for a best-effort answer from collected
    evidence rather than returning a "stuck" sentinel.
    """
    disabled = _disabled_tools()
    sub_tools = [t for t in TOOLS if t["function"]["name"] in EXPLORE_TOOL_NAMES
                 and t["function"]["name"] not in disabled]
    messages: list[dict] = [
        {"role": "system",
         "content": (f"{EXPLORE_SYSTEM_PROMPT}\n\n"
                     f"Working directory: {os.getcwd()}\n"
                     f"Step budget: {max_steps}.")},
        {"role": "user", "content": question},
    ]

    # Per-invocation tool-call cache. Keys are (fn, sorted-json args). Stops
    # the subagent from re-running the same grep/read within one exploration.
    run_cache: dict[tuple, str] = {}
    halftime_warned = False
    halftime_step = max(1, max_steps // 2)

    def cached_subcall(fn: str, args: dict) -> str:
        key = (fn, json.dumps(args, sort_keys=True, default=str))
        if key in run_cache:
            return ("[cached: this exact call returned the same result "
                    "earlier in this exploration. Change approach instead "
                    "of repeating.]\n\n" + run_cache[key])
        result = str(DISPATCH[fn](**args))
        run_cache[key] = result
        return result

    for step_idx in range(max_steps):
        # Self-compact at the explore-specific (lower) threshold so the
        # subagent stays lean. Main agent's threshold is reserved for the
        # main loop where 200k of accumulated context is sometimes warranted.
        compacted = maybe_compact(messages, threshold=EXPLORE_COMPACT_AT_TOKENS)
        if compacted is not None:
            messages = compacted

        # Halftime nudge: once you've burned half the budget, switch from
        # "investigate" mode to "commit" mode. Only fires once.
        if not halftime_warned and step_idx >= halftime_step:
            halftime_warned = True
            messages.append({
                "role": "user",
                "content": (
                    f"[BUDGET CHECK: you've used {step_idx}/{max_steps} steps. "
                    "If you have enough evidence, write the OUTPUT FORMAT "
                    "answer NOW and stop. Otherwise narrow your remaining "
                    "investigation to the SINGLE most important open question "
                    "and finalize within the remaining budget.]"
                )
            })

        try:
            resp = _post_chat(messages, sub_tools)
        except Exception as e:  # noqa: BLE001
            return f"[explore error] {type(e).__name__}: {e}"
        msg = resp["choices"][0]["message"]
        messages.append(msg)
        tcs = msg.get("tool_calls") or []
        if not tcs:
            return (msg.get("content") or "").strip() or "(no findings)"
        for tc in tcs:
            fn = tc["function"]["name"]
            if fn not in EXPLORE_TOOL_NAMES:
                result = f"[denied] explore subagent cannot use '{fn}'"
            else:
                try:
                    args = json.loads(tc["function"].get("arguments", "{}") or "{}")
                except json.JSONDecodeError:
                    args = {}
                try:
                    result = cached_subcall(fn, args)
                except Exception as e:  # noqa: BLE001
                    result = f"[tool error] {type(e).__name__}: {e}"
            if len(result) > 8000:
                result = result[:8000] + f"\n…[truncated {len(result) - 8000} chars]"
            messages.append(
                {
                    "role": "tool",
                    "tool_call_id": tc.get("id", ""),
                    "name": fn,
                    "content": result,
                }
            )

    # Step ceiling reached. Don't return a stuck sentinel — ask the model to
    # synthesize whatever it has into the OUTPUT FORMAT, with no tools so it
    # can't keep investigating. Partial information beats "[stuck]".
    messages.append({
        "role": "user",
        "content": (
            f"[BUDGET EXHAUSTED at {max_steps} steps. STOP investigating. "
            "Write your best answer NOW using the OUTPUT FORMAT and the "
            "evidence already gathered. Tag each claim Confirmed / Likely / "
            "Unverified. Do NOT request more tools.]"
        )
    })
    try:
        final_resp = _post_chat(messages, tools=None)  # no tools — force text
        final = (final_resp["choices"][0]["message"].get("content") or "").strip()
        if final:
            return ("[explore truncated at step ceiling — best answer with "
                    "collected evidence:]\n\n" + final)
    except Exception:  # noqa: BLE001
        pass
    return (f"[explore exhausted {max_steps}-step budget without a final "
            "answer. Try a narrower follow-up question.]")


# ---------- subagent_implement (write-capable subagent) --------------------

SUBAGENT_IMPLEMENT_SYSTEM_PROMPT = """\
Implementation subagent. Implement ONE code task and return a concise \
summary. The parent dispatched you so the read-write-test cycles stay out \
of its context.

# Budget
3-8 calls typical. By call 12 you must have called `done()`.

# Tools
read_file, list_files, grep, write_file, edit_file, apply_patch, \
write_file_verified, python_run, bash, test_run, done.

Prefer `apply_patch` over `write_file` for edits (5-20× faster). Read the \
affected lines first to confirm exact context before patching.

# Principles
1. Plan once — read the target, understand the change, then act.
2. First-draft fast — a 30%-correct edit beats no edit.
3. Validate non-trivial changes with python_run / bash / test_run.
4. Stop at first success.

# Output
`done(summary)` — name files touched and what changed. Not the journey."""


SUBAGENT_IMPLEMENT_TOOL_NAMES = {
    "read_file", "list_files", "grep",
    "write_file", "edit_file", "apply_patch", "write_file_verified",
    "append_finding",
    "python_run", "python_reset", "bash", "test_run",
    "done",
}

SUBAGENT_IMPLEMENT_MAX_STEPS = int(os.environ.get("QWEN_SUBAGENT_MAX_STEPS", "20"))
SUBAGENT_IMPLEMENT_COMPACT_AT = int(os.environ.get("QWEN_SUBAGENT_COMPACT_AT", "20000"))


def subagent_implement(task: str, files: str = "", max_steps: int = SUBAGENT_IMPLEMENT_MAX_STEPS) -> str:
    """Run a write-capable subagent in an isolated context for a focused
    code-gen / edit task. Returns the subagent's final done(summary) message.

    Use this when the parent agent's context would otherwise bloat from a
    multi-turn read-edit-test cycle. The subagent has access to write/edit/
    run tools but operates in its own conversation — only the final summary
    enters the parent's context.

    `task` is a self-contained imperative description of what to do.
    `files` is an optional list of relevant file paths (newline or comma
    separated) to bias the subagent's first reads.
    """
    disabled = _disabled_tools()
    sub_tools = [t for t in TOOLS if t["function"]["name"] in SUBAGENT_IMPLEMENT_TOOL_NAMES
                 and t["function"]["name"] not in disabled]

    user_content = task
    if files.strip():
        user_content = (f"{task}\n\nRelevant files (read these first if "
                        f"helpful):\n{files}")

    messages: list[dict] = [
        {"role": "system",
         "content": (f"{SUBAGENT_IMPLEMENT_SYSTEM_PROMPT}\n\n"
                     f"Working directory: {os.getcwd()}\n"
                     f"Step budget: {max_steps}.")},
        {"role": "user", "content": user_content},
    ]

    run_cache: dict[tuple, str] = {}

    def cached_subcall(fn: str, args: dict) -> str:
        key = (fn, json.dumps(args, sort_keys=True, default=str))
        if key in run_cache:
            return ("[cached: this exact call returned the same result "
                    "earlier in this subagent run. Change approach.]\n\n"
                    + run_cache[key])
        result = str(DISPATCH[fn](**args))
        run_cache[key] = result
        return result

    final_summary: str | None = None
    for step_idx in range(max_steps):
        compacted = maybe_compact(messages, threshold=SUBAGENT_IMPLEMENT_COMPACT_AT)
        if compacted is not None:
            messages = compacted

        try:
            resp = _post_chat(messages, sub_tools)
        except Exception as e:  # noqa: BLE001
            return f"[subagent_implement error] {type(e).__name__}: {e}"
        msg = resp["choices"][0]["message"]
        messages.append(msg)
        tcs = msg.get("tool_calls") or []
        if not tcs:
            content = (msg.get("content") or "").strip()
            return content or "(no summary)"

        for tc in tcs:
            fn = tc["function"]["name"]
            if fn not in SUBAGENT_IMPLEMENT_TOOL_NAMES:
                result = f"[denied] subagent cannot use '{fn}'"
            else:
                try:
                    args = json.loads(tc["function"].get("arguments", "{}") or "{}")
                except json.JSONDecodeError:
                    args = {}
                try:
                    result = cached_subcall(fn, args)
                except Exception as e:  # noqa: BLE001
                    result = f"[tool error] {type(e).__name__}: {e}"
            messages.append(
                {
                    "role": "tool",
                    "tool_call_id": tc.get("id", ""),
                    "name": fn,
                    "content": str(result),
                }
            )
            # If subagent called `done` and it was accepted, capture the
            # summary and return immediately.
            if fn == "done" and isinstance(result, str) and result.startswith("DONE accepted"):
                try:
                    summary_arg = args.get("summary", "")
                except Exception:
                    summary_arg = ""
                final_summary = summary_arg or "(done — no summary)"
                return f"[subagent_implement done] {final_summary}"

    # Step ceiling: ask for a final summary with no tools.
    messages.append({
        "role": "user",
        "content": (
            f"[BUDGET EXHAUSTED at {max_steps} steps. Summarize what you did "
            "(files touched + outcome) in 1-3 sentences. Do NOT request more "
            "tools.]"
        )
    })
    try:
        final_resp = _post_chat(messages, tools=None)
        final = (final_resp["choices"][0]["message"].get("content") or "").strip()
        if final:
            return f"[subagent_implement truncated at step ceiling] {final}"
    except Exception:  # noqa: BLE001
        pass
    return (f"[subagent_implement exhausted {max_steps}-step budget without "
            "calling done. The parent agent should re-decompose the task.]")


# ---------- sci-agent skills: arxiv, doi, github, csv, datetime ----------
# These are the tools the model would routinely encounter in real research
# workflows. They sit at a higher level than web_fetch — each one knows the
# shape of the data it returns, so the model gets clean structured output
# rather than a 200KB HTML dump it has to parse.

def _arxiv_id_from_input(s) -> str | None:
    """Extract a normalized arxiv id (e.g. '2401.12345' or 'cs/0301002')
    from any of: bare id, full URL, abs/PDF link, or title-search style.

    Coerces non-string input (e.g. a JSON number like 2401.12345 that the
    model emitted unquoted) to its string form before parsing — without
    this, `s.strip()` blows up with "'float' object has no attribute
    'strip'" and the user sees a tool error instead of a successful
    lookup. arXiv IDs always have a decimal so the JSON parser silently
    turns them into floats."""
    if s is None or s == "":
        return None
    if not isinstance(s, str):
        s = str(s)
    s = s.strip()
    # Common URL forms.
    m = re.search(r"arxiv\.org/(?:abs|pdf|html)/([\w./\-]+?)(?:v\d+)?(?:\.pdf)?(?:[#?]|$)", s)
    if m:
        return m.group(1)
    # Bare new-style id (yymm.nnnnn)
    m = re.fullmatch(r"\d{4}\.\d{4,5}(?:v\d+)?", s)
    if m:
        return s.split("v")[0] if "v" in s else s
    # Old-style cs/xxx.
    m = re.fullmatch(r"[a-z\-]+(?:\.[A-Z]{2})?/\d{7}(?:v\d+)?", s)
    if m:
        return s.split("v")[0] if "v" in s else s
    return None


def arxiv_search(query, max_results: int = 5,
                 sort_by: str = "relevance") -> str:
    """Search arXiv with the official Atom API.

    Use this instead of `web_search query="... arxiv"` whenever you want
    actual paper metadata (id, abstract, authors, categories) rather than
    Google-style result snippets. Much cleaner output and no scraping
    artifacts.

    sort_by ∈ {"relevance", "submittedDate", "lastUpdatedDate"}.
    """
    from urllib.parse import quote_plus
    # Coerce non-string input (e.g. an arxiv id JSON-parsed as a float
    # like 2401.12345) — without this, quote_plus() raises
    # "TypeError: quote_from_bytes() expected bytes" and the tool call
    # surfaces as a "[tool error]" to the model.
    if not isinstance(query, str):
        query = str(query) if query is not None else ""
    sort_by = sort_by if sort_by in ("relevance", "submittedDate",
                                      "lastUpdatedDate") else "relevance"
    url = (f"http://export.arxiv.org/api/query?"
           f"search_query=all:{quote_plus(query)}&"
           f"start=0&max_results={int(max(1, min(max_results, 20)))}&"
           f"sortBy={sort_by}&sortOrder=descending")
    try:
        with urllib.request.urlopen(url, timeout=15) as r:
            xml = r.read().decode("utf-8", errors="replace")
    except Exception as e:  # noqa: BLE001
        return f"[arxiv api error] {type(e).__name__}: {e}"
    try:
        from xml.etree import ElementTree as ET
        ns = {"a": "http://www.w3.org/2005/Atom",
              "ax": "http://arxiv.org/schemas/atom"}
        root = ET.fromstring(xml)
    except Exception as e:  # noqa: BLE001
        return f"[arxiv parse error] {e}"
    entries = root.findall("a:entry", ns)
    if not entries:
        return "(no arxiv results)"
    out = []
    for i, e in enumerate(entries, 1):
        link = e.findtext("a:id", default="", namespaces=ns)
        aid = _arxiv_id_from_input(link) or link.rsplit("/", 1)[-1]
        title = " ".join((e.findtext("a:title", default="", namespaces=ns) or "").split())
        published = (e.findtext("a:published", default="", namespaces=ns) or "")[:10]
        authors = []
        for au in e.findall("a:author", ns):
            n = au.findtext("a:name", default="", namespaces=ns) or ""
            if n:
                authors.append(n)
        primary = e.find("ax:primary_category", ns)
        category = primary.attrib.get("term", "") if primary is not None else ""
        summary = " ".join((e.findtext("a:summary", default="", namespaces=ns) or "").split())
        if len(summary) > 400:
            summary = summary[:400] + "…"
        author_str = ", ".join(authors[:5]) + (f" (+{len(authors)-5} more)"
                                                if len(authors) > 5 else "")
        out.append(f"{i}. {title}\n"
                   f"   id: {aid}  ·  {published}  ·  {category}\n"
                   f"   {author_str}\n"
                   f"   {summary}")
    return "\n\n".join(out)


def arxiv_fetch(id_or_url, what: str = "abstract",
                max_chars: int = 60000) -> str:
    """Fetch an arXiv paper by id or URL.

    `what` selects the format:
      - "abstract" — title, authors, abstract, categories (cheap, ~1 KB)
      - "html"     — arxiv's HTML5 rendering of the paper (best for sections)
      - "pdf"      — extract text from the PDF (slowest, but works for any
                     paper without an HTML rendering)
    """
    # `_arxiv_id_from_input` now coerces non-string inputs internally, but
    # also accept str/repr explicitly here so the error message below
    # shows the user-friendly form rather than e.g. `2401.12345 (float)`.
    if id_or_url is not None and not isinstance(id_or_url, str):
        id_or_url = str(id_or_url)
    aid = _arxiv_id_from_input(id_or_url)
    if not aid:
        return f"[arxiv] could not parse id from {id_or_url!r}"
    if what == "abstract":
        # The Atom API again, but with id_list for a precise lookup.
        url = f"http://export.arxiv.org/api/query?id_list={aid}"
        try:
            with urllib.request.urlopen(url, timeout=15) as r:
                xml = r.read().decode("utf-8", errors="replace")
        except Exception as e:  # noqa: BLE001
            return f"[arxiv api error] {type(e).__name__}: {e}"
        try:
            from xml.etree import ElementTree as ET
            ns = {"a": "http://www.w3.org/2005/Atom",
                  "ax": "http://arxiv.org/schemas/atom"}
            root = ET.fromstring(xml)
            e = root.find("a:entry", ns)
        except Exception as ex:  # noqa: BLE001
            return f"[arxiv parse error] {ex}"
        if e is None:
            return f"(no arxiv entry for {aid})"
        title = " ".join((e.findtext("a:title", default="", namespaces=ns) or "").split())
        authors = [au.findtext("a:name", default="", namespaces=ns) or ""
                   for au in e.findall("a:author", ns)]
        primary = e.find("ax:primary_category", ns)
        category = primary.attrib.get("term", "") if primary is not None else ""
        summary = (e.findtext("a:summary", default="", namespaces=ns) or "").strip()
        published = (e.findtext("a:published", default="", namespaces=ns) or "")[:10]
        return (f"# {title}\n\n"
                f"**arxiv:** {aid}  ·  **category:** {category}  ·  "
                f"**published:** {published}\n"
                f"**authors:** {', '.join(authors)}\n\n"
                f"## Abstract\n\n{summary}")
    if what == "html":
        url = f"https://arxiv.org/html/{aid}"
        return web_fetch(url, max_chars=max_chars)
    if what == "pdf":
        url = f"https://arxiv.org/pdf/{aid}.pdf"
        return pdf_extract(url, max_chars=max_chars)
    return f"[arxiv] unknown `what`: {what!r} (use abstract|html|pdf)"


def pdf_extract(path_or_url: str, pages: str = "",
                max_chars: int = 80000) -> str:
    """Extract text from a PDF — local path or HTTPS URL.

    `pages` is a 1-indexed range like "1-3" or "5,7,9-12" or "" for all
    pages. Useful for large papers where you only want the abstract+intro.

    Backed by pypdf (which is already in the venv); falls back to the
    URL fetch path for non-pypdf-compatible PDFs by surfacing the error
    so the model can swap to a different approach.
    """
    try:
        import pypdf  # type: ignore
    except ImportError:
        return "[error] pypdf not installed in venv"
    src = path_or_url.strip()
    if src.startswith(("http://", "https://")):
        try:
            import httpx
            r = httpx.get(src, timeout=30, follow_redirects=True,
                          headers={"User-Agent": "qwen-agent"})
            r.raise_for_status()
        except Exception as e:  # noqa: BLE001
            return f"[pdf fetch failed] {type(e).__name__}: {e}"
        import io
        try:
            reader = pypdf.PdfReader(io.BytesIO(r.content))
        except Exception as e:  # noqa: BLE001
            return f"[pdf parse failed] {e}"
        source_label = src
    else:
        p = Path(src).expanduser()
        if not p.is_file():
            return f"[error] no such file: {p}"
        try:
            reader = pypdf.PdfReader(str(p))
        except Exception as e:  # noqa: BLE001
            return f"[pdf parse failed] {e}"
        source_label = str(p)
    total = len(reader.pages)
    # Parse page ranges
    selected: list[int]
    if not pages.strip():
        selected = list(range(total))
    else:
        selected = []
        for part in pages.split(","):
            part = part.strip()
            if not part:
                continue
            if "-" in part:
                a, b = part.split("-", 1)
                try:
                    lo, hi = int(a), int(b)
                except ValueError:
                    return f"[error] bad page range: {part!r}"
                for i in range(lo - 1, min(hi, total)):
                    if 0 <= i < total:
                        selected.append(i)
            else:
                try:
                    i = int(part) - 1
                except ValueError:
                    return f"[error] bad page number: {part!r}"
                if 0 <= i < total:
                    selected.append(i)
    out: list[str] = [f"# PDF: {source_label}\n",
                      f"_{total} pages, extracting "
                      f"{len(selected)} page{'s' if len(selected) != 1 else ''}_\n"]
    body_chars = 0
    for idx in selected:
        try:
            text = (reader.pages[idx].extract_text() or "").strip()
        except Exception:  # noqa: BLE001
            text = ""
        if not text:
            continue
        chunk = f"\n--- page {idx + 1} ---\n{text}"
        out.append(chunk)
        body_chars += len(chunk)
        if body_chars > max_chars:
            out.append(f"\n…[stopped at page {idx + 1}; "
                       f"raise max_chars or narrow `pages` to read more]")
            break
    body = "\n".join(out)
    if len(body) > max_chars:
        body = _smart_truncate(body, max_chars)
    return body


def github_repo(repo: str, action: str = "info", path: str = "",
                ref: str = "", max_chars: int = 60000) -> str:
    """Read a public GitHub repo via the API (no auth required for public).

    `repo` is "owner/name" (or any github URL — we'll parse it).
    `action` ∈ {"info", "list", "read", "readme"}:
      - "info":   metadata (description, stars, default_branch, last commit)
      - "list":   directory listing at `path` (default: repo root)
      - "read":   raw file contents at `path`. Required for this action.
      - "readme": fetch README at any common name. Cheaper than `read`.

    `ref` lets you pin to a branch / tag / sha. Default = repo's default
    branch.
    """
    # Accept owner/name OR a full URL.
    m = re.match(r"^(?:https?://github\.com/)?([\w.\-]+)/([\w.\-]+?)(?:\.git)?(?:/.*)?$",
                 repo.strip())
    if not m:
        return f"[error] could not parse repo from {repo!r}"
    owner, name = m.group(1), m.group(2)
    api_base = f"https://api.github.com/repos/{owner}/{name}"

    def _api(url: str) -> tuple[int, dict | list | str]:
        req = urllib.request.Request(url, headers={
            "User-Agent": "qwen-agent",
            "Accept": "application/vnd.github.v3+json",
        })
        try:
            with urllib.request.urlopen(req, timeout=20) as r:
                body = r.read()
                try:
                    return r.status, json.loads(body)
                except ValueError:
                    return r.status, body.decode("utf-8", errors="replace")
        except urllib.error.HTTPError as e:
            try:
                detail = json.loads(e.read())
            except Exception:  # noqa: BLE001
                detail = str(e)
            return e.code, detail
        except Exception as e:  # noqa: BLE001
            return 0, f"{type(e).__name__}: {e}"

    if action == "info":
        code, data = _api(api_base)
        if code != 200 or not isinstance(data, dict):
            return f"[github info error] HTTP {code}: {str(data)[:200]}"
        desc = data.get("description") or "(no description)"
        out = [
            f"# {owner}/{name}",
            f"{desc}",
            f"",
            f"- stars: {data.get('stargazers_count', 0)}  ·  forks: {data.get('forks_count', 0)}  ·  language: {data.get('language') or 'mixed'}",
            f"- default_branch: {data.get('default_branch', '?')}",
            f"- updated: {data.get('updated_at', '?')[:10]}",
            f"- topics: {', '.join(data.get('topics') or []) or '(none)'}",
            f"- url: {data.get('html_url', '')}",
        ]
        if data.get("license"):
            out.append(f"- license: {data['license'].get('spdx_id', '?')}")
        return "\n".join(out)
    if action == "readme":
        url = f"{api_base}/readme"
        if ref:
            url += f"?ref={ref}"
        code, data = _api(url)
        if code != 200 or not isinstance(data, dict):
            return f"[github readme error] HTTP {code}: {str(data)[:200]}"
        import base64
        try:
            content = base64.b64decode(data.get("content") or "").decode(
                "utf-8", errors="replace")
        except Exception as e:  # noqa: BLE001
            return f"[readme decode failed] {e}"
        if len(content) > max_chars:
            content = _smart_truncate(content, max_chars)
        return f"# README — {owner}/{name}\n\n{content}"
    if action == "list":
        url = f"{api_base}/contents/{path.lstrip('/')}"
        if ref:
            url += f"?ref={ref}"
        code, data = _api(url)
        if code != 200:
            return f"[github list error] HTTP {code}: {str(data)[:200]}"
        if isinstance(data, dict):
            data = [data]
        if not isinstance(data, list):
            return f"[github] unexpected response shape"
        out = [f"# {owner}/{name}/{path or ''}"]
        dirs, files = [], []
        for entry in data:
            if not isinstance(entry, dict):
                continue
            name_ = entry.get("name", "?")
            if entry.get("type") == "dir":
                dirs.append(f"  📁 {name_}/")
            else:
                size = entry.get("size") or 0
                files.append(f"  📄 {name_}  ({size} B)")
        return "\n".join(out + sorted(dirs) + sorted(files))
    if action == "read":
        if not path:
            return "[error] path required for action='read'"
        url = f"{api_base}/contents/{path.lstrip('/')}"
        if ref:
            url += f"?ref={ref}"
        code, data = _api(url)
        if code != 200 or not isinstance(data, dict):
            return f"[github read error] HTTP {code}: {str(data)[:200]}"
        if data.get("type") != "file":
            return f"[error] {path} is not a file (type={data.get('type', '?')})"
        import base64
        try:
            content = base64.b64decode(data.get("content") or "").decode(
                "utf-8", errors="replace")
        except Exception as e:  # noqa: BLE001
            return f"[file decode failed] {e}"
        if len(content) > max_chars:
            content = _smart_truncate(content, max_chars)
        return f"# {owner}/{name}/{path}\n\n```\n{content}\n```"
    return f"[error] unknown action: {action!r} (use info|list|read|readme)"


def doi_resolve(doi: str) -> str:
    """Resolve a DOI via the doi.org content negotiation API.

    Returns formatted citation metadata (authors, title, container, year,
    DOI, URL). Use this instead of fetching the publisher landing page —
    cleaner output, no paywalls / cookie banners.
    """
    doi = doi.strip()
    # Strip a leading https://doi.org/ if present.
    doi = re.sub(r"^https?://(?:dx\.)?doi\.org/", "", doi).strip()
    if not doi or not re.match(r"^10\.[\d.]+/", doi):
        return f"[error] does not look like a DOI: {doi!r}"
    url = f"https://doi.org/{doi}"
    req = urllib.request.Request(url, headers={
        "Accept": "application/vnd.citationstyles.csl+json",
        "User-Agent": "qwen-agent",
    })
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            data = json.loads(r.read())
    except Exception as e:  # noqa: BLE001
        return f"[doi resolve failed] {type(e).__name__}: {e}"
    out = []
    title = data.get("title")
    if isinstance(title, list):
        title = " ".join(title)
    out.append(f"# {title or '(no title)'}")
    authors = data.get("author") or []
    auth_strs = []
    for a in authors[:8]:
        family = a.get("family", "")
        given = a.get("given", "")
        suffix = a.get("suffix", "")
        if family or given:
            auth_strs.append(" ".join(p for p in (given, family, suffix) if p).strip())
    if len(authors) > 8:
        auth_strs.append(f"(+{len(authors) - 8} more)")
    if auth_strs:
        out.append(f"**authors:** {', '.join(auth_strs)}")
    container = data.get("container-title")
    if isinstance(container, list):
        container = container[0] if container else ""
    if container:
        out.append(f"**venue:** {container}")
    issued = (data.get("issued") or {}).get("date-parts") or []
    if issued and isinstance(issued[0], list) and issued[0]:
        out.append(f"**year:** {issued[0][0]}")
    out.append(f"**doi:** {doi}")
    out.append(f"**url:** {url}")
    abstract = data.get("abstract")
    if abstract:
        # Clean stray HTML markup that some publishers put in CSL abstracts.
        abstract = re.sub(r"<[^>]+>", "", abstract)
        if len(abstract) > 1500:
            abstract = abstract[:1500] + "…"
        out.append("\n## Abstract\n\n" + abstract.strip())
    return "\n".join(out)


# ---------- SEC EDGAR -------------------------------------------------------
#
# Deterministic discovery of SEC filings for any US-listed company. This
# avoids the failure mode where the model burns its web_search budget
# trying to find a 10-K/10-Q URL through DDG (which surfaces third-party
# scrapers far above the actual sec.gov filing). Fully general — works for
# any ticker that has an SEC CIK, which is every company that files with
# the SEC. Mirrors arxiv_search/doi_resolve in spirit: a free public API
# wrapped as a single tool that returns ready-to-fetch URLs.

_SEC_TICKER_CACHE: dict[str, str] | None = None  # ticker (uppercase) -> CIK str
_SEC_USER_AGENT = os.environ.get(
    "QWEN_SEC_USER_AGENT", "qwen-agent finance-eval contact@example.com")


def _sec_load_ticker_map() -> dict[str, str]:
    """Lazy-fetch the SEC's ticker→CIK mapping.

    SEC publishes this as a flat JSON at company_tickers.json — keyed by
    integer index, value is {cik_str, ticker, title}. We invert to ticker
    upper → 10-digit CIK string. Cached process-wide; ~600 KB / one HTTP
    call per process.
    """
    global _SEC_TICKER_CACHE
    if _SEC_TICKER_CACHE is not None:
        return _SEC_TICKER_CACHE
    try:
        req = urllib.request.Request(
            "https://www.sec.gov/files/company_tickers.json",
            headers={"User-Agent": _SEC_USER_AGENT},
        )
        with urllib.request.urlopen(req, timeout=20) as r:
            data = json.loads(r.read())
    except Exception:  # noqa: BLE001
        _SEC_TICKER_CACHE = {}
        return _SEC_TICKER_CACHE
    out: dict[str, str] = {}
    for v in data.values():
        ticker = (v.get("ticker") or "").upper()
        cik = v.get("cik_str")
        if ticker and cik is not None:
            out[ticker] = f"{int(cik):010d}"
    _SEC_TICKER_CACHE = out
    return out


def sec_filings(ticker: str, form: str = "10-K", limit: int = 5,
                year: str = "") -> str:
    """List recent SEC EDGAR filings for a US-listed company.

    Returns a compact list of filings with direct URLs the model can
    `web_fetch`. This is the deterministic path to authoritative source
    documents (10-K annual, 10-Q quarterly, 8-K current, DEF 14A proxy,
    etc.) — much more reliable than `web_search` for finding the right
    filing of a specific period.

    Args:
      ticker: stock ticker (case-insensitive). E.g. "NFLX", "AAPL", "TSLA".
      form:   filing type. Common values: "10-K", "10-Q", "8-K", "DEF 14A",
              "S-1". Default "10-K".
      limit:  how many recent filings to return (default 5).
      year:   optional 4-digit year filter (e.g. "2024"). Filings outside
              the year are skipped before the limit cap is applied.

    Returns markdown-formatted list. Each entry includes:
      - filing date + period of report
      - direct URL to the index page (use web_fetch to get list of exhibits)
      - direct URL to the primary document HTML

    Examples:
      sec_filings("NFLX", "10-K", 3)         # last 3 annual reports
      sec_filings("NFLX", "10-Q", 4, "2024") # 2024's quarterly reports
      sec_filings("AAPL", "8-K", 10)         # last 10 current reports
    """
    ticker = (ticker or "").strip().upper()
    if not ticker:
        return "[error] sec_filings: ticker is required"
    form = str(form or "10-K").strip().upper()
    # Coerce year to string up front — the model often passes it as an int
    # (the JSON schema declares it string, but tool-call argument JSON is
    # generated by the model and an int slips through). Bare `if year` then
    # `re.fullmatch(year)` previously raised TypeError on int.
    if year is None or year == "":
        year = ""
    else:
        year = str(year).strip()
        if not re.fullmatch(r"\d{4}", year):
            return f"[error] sec_filings: year must be 4 digits, got {year!r}"
    tickers = _sec_load_ticker_map()
    if not tickers:
        return ("[error] sec_filings: could not load SEC ticker map. "
                "Network down or SEC unreachable; try web_search instead.")
    cik = tickers.get(ticker)
    if not cik:
        return (f"[error] sec_filings: ticker {ticker!r} not found in SEC "
                f"EDGAR. Confirm the ticker is correct (US-listed only).")
    try:
        req = urllib.request.Request(
            f"https://data.sec.gov/submissions/CIK{cik}.json",
            headers={"User-Agent": _SEC_USER_AGENT},
        )
        with urllib.request.urlopen(req, timeout=30) as r:
            data = json.loads(r.read())
    except Exception as e:  # noqa: BLE001
        return (f"[error] sec_filings: SEC API error for {ticker} "
                f"(CIK {cik}): {type(e).__name__}: {e}")
    name = data.get("name") or ticker
    recent = (data.get("filings") or {}).get("recent") or {}
    forms = recent.get("form") or []
    accessions = recent.get("accessionNumber") or []
    primary_docs = recent.get("primaryDocument") or []
    primary_descs = recent.get("primaryDocDescription") or []
    filing_dates = recent.get("filingDate") or []
    periods = recent.get("reportDate") or []
    out_lines = [f"# {name} ({ticker}, CIK {cik}) — recent {form} filings"]
    n_kept = 0
    for i, f_type in enumerate(forms):
        if f_type.upper() != form:
            continue
        if year and not (filing_dates[i].startswith(year) or
                         (i < len(periods) and periods[i].startswith(year))):
            continue
        accession = accessions[i].replace("-", "")
        primary = primary_docs[i] if i < len(primary_docs) else ""
        desc = primary_descs[i] if i < len(primary_descs) else ""
        idx_url = (f"https://www.sec.gov/Archives/edgar/data/"
                   f"{int(cik)}/{accession}/")
        doc_url = idx_url + primary if primary else idx_url
        out_lines.append(
            f"\n## {filing_dates[i]} — period: {periods[i] or 'n/a'}"
            + (f" — {desc}" if desc else "") +
            f"\n- index:  {idx_url}"
            f"\n- doc:    {doc_url}"
        )
        n_kept += 1
        if n_kept >= limit:
            break
    if n_kept == 0:
        suffix = f" in {year}" if year else ""
        return (f"[no {form} filings found for {ticker}{suffix}. The recent-"
                f"filings list contains forms: "
                f"{', '.join(sorted(set(forms))[:20])}]")
    out_lines.append(
        f"\n_To read a filing, call_ `web_fetch(<doc URL>)` _from above. "
        f"For the full exhibit list of one filing, fetch its index URL._"
    )
    return "\n".join(out_lines)


def csv_summary(path: str, max_rows: int = 20,
                describe: bool = True) -> str:
    """Quick stats + preview for a CSV/TSV/JSONL file.

    Cheaper than `python_run` + pandas boilerplate for the common "what's
    in this file?" question. Returns: shape, column types, first N rows,
    plus per-numeric-column min/max/mean/stddev when `describe=True`.

    Auto-detects delimiter for CSV/TSV; JSONL is also supported.
    """
    p = Path(path).expanduser()
    if not p.is_file():
        return f"[error] no such file: {p}"
    ext = p.suffix.lower()
    try:
        if ext in (".jsonl", ".ndjson"):
            rows: list[dict] = []
            with p.open(encoding="utf-8", errors="replace") as f:
                for i, line in enumerate(f):
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        rows.append(json.loads(line))
                    except json.JSONDecodeError:
                        continue
                    if i >= 50000:  # hard cap on lines processed
                        break
            cols = sorted({k for r in rows for k in r.keys()
                           if isinstance(r, dict)})
        else:
            import csv
            with p.open(encoding="utf-8", errors="replace", newline="") as f:
                sample = f.read(8192)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
                except csv.Error:
                    dialect = csv.excel
                reader = csv.DictReader(f, dialect=dialect)
                rows = []
                for i, row in enumerate(reader):
                    rows.append(row)
                    if i >= 50000:
                        break
                cols = list(reader.fieldnames or [])
    except Exception as e:  # noqa: BLE001
        return f"[csv read failed] {type(e).__name__}: {e}"
    if not rows:
        return "(empty file)"
    n = len(rows)
    # Column types: try numeric, then date, else string.
    out = [f"# {p}", f"rows: {n}  ·  cols: {len(cols)}", ""]
    col_types: dict[str, str] = {}
    col_numeric: dict[str, list[float]] = {}
    for c in cols:
        seen_num = seen_str = 0
        for r in rows[:1000]:
            v = r.get(c) if isinstance(r, dict) else None
            if v is None or v == "":
                continue
            try:
                f = float(v)
                col_numeric.setdefault(c, []).append(f)
                seen_num += 1
            except (TypeError, ValueError):
                seen_str += 1
        if seen_num and not seen_str:
            col_types[c] = "number"
        elif seen_num and seen_str:
            col_types[c] = "mixed"
        else:
            col_types[c] = "string"
    # Header row.
    out.append("## Columns")
    for c in cols:
        out.append(f"- `{c}` ({col_types.get(c, '?')})")
    out.append("")
    # Preview.
    out.append(f"## First {min(max_rows, n)} rows")
    for r in rows[:max_rows]:
        if isinstance(r, dict):
            preview = {c: (str(r.get(c, ""))[:60]) for c in cols[:8]}
            out.append("- " + ", ".join(f"{k}={v!r}" for k, v in preview.items()))
    if describe and col_numeric:
        out.append("")
        out.append("## Numeric describe")
        for c, values in col_numeric.items():
            if not values:
                continue
            mean = sum(values) / len(values)
            mn, mx = min(values), max(values)
            var = sum((v - mean) ** 2 for v in values) / max(1, len(values) - 1)
            std = var ** 0.5
            out.append(f"- `{c}`: n={len(values)}  min={mn:.4g}  "
                       f"max={mx:.4g}  mean={mean:.4g}  std={std:.4g}")
    return "\n".join(out)


def make_table(headers: list, rows: list, align: str = "left",
                title: str = "", numbered: bool = False) -> str:
    """Build a clean Markdown table from rows + headers.

    The model often produces tables, but does it inconsistently — drift on
    column widths, missing pipes, mismatched cell counts. Calling this tool
    yields a deterministic, render-ready table.

    Args:
        headers: list of column titles.
        rows: list of rows; each row is a list whose length should match
            `headers` (excess truncated; missing padded with empty string).
        align: per-column alignment. Either a single string ("left"|"right"|
            "center") applied to all columns, or a comma-separated list
            ("left,right,center") matching the column count.
        title: optional title (rendered as a Markdown ## header above).
        numbered: if true, prepend a "#" column with 1-based row numbers.
    """
    if not isinstance(headers, list):
        return "[error] headers must be a list"
    if not isinstance(rows, list):
        return "[error] rows must be a list"
    if numbered:
        headers = ["#"] + list(headers)
        rows = [[str(i + 1)] + list(r) for i, r in enumerate(rows)]
    n_cols = len(headers)
    if n_cols == 0:
        return "[error] at least one header is required"

    # Normalize cells to strings, pad/truncate rows to n_cols.
    norm_rows: list[list[str]] = []
    for r in rows:
        cells = [str(c) if c is not None else "" for c in (r or [])]
        if len(cells) < n_cols:
            cells = cells + [""] * (n_cols - len(cells))
        elif len(cells) > n_cols:
            cells = cells[:n_cols]
        # Replace pipes and newlines so the row stays on one line.
        cells = [c.replace("\\", "\\\\").replace("|", "\\|").replace("\n", " ")
                 for c in cells]
        norm_rows.append(cells)

    # Resolve per-column alignment string.
    if "," in align:
        per_col = [a.strip().lower() for a in align.split(",")]
    else:
        per_col = [align.strip().lower()] * n_cols
    if len(per_col) < n_cols:
        per_col = per_col + ["left"] * (n_cols - len(per_col))
    align_token = {"left": ":---", "center": ":---:", "right": "---:"}
    sep_cells = [align_token.get(a, ":---") for a in per_col[:n_cols]]

    # Compute column widths for tidy output (cosmetic, doesn't affect render).
    widths = [len(str(h)) for h in headers]
    for cells in norm_rows:
        for i, c in enumerate(cells):
            widths[i] = max(widths[i], len(c))
    widths = [max(w, 3) for w in widths]

    def _fmt_row(cells: list[str]) -> str:
        out = []
        for i, c in enumerate(cells):
            w = widths[i]
            a = per_col[i] if i < len(per_col) else "left"
            if a == "right":
                out.append(c.rjust(w))
            elif a == "center":
                out.append(c.center(w))
            else:
                out.append(c.ljust(w))
        return "| " + " | ".join(out) + " |"

    lines: list[str] = []
    if title:
        lines.append(f"## {title}")
        lines.append("")
    lines.append(_fmt_row([str(h) for h in headers]))
    # Pad separator tokens to column widths
    sep_padded = []
    for i, t in enumerate(sep_cells[:n_cols]):
        w = widths[i]
        if t == ":---:":
            sep_padded.append(":" + "-" * (w - 2) + ":")
        elif t == "---:":
            sep_padded.append("-" * (w - 1) + ":")
        else:
            sep_padded.append(":" + "-" * (w - 1))
    lines.append("| " + " | ".join(sep_padded) + " |")
    for cells in norm_rows:
        lines.append(_fmt_row(cells))
    return "\n".join(lines)


def inspect_data(path: str, max_chars: int = 4000) -> str:
    """Auto-detect a data file's format and return a compact summary.

    Routing by extension (lowercase):
      .csv .tsv  → reuses csv_summary
      .json      → root type + key list (or array len) + first value preview
      .parquet   → pyarrow schema + nrows + first 5 rows (if pyarrow installed)
      .ipynb     → cell counts + outline of code cells (first 80 chars each)
      .yaml .yml → top-level keys + a 1-line type sketch per key
      .xlsx      → sheet names + per-sheet shape (if openpyxl installed)
      anything else → first `max_chars` characters as plain text

    Designed for ingestion: hand it a path, get a structured summary in
    one shot — no need for the model to decide what tool to use, or to
    pull bytes into context only to throw most of them away.
    """
    p = path.strip()
    if not os.path.exists(p):
        return f"[error] no file at {p!r}"
    if os.path.isdir(p):
        try:
            entries = sorted(os.listdir(p))[:50]
        except OSError as e:
            return f"[error] {e}"
        return f"directory: {p}\n" + "\n".join(f"  {e}" for e in entries)
    ext = os.path.splitext(p)[1].lower()

    if ext in (".csv", ".tsv"):
        try:
            return csv_summary(p)
        except Exception as e:  # noqa: BLE001
            return f"[csv_summary error] {type(e).__name__}: {e}"

    if ext == ".json":
        try:
            with open(p, encoding="utf-8") as f:
                data = json.load(f)
        except (OSError, json.JSONDecodeError) as e:
            return f"[json error] {e}"
        if isinstance(data, dict):
            keys = list(data.keys())
            preview_keys = keys[:30]
            sample = {k: data[k] for k in preview_keys[:5]}
            try:
                sample_repr = json.dumps(sample, indent=2, ensure_ascii=False, default=str)[:1500]
            except Exception:  # noqa: BLE001
                sample_repr = str(sample)[:1500]
            extra = f"\n... and {len(keys) - 30} more" if len(keys) > 30 else ""
            return (f"json object  ({len(keys)} keys)\n"
                    f"keys: {preview_keys}{extra}\n"
                    f"--- first 5 values ---\n{sample_repr}")
        if isinstance(data, list):
            sample = data[:5]
            try:
                sample_repr = json.dumps(sample, indent=2, ensure_ascii=False, default=str)[:1500]
            except Exception:  # noqa: BLE001
                sample_repr = str(sample)[:1500]
            return f"json array  (length {len(data)})\n--- first 5 items ---\n{sample_repr}"
        return f"json scalar: {data!r}"

    if ext == ".parquet":
        try:
            import pyarrow.parquet as pq  # type: ignore
        except ImportError:
            return "[parquet] pyarrow not installed"
        try:
            tbl = pq.read_table(p)
        except Exception as e:  # noqa: BLE001
            return f"[parquet read error] {e}"
        head = tbl.slice(0, 5).to_pylist()
        head_repr = json.dumps(head, indent=2, default=str, ensure_ascii=False)[:2000]
        return (f"parquet  rows={tbl.num_rows}  cols={tbl.num_columns}\n"
                f"schema:\n{tbl.schema}\n--- first 5 rows ---\n{head_repr}")

    if ext == ".ipynb":
        try:
            with open(p, encoding="utf-8") as f:
                nb = json.load(f)
        except (OSError, json.JSONDecodeError) as e:
            return f"[ipynb error] {e}"
        cells = nb.get("cells") or []
        n_code = sum(1 for c in cells if c.get("cell_type") == "code")
        n_md = sum(1 for c in cells if c.get("cell_type") == "markdown")
        out: list[str] = [f"notebook  total_cells={len(cells)}  code={n_code}  markdown={n_md}"]
        out.append("--- cell outline (first 80 chars per code cell) ---")
        shown = 0
        for i, c in enumerate(cells):
            if c.get("cell_type") != "code":
                continue
            shown += 1
            if shown > 25:
                out.append(f"  ... and {n_code - 25} more code cells")
                break
            src = c.get("source") or []
            if isinstance(src, list):
                src = "".join(src)
            head = " ".join(str(src).split())[:80]
            out.append(f"  [{i:3d}] {head}")
        return "\n".join(out)

    if ext in (".yaml", ".yml"):
        try:
            import yaml  # type: ignore
        except ImportError:
            return "[yaml] pyyaml not installed"
        try:
            with open(p, encoding="utf-8") as f:
                data = yaml.safe_load(f)
        except Exception as e:  # noqa: BLE001
            return f"[yaml error] {e}"
        if isinstance(data, dict):
            lines = [f"yaml object ({len(data)} keys)"]
            for k, v in list(data.items())[:30]:
                t = type(v).__name__
                if isinstance(v, (dict, list)):
                    extra = f" len={len(v)}"
                else:
                    extra = ""
                lines.append(f"  {k}: <{t}>{extra}")
            return "\n".join(lines)
        return f"yaml: {type(data).__name__}\n{str(data)[:1500]}"

    if ext == ".xlsx":
        try:
            import openpyxl  # type: ignore
        except ImportError:
            return "[xlsx] openpyxl not installed"
        try:
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        except Exception as e:  # noqa: BLE001
            return f"[xlsx error] {e}"
        out = [f"xlsx workbook  sheets={wb.sheetnames}"]
        for s in wb.sheetnames[:10]:
            ws = wb[s]
            out.append(f"  sheet {s!r}: dim={ws.dimensions}  max_row={ws.max_row}  max_col={ws.max_column}")
        return "\n".join(out)

    # Plain text fallback
    try:
        with open(p, "rb") as f:
            raw = f.read(max_chars + 256)
    except OSError as e:
        return f"[error] {e}"
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError:
        try:
            text = raw.decode("utf-16")
        except UnicodeDecodeError:
            text = raw.decode("latin-1", "replace")
    if len(text) > max_chars:
        text = text[:max_chars] + f"\n…[truncated {len(raw) - max_chars} bytes]"
    return f"text  ext={ext or '(none)'}  size={os.path.getsize(p)}\n--- content ---\n{text}"


def now(tz: str = "UTC", fmt: str = "iso") -> str:
    """Return the current date/time in the requested timezone.

    Models routinely struggle with date math (training cutoff vs the
    user's "today") and timezones are an even worse blind spot. This tool
    surfaces both at once. `tz` accepts IANA names ("America/New_York",
    "Europe/London", etc.); `fmt` ∈ {"iso", "rfc", "date", "weekday",
    "epoch"}. Default ISO 8601 with timezone offset.
    """
    import datetime as _dt
    try:
        from zoneinfo import ZoneInfo
        zone = ZoneInfo(tz) if tz else ZoneInfo("UTC")
    except Exception:  # noqa: BLE001
        return f"[error] unknown timezone: {tz!r}"
    nowdt = _dt.datetime.now(zone)
    if fmt == "iso":
        return nowdt.isoformat(timespec="seconds")
    if fmt == "rfc":
        return nowdt.strftime("%a, %d %b %Y %H:%M:%S %z")
    if fmt == "date":
        return nowdt.date().isoformat()
    if fmt == "weekday":
        return nowdt.strftime("%A, %B %-d, %Y")
    if fmt == "epoch":
        return str(int(nowdt.timestamp()))
    return f"[error] unknown fmt: {fmt!r} (use iso|rfc|date|weekday|epoch)"


# ---------- shared cached + parallel dispatcher ---------------------------
# Both the agent CLI (agent.py) and the chat UI (qwen_ui.py) need:
#  (1) cross-turn cache of read tools so the model isn't punished for asking
#      the same thing twice
#  (2) intra-turn dedup so parallel duplicate tool_calls dispatch once
#  (3) parallel execution of distinct calls in one turn
# Putting it here means both consumers share identical semantics, and the
# class is also what we expose via the LangGraph / Pydantic AI converters.

_FS_READ_TOOLS = {"grep", "list_files", "read_file"}
_WEB_READ_TOOLS = {"web_search", "web_fetch", "web_outline",
                   "arxiv_search", "arxiv_fetch", "github_repo",
                   "doi_resolve"}
_FS_WRITE_TOOLS = {"write_file", "edit_file", "bash", "apply_patch",
                   "append_finding", "write_file_verified", "notebook_edit"}


def _arg_key(args: dict) -> str:
    """Order-insensitive stable hashable key for tool arguments."""
    try:
        return json.dumps(args, sort_keys=True, default=str)
    except Exception:  # noqa: BLE001
        return repr(sorted(args.items()) if isinstance(args, dict) else args)


_DISPATCHER_CACHE_MAX_ENTRIES = int(
    os.environ.get("QWEN_DISPATCHER_CACHE_MAX", "256"))


# Free-form URL extraction. Trailing punctuation that's commonly NOT part of
# the URL (closing brackets/parens, periods at sentence end, commas) is
# stripped post-match — handles markdown links and prose alike.
_URL_RE = re.compile(r"https?://[^\s<>\"'\\\)\]\}\|`]+")
_URL_TRAIL_STRIP = ".,;:!?)]}>"

# When the model fetches a URL and gets cross-host-redirected to a path that
# matches one of these patterns, it's a bot-block / captcha page rather than
# real content. Returning the blocked HTML poisons context — return a short
# explanation instead so the model knows to try a different source.
_BLOCK_REDIRECT_PATH_PATTERNS = (
    "/blocked",
    "/captcha",
    "/robot-check",
    "/access-denied",
    "/bot-check",
    "/cf-error",
    "/cdn-cgi/error",
    "/cdn-cgi/challenge",
    "/_incapsula_resource",
    "/distil-",
    "/checkjavascript",
    "/bm-verify",
    "/px-captcha",
)

# Hosts where same-host fetching of NEW slugs is plausible without a prior
# search hit — official docs, public APIs, well-known schemes. The URL guard
# (require URL to have appeared in a prior search/user message) skips these
# so the model can chain `docs.python.org/3/library/x.html` → `.../y.html`
# without burning a search call. This is Claude Code's "preapproved" pattern,
# trimmed to the basics; expand via QWEN_WEB_PREAPPROVED_HOSTS.
_DEFAULT_PREAPPROVED_HOSTS = frozenset({
    "docs.python.org", "developer.mozilla.org", "react.dev", "nodejs.org",
    "doc.rust-lang.org", "go.dev", "pkg.go.dev", "kotlinlang.org",
    "swift.org", "docs.swift.org", "www.typescriptlang.org",
    "tailwindcss.com", "vuejs.org", "nextjs.org", "expressjs.com",
    "fastapi.tiangolo.com", "flask.palletsprojects.com",
    "github.com",  # public repos; private pages WebFetch will fail anyway
    "raw.githubusercontent.com", "gist.github.com",
    "stackoverflow.com", "wikipedia.org", "en.wikipedia.org",
    "arxiv.org", "fdc.nal.usda.gov",  # USDA FoodData Central
    "huggingface.co",
    # SEC EDGAR — public US-government corporate-filing archive. Every URL
    # under www.sec.gov/ is intended to be fetchable; no auth, no captcha,
    # stable URL patterns (cgi-bin/browse-edgar, /Archives/edgar/data/...).
    # Whitelisting here lets the model use EDGAR's filing-finder endpoints
    # instead of trying to web_search for specific 10-K/10-Q URLs.
    "www.sec.gov", "efts.sec.gov", "data.sec.gov",
    # Iter 29: finance press / investor-relations / market-data hosts that
    # appear constantly in finance and trading workloads. Adding here so
    # the model can fetch press releases and market data without needing
    # the URL to be in a prior search result first. URL guard still
    # protects against fabricated URLs at non-preapproved hosts.
    "businesswire.com", "www.businesswire.com",
    "prnewswire.com", "www.prnewswire.com",
    "globenewswire.com", "www.globenewswire.com",
    "finance.yahoo.com", "www.marketwatch.com",
    "www.federalreserve.gov", "fred.stlouisfed.org",
    "www.bls.gov", "www.bea.gov",
    # Iter 37 (loosen-for-finance): major business news + data aggregators.
    # Most of these are paywalled, but the URL guard's purpose is to catch
    # fabricated slugs — not gate access. A 403/paywall return is fine; the
    # model marks the URL dead and pivots. Without these in the allowlist,
    # the model can't even attempt the fetch from a search result it just
    # got back, because the URL guard fires before _check_url_seen does.
    "www.cnbc.com", "cnbc.com",
    "www.reuters.com", "reuters.com",
    "www.bloomberg.com", "bloomberg.com",
    "www.ft.com", "ft.com",
    "www.wsj.com", "wsj.com",
    "www.barrons.com", "barrons.com",
    "fortune.com", "www.fortune.com",
    "www.forbes.com", "forbes.com",
    "seekingalpha.com", "www.seekingalpha.com",
    "www.macrotrends.net", "macrotrends.net",
    "stockanalysis.com", "www.stockanalysis.com",
    "simplywall.st", "www.simplywall.st",
    "www.nasdaq.com", "nasdaq.com",
    "www.nyse.com", "nyse.com",
    "finance.google.com", "www.google.com/finance",
    "investor.gov", "www.investor.gov",
    "sec.report", "www.sec.report",
    "www.treasury.gov", "treasury.gov",
    "www.imf.org", "imf.org", "www.worldbank.org", "worldbank.org",
    "www.oecd.org", "oecd.org",
    "www.morningstar.com", "morningstar.com",
    "tradingview.com", "www.tradingview.com",
})

# Subdomain prefixes that nearly always point to a company's own
# investor-relations / corporate-affairs surface. A URL like
# `ir.netflix.com/financial-info/quarterly-results` is essentially as
# trustworthy as a sec.gov fetch: the host is the company itself,
# the path is structured. Without this prefix-match the model would
# need to find the IR URL via a search first, even when the structure
# is canonical.
_IR_SUBDOMAIN_PREFIXES = (
    "ir.", "investor.", "investors.", "investorrelations.",
    "investor-relations.", "investorrelations-", "corp.", "corporate.",
)


def _normalize_url(url: str) -> str:
    """Lowercase host + strip fragment/trailing-slash for set-membership.
    Two URLs that differ only by `?utm_source=` or trailing `/` should match.
    """
    try:
        from urllib.parse import urlparse, urlunparse
        p = urlparse(url.strip())
        if not p.hostname:
            return ""
        host = p.hostname.lower()
        path = p.path.rstrip("/") or ""
        # Drop fragment + commonly-noisy tracking params; keep the rest of
        # query as-is since it's often load-bearing (e.g., page IDs).
        return urlunparse((p.scheme.lower(), host, path, "", p.query, ""))
    except Exception:  # noqa: BLE001
        return ""


def _extract_urls(text: str) -> list[str]:
    """Pull all http(s) URLs from a blob of text, normalized for matching."""
    if not text:
        return []
    urls: list[str] = []
    for raw in _URL_RE.findall(text):
        cleaned = raw.rstrip(_URL_TRAIL_STRIP)
        norm = _normalize_url(cleaned)
        if norm:
            urls.append(norm)
    return urls


def _preapproved_hosts() -> frozenset[str]:
    extra = os.environ.get("QWEN_WEB_PREAPPROVED_HOSTS", "").strip()
    if not extra:
        return _DEFAULT_PREAPPROVED_HOSTS
    extra_set = {h.strip().lower() for h in extra.split(",") if h.strip()}
    return _DEFAULT_PREAPPROVED_HOSTS | frozenset(extra_set)


class CachedDispatcher:
    """Stateful tool-dispatch helper with cross-call cache + intra-turn dedup.

    State is per-instance, so the agent CLI uses one global instance for the
    duration of its interactive session, and the chat UI keeps one per
    session_id (so two browser tabs don't poison each other's web cache).

    The cache stores are bounded LRU dicts (`_DISPATCHER_CACHE_MAX_ENTRIES`
    each, default 256). Without this bound, a long-running chat that fetches
    hundreds of unique URLs would let the dict grow unbounded and pin every
    response in memory forever; with it, oldest entries evict naturally and
    the cache stays at ~MB-scale.
    """

    def __init__(self,
                 max_entries: int = _DISPATCHER_CACHE_MAX_ENTRIES) -> None:
        from collections import OrderedDict
        self._max_entries = max(16, int(max_entries))
        self.fs_generation = 0
        self.fs_cache: "OrderedDict[tuple, tuple[int, str]]" = OrderedDict()
        self.web_cache: "OrderedDict[tuple, str]" = OrderedDict()
        # Per-key call counts for cached read tools. We track separately from
        # the LRU caches so eviction doesn't reset the counter; that way a
        # model that loops 50× on one query while occasionally fetching new
        # URLs still gets escalated to a hard refusal on its 3rd hit.
        self.web_call_counts: dict[tuple, int] = {}
        self.fs_call_counts: dict[tuple, int] = {}
        # Per-(args) bash invocation count. `bash` lives in _FS_WRITE_TOOLS,
        # so it bypasses the LRU cache AND fs_call_counts gets cleared on
        # every write (which bash itself triggers). Without a separate
        # counter, the model can fire byte-identical bash commands forever
        # — observed in session 20260514-085942 where the same 779-char
        # `find ... *.json -not -path ...` ran 13× consecutively. This
        # counter survives writes and is only reset by start_new_task.
        self.bash_call_counts: dict[str, int] = {}
        self._bash_dup_max = int(os.environ.get("QWEN_BASH_DUP_MAX", "3"))
        # Total web_search calls this session — capped per Claude Code's
        # `max_uses: 8` pattern so a model that "needs one more search" 50
        # times in a row gets refused after 8.
        self.web_search_count = 0
        # Raised 8 → 15 so multi-period research questions (e.g. quarterly
        # guidance series, multi-year ARPU/metric series) can gather enough
        # data without the model hitting the cap mid-task.
        # Raised 15 → 100 to support long-horizon code-reading tasks
        # (e.g. SWE-bench Pro repo navigation) where many distinct
        # searches are legitimately needed.
        self._web_search_max = int(os.environ.get("QWEN_WEB_SEARCH_MAX", "100"))
        self._web_search_guard_enabled = (
            os.environ.get("QWEN_WEB_SEARCH_GUARD", "1") == "1")
        # Total web_fetch calls this session. Each fetch carries far more
        # context cost than a search (full page body), so the cap is on
        # successful fetches, not retried-failed ones — unlike search where
        # every attempt counts. Default 15 covers research workflows
        # (search + 5-10 fetches typical) while catching the runaway
        # "fetch every product page in the catalog" case.
        self.web_fetch_count = 0
        # Raised 15 → 25 to match the new search cap; multi-document
        # research often needs to fetch a 10-K + 10-Q + earnings release
        # for each of several periods.
        # Raised 25 → 100 for long-horizon code-reading tasks where
        # the agent must read many source files from a real repo.
        self._web_fetch_max = int(os.environ.get(
            "QWEN_WEB_FETCH_MAX_CALLS",
            # Backward-compatible alias for older shells/configs. New configs
            # should use QWEN_WEB_FETCH_MAX_CALLS so it cannot be confused with
            # QWEN_WEB_FETCH_MAX_CHARS.
            os.environ.get("QWEN_WEB_FETCH_MAX", "100"),
        ))
        self._web_fetch_guard_enabled = (
            os.environ.get("QWEN_WEB_FETCH_GUARD", "1") == "1")
        # Iter 29 efficiency: per-URL fetch counter (ignores max_chars/
        # other args). When a model fetches the SAME URL with different
        # max_chars values to "see more," each variation bypasses the
        # cache key. After 2 same-URL fetches we point to find_in_url /
        # web_outline; after 3 we hard-refuse. Caught the iter 28 pattern
        # of fetching the same DEF 14A URL 4× with max_chars variations.
        self._url_fetch_count: dict[str, int] = {}
        self._url_fetch_max = int(os.environ.get("QWEN_URL_FETCH_MAX", "3"))
        # Iter 30 efficiency: separate near-dup web_search counter. The
        # main cap excludes near-dups so legit reformulations don't burn
        # quota. Evidence (sessions observed at 13× and 19× near-dup
        # search rates in iter 28) shows the model spams near-dups to
        # dodge the cap. Iter 29 found cap=2 too tight (some sessions
        # regressed when a 3rd refinement was needed for a specific
        # quantitative needle). cap=4 with the Jaccard fallback (which
        # catches MANY more reformulations than cosine 0.97 alone) gives
        # the model room to do 4 thoughtful refinements before being told
        # to commit, while still bounding the worst-case 17-search spam
        # pattern at 5 attempts.
        self._web_search_near_dup_count = 0
        self._web_search_near_dup_max = int(os.environ.get(
            "QWEN_WEB_SEARCH_NEAR_DUP_MAX", "4"))
        # URLs that returned 4xx/5xx or were detected as block pages. The
        # next fetch attempt for any of these short-circuits with a refusal
        # — without this, the model would otherwise retry the same dead
        # URLs in slightly-different forms forever.
        self._dead_urls: set[str] = set()
        # Cumulative chars saved by triage / condense across all dispatches
        # this session. Surfaced via stats() → /api/health/extended.
        self.triage_chars_saved = 0
        self.condense_chars_saved = 0
        self.triage_pruned_count = 0
        self.condense_count = 0
        # URLs the model is "allowed" to fetch — populated from user/system
        # messages and from prior tool results (web_search hits, web_fetch
        # responses that mention URLs). web_fetch on an unseen URL is
        # refused unless the host is in the preapproved set.
        self._seen_urls: set[str] = set()
        self._url_guard_enabled = (
            os.environ.get("QWEN_URL_GUARD_ENABLE", "1") == "1")
        # URL-guard refusal counter — triggers a stronger "stop hallucinating"
        # message after the model has been refused several times in a row,
        # rather than the same message every time.
        self._url_guard_refusals = 0

    def note_text(self, text: str) -> None:
        """Seed the seen-URLs set from arbitrary text (user message, system
        prompt, graph node goal, etc.). Call this when new content enters
        the conversation so the URL guard can recognize it as legitimate.
        Idempotent — safe to call repeatedly with the same text."""
        if not text or not self._url_guard_enabled:
            return
        for u in _extract_urls(text):
            self._seen_urls.add(u)

    def start_new_task(self) -> None:
        """Reset per-task counters. Call this when a fresh user message
        arrives so the web_search/web_fetch budgets restart — otherwise
        the model gets cumulatively poorer over a long chat session.

        DOES NOT reset the seen-URLs / dead-URLs sets or the LRU response
        caches: those are session-wide context the model should keep
        accumulating (URLs the user pasted earlier are still allowed; URLs
        that returned 404 still shouldn't be retried).
        """
        self.web_search_count = 0
        self.web_fetch_count = 0
        self._url_guard_refusals = 0
        # Iter 29: reset efficiency counters per session.
        self._web_search_near_dup_count = 0
        self._url_fetch_count = {}
        # Reset bash duplicate counters per task — running the same
        # diagnostic command across separate user requests is fine.
        self.bash_call_counts = {}

    def _check_web_search_cap(self) -> str | None:
        """Return a refusal string if the per-session search cap is hit,
        else None. Called before dispatching a web_search call."""
        if not self._web_search_guard_enabled:
            return None
        # Iter 30: near-dup spam guard fires before the main cap. The
        # dispatcher has already counted enough near-dup hits to call
        # the model "stuck reformulating." Refuse the next call with
        # a sharply-actionable message: COMMIT what you have to disk
        # and call done. Iter 29 versions just said "STOP" and the
        # model often got stuck at "search-refused, no new data, no
        # synthesis" — net result FAIL because no artifact was
        # written. The new text orders the exact next two tool calls.
        if self._web_search_near_dup_count >= self._web_search_near_dup_max:
            return (
                f"{REFUSED_PREFIX}{self._web_search_near_dup_count} near-duplicate "
                "web_search reformulations. Reformulating further yields "
                "the same results. Your VERY NEXT actions MUST be: "
                "(1) `write_file` an artifact summarizing what you found "
                "+ what couldn't be retrieved, (2) `done(summary=...)`. "
                "Do NOT issue more web_search calls.]"
            )
        if self.web_search_count < self._web_search_max:
            return None
        # Tight, model-actionable refusal: explicit STOP signal, then a
        # single concrete instruction. Short message keeps token cost of
        # repeated refusals low, and the harness counter triggers an
        # immediate synthesize-now nudge after a single all-refused turn.
        return (
            f"{REFUSED_PREFIX}web_search cap of {self._web_search_max} reached. "
            "STOP searching. Synthesize from results already gathered + "
            "call done(). Do NOT issue more web_search calls.]"
        )

    def _check_bash_url_bypass(self, args: dict) -> str | None:
        """Iter 29: detect `bash curl URL` (or wget) for a URL already
        fetched via web_fetch in this session. Refuse and point at
        find_in_url. Without this, models that don't like web_fetch's
        condensation re-grab the raw HTML via bash, defeating cache,
        condense pipeline, and the URL-guard.

        Targets the EXACT pattern observed in an iter 28 session trace
        (4 consecutive bash curls). Generic shell tasks (`ls`, `python -m
        venv`, `git status`, etc.) are unaffected — only commands that
        START with curl/wget on an already-fetched URL fire the refusal.
        """
        cmd = (args.get("command") or "").strip()
        if not cmd:
            return None
        # Cheap pre-check: must start with curl or wget (with optional
        # leading flags like `set -e`, `time`, etc. ignored — those
        # patterns are vanishingly rare in the data).
        first = cmd.split(None, 1)[0].lower()
        if first not in ("curl", "wget"):
            return None
        # Find first URL in the command. Conservative regex to avoid
        # false positives on file paths or git URLs.
        m = _URL_RE.search(cmd)
        if not m:
            return None
        url = m.group(0).rstrip(_URL_TRAIL_STRIP).strip("\"'")
        norm = _normalize_url(url)
        if not norm:
            return None
        if norm not in self._url_fetch_count:
            return None
        return (
            f"{REFUSED_PREFIX}bash {first} {url!r} would bypass web_fetch's "
            "cache, condense pipeline, and URL guard for a URL already "
            "fetched in this session. To find a specific value within "
            "the page, call `find_in_url(url, needle)`. To navigate to "
            "a sub-section, call `web_outline(url)` and fetch a more "
            "targeted sub-URL.]"
        )

    def _check_url_refetch(self, url: str) -> str | None:
        """Iter 29: refuse repeated fetches of the same URL once it's
        been pulled `_url_fetch_max` times in this session, regardless
        of `max_chars` or other args. The cache key includes max_chars,
        so a model can vary that value to bypass cache; this guard sits
        OUTSIDE the cache key and points the model at the right next
        action (find_in_url for a needle, web_outline for structure).
        """
        if not self._web_fetch_guard_enabled:
            return None
        norm = _normalize_url(url)
        if not norm:
            return None
        n = self._url_fetch_count.get(norm, 0)
        if n < self._url_fetch_max:
            return None
        return (
            f"{REFUSED_PREFIX}URL fetched {n} times already in this session: "
            f"{url!r}. The body is in your context. To find a specific "
            "phrase or value, call `find_in_url(url, needle)` instead. "
            "To navigate within a long doc, call `web_outline(url)` and "
            "fetch a more targeted sub-URL.]"
        )

    def _check_web_fetch_cap(self) -> str | None:
        """Return a refusal string if the per-session fetch cap is hit."""
        if not self._web_fetch_guard_enabled:
            return None
        if self.web_fetch_count < self._web_fetch_max:
            return None
        return (
            f"{REFUSED_PREFIX}web_fetch cap of {self._web_fetch_max} reached. "
            "STOP fetching. Synthesize from pages already fetched + "
            "call done(). Do NOT issue more web_fetch calls.]"
        )

    def _check_url_dead(self, url: str) -> str | None:
        """Refuse retry of a URL that already returned 4xx/5xx or was
        flagged as a block page. Stops the 'try slightly different slug'
        loop dead."""
        if not url:
            return None
        norm = _normalize_url(url)
        if norm and norm in self._dead_urls:
            return (
                f"{REFUSED_PREFIX}URL {url!r} previously returned an error or "
                "was a block page in this session. Don't retry. Use a "
                "different source.]"
            )
        return None

    def mark_url_dead(self, url: str) -> None:
        """Public hook for fetch-side code to flag a URL as known-dead."""
        norm = _normalize_url(url)
        if norm:
            self._dead_urls.add(norm)

    def _check_url_seen(self, url: str) -> str | None:
        """Return a refusal string if the URL hasn't appeared in any prior
        search/fetch/user-message + isn't on a preapproved host. Else None.

        Most URL hallucinations look like real product pages
        (`walmart.com/ip/.../12345678`) but the slugs/IDs are fabricated.
        Refusing them here forces the model to use web_search first."""
        if not self._url_guard_enabled or not url:
            return None
        norm = _normalize_url(url)
        if not norm:
            return None
        if norm in self._seen_urls:
            return None
        # Preapproved hosts — same-host slug navigation is fine without
        # a prior search hit.
        try:
            from urllib.parse import urlparse
            host = (urlparse(url).hostname or "").lower()
        except Exception:  # noqa: BLE001
            host = ""
        approved = _preapproved_hosts()
        if host in approved:
            return None
        # Also accept subdomains of preapproved suffixes (e.g.,
        # `something.docs.python.org`) — keeps the list short.
        if any(host.endswith("." + h) or host == h for h in approved):
            return None
        # Iter 37 (loosen-for-finance): IR / investor-relations subdomains
        # follow canonical patterns across virtually all public companies.
        # `ir.netflix.com`, `investor.apple.com`, `investors.tjx.com`,
        # `corp.aapl.com` etc. — the host is the company itself, no slug
        # to fabricate. The narrowness of the prefix list (must START with
        # one of these tokens) keeps the false-positive surface tiny:
        # generic hosts like `news.example.com` are still refused.
        if any(host.startswith(p) for p in _IR_SUBDOMAIN_PREFIXES):
            return None
        self._url_guard_refusals += 1
        if self._url_guard_refusals >= 3:
            tail = (" You have been refused multiple times — STOP fabricating "
                    "URLs. Run web_search and copy a real URL from its results.")
        else:
            tail = ""
        return (
            f"{REFUSED_PREFIX}URL {url!r} not seen in any prior search result, "
            "fetch result, or user message. The model often hallucinates "
            "product/page slugs that look real but 404. Run web_search "
            "first, then web_fetch a URL from those results verbatim." + tail
            + "]"
        )

    def _harvest_urls_from_result(self, result: str) -> None:
        """After a web tool returns, scan the result text for URLs and add
        them to the seen-set so the model can follow links from there."""
        if not self._url_guard_enabled or not result:
            return
        for u in _extract_urls(result):
            self._seen_urls.add(u)

    def _put(self, store, key, value) -> None:
        if key in store:
            store.move_to_end(key)
        store[key] = value
        while len(store) > self._max_entries:
            store.popitem(last=False)  # evict LRU

    def _get(self, store, key):
        v = store.get(key)
        if v is not None:
            store.move_to_end(key)
        return v

    def dispatch(self, fn: str, args: dict) -> tuple[str, bool]:
        """Dispatch one call. Returns (result, was_cached).

        On the 3rd+ identical call to a cacheable read tool, returns a SHORT
        refusal string instead of the full cached payload. This stops the
        common loop where the model re-issues the same web_search 15× because
        the soft "[cached…]" prefix on a full re-rendered result still looks
        like fresh information to it. With the refusal-only mode the result
        adds zero new tokens and is unmistakably "STOP".
        """
        if fn in _FS_WRITE_TOOLS:
            # Iter 29: detect bash-curl bypass of web_fetch. Models that
            # don't like web_fetch's condensation sometimes call
            # `bash curl URL` to grab raw HTML — caught in the iter 28
            # pattern (4 consecutive bash curls of an SEC URL already
            # fetched via web_fetch). Refuse the curl, point at
            # find_in_url + web_outline. Generic shell command via bash
            # otherwise unaffected.
            if fn == "bash":
                refusal = self._check_bash_url_bypass(args)
                if refusal is not None:
                    return refusal, False
                # Exact-duplicate bash detection. Increment FIRST so the
                # threshold counts THIS call. At threshold, return a refusal
                # explaining what the model should do instead — re-running
                # an identical shell command never yields different output
                # (filesystem unchanged between writes is the common case),
                # so the loop is wasted compute.
                bkey = _arg_key(args)
                self.bash_call_counts[bkey] = self.bash_call_counts.get(bkey, 0) + 1
                n = self.bash_call_counts[bkey]
                if n >= self._bash_dup_max:
                    cmd_preview = args.get("command", "")
                    if isinstance(cmd_preview, str) and len(cmd_preview) > 120:
                        cmd_preview = cmd_preview[:117] + "..."
                    return (
                        f"{REFUSED_PREFIX}duplicate bash invocation #{n}. "
                        f"You already ran this exact command {n - 1} time(s) "
                        "this task with identical arguments — re-running it "
                        "won't change the output. The result is in your "
                        "context above. Either (a) change the command "
                        "(different path, different flags, broader/narrower "
                        "glob), (b) use a dedicated tool instead of bash "
                        "(write_file, read_file, find_in_file, web_fetch, "
                        "mcp_register, etc.), or (c) accept the empty result "
                        "and proceed with the data you have. Command was: "
                        f"`{cmd_preview}`]"
                    ), False
            # State-changing call invalidates the FS read cache. Counters too —
            # a write makes prior FS reads stale, so reset the call count for
            # the next read attempt to take its lumps cleanly.
            self.fs_generation += 1
            self.fs_call_counts.clear()
            return dispatch(fn, args), False
        key = (fn, _arg_key(args))
        # Iter 37 (loosen-for-finance): exact-duplicate call cap is the LAST
        # backstop after the cache markers + near-dup detection. Hardcoded
        # cap=3 was too tight for multi-period finance research where a
        # model legitimately retries the same query after a related variant
        # got near-dup-blocked. Bumped 3 → 5 (env-tunable). The cached-marker
        # response at counts 1-4 still tells the model the call was a no-op
        # — only at 5 does it hard-refuse.
        _dup_call_max = int(os.environ.get("QWEN_DUP_CALL_MAX", "5"))
        if fn in _FS_READ_TOOLS:
            entry = self._get(self.fs_cache, key)
            if entry is not None and entry[0] == self.fs_generation:
                self.fs_call_counts[key] = self.fs_call_counts.get(key, 0) + 1
                n = self.fs_call_counts[key]
                if n >= _dup_call_max:
                    return (
                        f"{REFUSED_PREFIX}duplicate filesystem read #{n} of {fn}({_arg_key(args)}). "
                        "Same result already in this conversation. Stop re-reading. "
                        "If you need different info, change the args; otherwise synthesize."
                    ), True
                marker = ("[cached: this exact call returned the same result "
                          "earlier in this session, no filesystem writes since. "
                          "Do NOT call again with these args.]\n\n")
                return marker + entry[1], True
            result = dispatch(fn, args)
            self._put(self.fs_cache, key, (self.fs_generation, result))
            self.fs_call_counts[key] = 1
            return result, False
        if fn in _WEB_READ_TOOLS:
            # Pre-call guards (Claude Code-style):
            # 1. web_search has a hard cap per session (max_uses: 8).
            # 2. web_fetch refuses URLs that haven't appeared in any prior
            #    search/fetch/user-message context (anti-hallucination).
            if fn == "web_search":
                refusal = self._check_web_search_cap()
                if refusal is not None:
                    return refusal, False
            elif fn == "web_fetch":
                refusal = self._check_web_fetch_cap()
                if refusal is not None:
                    return refusal, False
                url_arg = args.get("url", "")
                refusal = self._check_url_dead(url_arg)
                if refusal is not None:
                    return refusal, False
                refusal = self._check_url_seen(url_arg)
                if refusal is not None:
                    return refusal, False
                # Iter 29: per-URL refetch cap. Independent of cache key
                # because the model bypasses the cache by varying
                # max_chars (4× same DEF 14A URL observed). Count first, then
                # check; the increment happens AFTER successful dispatch
                # so a refused-and-retried URL gets its full quota.
                refusal = self._check_url_refetch(url_arg)
                if refusal is not None:
                    return refusal, False
            cached = self._get(self.web_cache, key)
            if cached is not None:
                self.web_call_counts[key] = self.web_call_counts.get(key, 0) + 1
                n = self.web_call_counts[key]
                if n >= _dup_call_max:
                    return (
                        f"{REFUSED_PREFIX}duplicate web call #{n} of {fn}({_arg_key(args)}). "
                        "Same query already answered above; further repeats blocked. "
                        "Move on to a different question or synthesize an answer "
                        "from what's already gathered."
                    ), True
                return ("[cached web result from earlier in this session — same "
                        "query already returned this; do NOT call again with these "
                        "args.]\n\n" + cached), True
            result = dispatch(fn, args)
            self._put(self.web_cache, key, result)
            self.web_call_counts[key] = 1
            # Post-call bookkeeping for guards:
            #   web_search counter ticks ONLY when the call actually executed
            #     a new search. Near-duplicate dedup notices return early
            #     inside web_search() with the prior summary; they didn't
            #     consume backend quota, so they shouldn't burn the cap.
            #     Without this gate, a few legit reformulations get blocked
            #     by dedup and still eat the budget, leaving the model with
            #     no searches left when it actually needs new data.
            #   URL harvest happens for any web read so the model can follow
            #     hyperlinks discovered in fetched pages without re-searching.
            if fn == "web_search":
                is_near_dup = result.lstrip().startswith(
                    "[near-duplicate of earlier search")
                if not is_near_dup:
                    self.web_search_count += 1
                else:
                    # Iter 29: track near-dup attempts in their own
                    # counter. After threshold the next near-dup web_search
                    # is hard-refused — prevents the spam pattern where
                    # model reformulates 13× to dodge the main cap.
                    self._web_search_near_dup_count += 1
            elif fn == "web_fetch":
                # Only count successful fetches against the cap so a model
                # whose first fetch was refused for hallucinated URL still
                # gets its full budget once it switches to real ones.
                is_failure = result.startswith((
                    "[blocked:", "[tool error]", "[error]",
                    "[refused]", "[REFUSED", "[empty:"))
                if not is_failure:
                    self.web_fetch_count += 1
                    # Iter 29: tick per-URL refetch counter so subsequent
                    # fetches of the SAME URL (regardless of max_chars)
                    # hit a hard cap. Caught the iter 28 same-URL pattern.
                    norm = _normalize_url(args.get("url", ""))
                    if norm:
                        self._url_fetch_count[norm] = (
                            self._url_fetch_count.get(norm, 0) + 1)
                else:
                    # Mark dead so the same URL doesn't get retried later.
                    self.mark_url_dead(args.get("url", ""))
            self._harvest_urls_from_result(result)
            return result, False
        # Unknown / non-cacheable (explore, python_run, done, ...): passthrough.
        return dispatch(fn, args), False

    def stats(self) -> dict:
        """For introspection — surfaced via /api/health if a UI wants it."""
        return {
            "fs_entries": len(self.fs_cache),
            "web_entries": len(self.web_cache),
            "fs_generation": self.fs_generation,
            "max_entries": self._max_entries,
            "web_search_count": self.web_search_count,
            "web_search_max": self._web_search_max,
            "web_fetch_count": self.web_fetch_count,
            "web_fetch_max": self._web_fetch_max,
            "web_fetch_max_calls": self._web_fetch_max,
            "web_fetch_max_chars": WEB_FETCH_DEFAULT_MAX,
            "seen_urls": len(self._seen_urls),
            "dead_urls": len(self._dead_urls),
            "triage_pruned_count": self.triage_pruned_count,
            "triage_chars_saved": self.triage_chars_saved,
            "condense_count": self.condense_count,
            "condense_chars_saved": self.condense_chars_saved,
        }

    def record_reduction(self, info: dict) -> None:
        """Caller-side hook so the agent/UI loop can fold per-tool-result
        triage/condense info dicts back into per-session totals."""
        if not isinstance(info, dict):
            return
        verdict = info.get("verdict")
        chars_in = int(info.get("chars_in") or 0)
        chars_out = int(info.get("chars_out") or 0)
        saved = max(0, chars_in - chars_out)
        if verdict == "low_relevance":
            self.triage_pruned_count += 1
            self.triage_chars_saved += saved
        elif verdict == "condensed":
            self.condense_count += 1
            self.condense_chars_saved += saved

    def dispatch_batch(self, calls: list[tuple[str, dict]],
                       max_workers: int = 6) -> list[tuple[str, bool]]:
        """Dispatch a list of tool calls. Identical (fn, args) pairs are
        deduped before dispatch; distinct calls run in parallel.

        Parallel-duplicate detection: when the model issues N>1 byte-
        identical web read calls in ONE turn (iter 23 had a session that
        fired 8 of the same web_search query in turn 13 and 4 each in
        turns 22+23), only
        the FIRST input gets the dispatch result. Inputs 2..N get a tight
        refusal marker so the model sees the redundancy explicitly rather
        than 8 copies of the same cached payload — which it tends to read
        as 'still searching' and continues looping. Per-key counter is
        incremented by the full input multiplicity, not the unique count,
        so a single 8× burst pushes past the cross-turn refusal threshold
        on its own.

        Returns one (result, was_cached) tuple per *input* call (duplicates
        share the same result object reference)."""
        if not calls:
            return []
        # Count per-key input multiplicity for web reads — this is the
        # signal we lacked before. Even purely intra-turn floods now
        # consume the duplicate budget.
        from collections import Counter
        input_counts = Counter((fn, _arg_key(args)) for fn, args in calls)
        unique_keys: dict[tuple, int] = {}
        unique_calls: list[tuple[str, dict]] = []
        for fn, args in calls:
            k = (fn, _arg_key(args))
            if k not in unique_keys:
                unique_keys[k] = len(unique_calls)
                unique_calls.append((fn, args))
        if len(unique_calls) == 1:
            try:
                unique_results: list[tuple[str, bool]] = [
                    self.dispatch(unique_calls[0][0], unique_calls[0][1])
                ]
            except Exception as e:  # noqa: BLE001
                unique_results = [(f"[tool error] {type(e).__name__}: {e}", False)]
        else:
            from concurrent.futures import ThreadPoolExecutor
            with ThreadPoolExecutor(max_workers=min(len(unique_calls), max_workers)) as ex:
                futures = [ex.submit(self.dispatch, fn, args)
                           for fn, args in unique_calls]
                unique_results = []
                for f in futures:
                    try:
                        unique_results.append(f.result())
                    except Exception as e:  # noqa: BLE001
                        unique_results.append((f"[tool error] {type(e).__name__}: {e}", False))
        # Charge cross-turn counter for the EXTRA copies (input N - 1)
        # AFTER dispatch ran (dispatch sets count to 1 on first hit, so
        # adding here preserves both signals). Net effect: a single 8×
        # parallel burst leaves web_call_counts at 8 for that key, which
        # blows past the cross-turn refusal threshold (>=3). On the NEXT
        # turn even a single re-issue of the same query will be REFUSED.
        for k, n in input_counts.items():
            if n > 1 and k[0] in _WEB_READ_TOOLS:
                self.web_call_counts[k] = self.web_call_counts.get(k, 0) + (n - 1)
        # Build per-input-position results. The FIRST occurrence of a key
        # gets the dispatch result; later occurrences of the SAME key
        # (within this batch) get a short parallel-duplicate refusal so
        # the model sees the burst as wasted, not as fresh evidence.
        out: list[tuple[str, bool]] = []
        seen_position: dict[tuple, int] = {}  # key -> first input index in `out`
        for idx, (fn, args) in enumerate(calls):
            k = (fn, _arg_key(args))
            if k not in seen_position:
                seen_position[k] = idx
                out.append(unique_results[unique_keys[k]])
                continue
            # Duplicate within the same turn. For web reads, force a
            # refusal marker so the duplicate is unmistakably wasted.
            # For non-web tools (e.g. read_file the same path twice in
            # one turn), keep the prior shared-result behavior — those
            # are usually accidental and the cache already short-circuits
            # them inside dispatch().
            if fn in _WEB_READ_TOOLS:
                out.append((
                    f"{REFUSED_PREFIX}parallel duplicate of {fn}({_arg_key(args)}) "
                    "issued earlier in THIS turn. You fired the same call "
                    "more than once in parallel; the result is already in "
                    "your context. Stop replicating the same query and either "
                    "synthesize from what's there or change the args.]",
                    True,
                ))
            else:
                out.append(unique_results[unique_keys[k]])
        return out


# ---------- registry -------------------------------------------------------

DISPATCH: dict[str, Any] = {
    "web_search": web_search,
    "web_fetch": web_fetch,
    "web_outline": web_outline,
    "find_in_url": find_in_url,
    "arxiv_search": arxiv_search,
    "arxiv_fetch": arxiv_fetch,
    "pdf_extract": pdf_extract,
    "github_repo": github_repo,
    "doi_resolve": doi_resolve,
    "sec_filings": sec_filings,
    "csv_summary": csv_summary,
    "now": now,
    "make_table": make_table,
    "inspect_data": inspect_data,
    "read_file": read_file,
    "list_files": list_files,
    "grep": grep,
    "write_file": write_file,
    "edit_file": edit_file,
    "bash": bash,
    "explore": explore,
    "memory_save": memory_save,
    "memory_search": memory_search,
    "memory_get": memory_get,
    "memory_list": memory_list,
    "memory_delete": memory_delete,
    "todo_write": todo_write,
    "enter_worktree": enter_worktree,
    "exit_worktree": exit_worktree,
    "notebook_edit": notebook_edit,
    "notebook_run": notebook_run,
    "python_run": python_run,
    "python_reset": python_reset,
    "test_run": test_run,
    "append_finding": append_finding,
    "write_file_verified": write_file_verified,
    "apply_patch": apply_patch,
    "subagent_implement": subagent_implement,
    "scratchpad": scratchpad,
    "ask_user": ask_user,
    "graph_compose": graph_compose,
    "done": done,
}


# Disable list — set QWEN_AGENT_TOOLS_DISABLE="todo_write,enter_worktree,..." to hide
# specific tools from the model. Used for baseline-vs-improved A/B comparison.
def agent_graph_list() -> str:
    """List available agent-graph definitions in the project's examples/ directory.

    Returns a short table: name, file path, node count. The chat agent can
    then call agent_graph_run with one of these names to launch a multi-agent
    workflow whose context cost stays compartmentalized per node.
    """
    proj = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    ex_dir = os.path.join(proj, "examples")
    if not os.path.isdir(ex_dir):
        return "[no examples/ directory]"
    out = ["available graphs (call agent_graph_run with the `name` column):"]
    for fn in sorted(os.listdir(ex_dir)):
        if not fn.endswith("_graph.py"):
            continue
        name = fn[:-len("_graph.py")]
        path = os.path.join(ex_dir, fn)
        n_nodes = 0
        try:
            with open(path, encoding="utf-8") as f:
                src = f.read()
            n_nodes = src.count("graph.add_node(")
        except OSError:
            pass
        out.append(f"  - name={name!r:24s}  nodes={n_nodes}  file={fn}")
    if len(out) == 1:
        return "[no *_graph.py files found in examples/]"
    return "\n".join(out)


def agent_graph_run(graph: str, inputs: str = "{}",
                     max_parallel: int = 4) -> str:
    """Run a defined agent graph and return its outputs.

    Args:
        graph: Either a graph name (e.g. 'market_research') resolved against
            examples/<name>_graph.py, OR an absolute path to a Python file
            defining a top-level `graph` AgentGraph.
        inputs: JSON string with the initial input dict, e.g. '{"topic": "..."}'.
        max_parallel: Max nodes to run in parallel (default 4).

    Returns a compact JSON summary: per-node outputs (truncated previews),
    skip flags, total wall time, and per-node tool-call counts. Use this
    when a question naturally decomposes into research → analyze → produce
    steps; each agent's context stays small so the 60k window holds.
    """
    proj = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if os.path.isabs(graph) and os.path.exists(graph):
        path = graph
    else:
        # Resolve a name like "market_research" against examples/
        ex_dir = os.path.join(proj, "examples")
        candidate = os.path.join(ex_dir, f"{graph}_graph.py")
        if not os.path.exists(candidate) and graph.endswith(".py"):
            candidate = os.path.join(ex_dir, graph)
        path = candidate
    if not os.path.exists(path):
        return f"[error] no graph at {path!r}. Use agent_graph_list to see options."
    try:
        ins = json.loads(inputs) if isinstance(inputs, str) else inputs
        if not isinstance(ins, dict):
            return f"[error] inputs must be a JSON object, got {type(ins).__name__}"
    except json.JSONDecodeError as e:
        return f"[error] inputs is not valid JSON: {e}"

    # Lazy import — keeps cold-start cheap if the chat never invokes a graph.
    sys.path.insert(0, os.path.join(proj, "scripts"))
    try:
        from agent_graph import _load_graph_module  # type: ignore
    except Exception as e:  # noqa: BLE001
        return f"[error] cannot import agent_graph: {type(e).__name__}: {e}"

    try:
        g = _load_graph_module(path)
    except Exception as e:  # noqa: BLE001
        return f"[error] failed to load graph: {type(e).__name__}: {e}"

    import time as _time
    t0 = _time.monotonic()
    try:
        out = g.run(ins, verbose=False, max_parallel=max_parallel)
    except Exception as e:  # noqa: BLE001
        return f"[error] graph.run raised: {type(e).__name__}: {e}"
    wall = _time.monotonic() - t0

    # Render a HUMAN-READABLE Markdown report so the chat model (and the
    # downstream UI which renders Markdown) gets a clean summary it can
    # quote or hand back to the user. Per-node sections, output values
    # truncated to 600 chars to keep token cost bounded.
    name = getattr(g, "name", os.path.basename(path))
    lines: list[str] = []
    lines.append(f"### Graph result: `{name}`  ({round(wall, 2)}s)")
    skipped: list[str] = []
    for nname, output in out.items():
        if isinstance(output, dict) and output.get("_skipped"):
            skipped.append(f"`{nname}` ({output.get('_reason', 'skipped')[:80]})")
            continue
        lines.append("")
        lines.append(f"#### {nname}")
        for k, v in (output or {}).items():
            if k.startswith("_"):
                continue
            if isinstance(v, str):
                txt = v.strip()
            else:
                try:
                    txt = json.dumps(v, ensure_ascii=False, indent=2, default=str)
                except Exception:  # noqa: BLE001
                    txt = str(v)
            if len(txt) > 600:
                txt = txt[:600] + f"\n…[truncated {len(txt) - 600} chars]"
            lines.append(f"**{k}**:")
            if "\n" in txt or len(txt) > 80:
                lines.append("```")
                lines.append(txt)
                lines.append("```")
            else:
                lines.append(txt)
    if skipped:
        lines.append("")
        lines.append(f"_skipped: {', '.join(skipped)}_")
    return "\n".join(lines)


def _disabled_tools() -> set[str]:
    raw = os.environ.get("QWEN_AGENT_TOOLS_DISABLE", "")
    return {t.strip() for t in raw.split(",") if t.strip()}


# Post-init: register the agent-graph tools once their definitions exist.
# They live after DISPATCH so they can refer to other top-level helpers
# defined later (json, sys, etc. already imported but the ordering keeps
# DISPATCH free of forward references).
DISPATCH["agent_graph_list"] = agent_graph_list
DISPATCH["agent_graph_run"] = agent_graph_run


# ----------- MCP harness ---------------------------------------------------
# Minimal custom-MCP support: users can register HTTP-backed tool servers
# at runtime via /api/mcps/register. Each registered MCP exposes one or
# more tools; we name them `mcp_<server>__<tool>` and route invocations to
# `<url>/tools/<tool>` with the args as JSON body.
#
# Storage: $HOME/.qwen/mcps.json. Loaded at module import; updated by
# mcp_register / mcp_unregister. The dispatcher and TOOLS list both pick up
# the latest registrations on each call (no UI restart required).

def _mcp_registry_path() -> str:
    home = os.environ.get("QWEN_HOME") or os.path.expanduser("~/.qwen")
    return os.path.join(home, "mcps.json")


def _mcp_load() -> dict:
    path = _mcp_registry_path()
    if not os.path.exists(path):
        return {"mcps": []}
    try:
        with open(path, encoding="utf-8") as f:
            d = json.load(f)
        if not isinstance(d, dict):
            return {"mcps": []}
        d.setdefault("mcps", [])
        return d
    except (OSError, json.JSONDecodeError):
        return {"mcps": []}


def _mcp_save(reg: dict) -> None:
    path = _mcp_registry_path()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(reg, f, indent=2)


_MCP_NAME_RE = re.compile(r"^[a-z][a-z0-9_]{0,30}$")


def mcp_register(name: str, url: str,
                  headers: dict | None = None,
                  tools: list | None = None) -> dict:
    """Add an MCP entry to the registry. Returns the saved record."""
    name = (name or "").strip().lower()
    if not _MCP_NAME_RE.match(name):
        raise ValueError("name must match [a-z][a-z0-9_]{0,30}")
    url = (url or "").strip()
    if not url.startswith("http://") and not url.startswith("https://"):
        raise ValueError("url must start with http:// or https://")
    headers = headers or {}
    tools = tools or []
    if not isinstance(headers, dict) or not isinstance(tools, list):
        raise ValueError("headers must be dict, tools must be list")
    reg = _mcp_load()
    reg["mcps"] = [m for m in reg["mcps"] if m.get("name") != name]
    entry = {"name": name, "url": url.rstrip("/"),
             "headers": headers, "tools": tools}
    reg["mcps"].append(entry)
    _mcp_save(reg)
    # Auto-discover tools from the MCP server if the tools list is empty.
    if not tools:
        discovered = _mcp_discover_tools(url)
        if discovered:
            # Update the saved entry with discovered tools.
            reg = _mcp_load()
            for m in reg["mcps"]:
                if m.get("name") == name:
                    m["tools"] = discovered
                    break
            _mcp_save(reg)
            entry["tools"] = discovered
    return entry


def mcp_unregister(name: str) -> bool:
    reg = _mcp_load()
    before = len(reg["mcps"])
    reg["mcps"] = [m for m in reg["mcps"] if m.get("name") != name]
    if len(reg["mcps"]) == before:
        return False
    _mcp_save(reg)
    return True


def mcp_list() -> str:
    """Return a short summary of registered MCPs and their exposed tools."""
    reg = _mcp_load()
    if not reg["mcps"]:
        return (
            "[no MCPs registered.]\n"
            "To add one, call the `mcp_register` tool directly — do NOT "
            "search the filesystem for a config file. There is no config "
            "file to edit; the registry lives at ~/.qwen/mcps.json and is "
            "managed entirely through `mcp_register` / `mcp_unregister`. "
            "Example: mcp_register(name='weather', url='http://127.0.0.1:9100')."
        )
    out = ["registered MCPs:"]
    for m in reg["mcps"]:
        tnames = [t.get("name") for t in (m.get("tools") or [])
                  if isinstance(t, dict)]
        out.append(f"  - {m['name']:16s}  url={m['url']}  tools={tnames}")
    return "\n".join(out)


def _mcp_discover_tools(url: str) -> list[dict]:
    """Call the MCP server's tools/list endpoint and return the tool list."""
    base_url = url.rstrip("/")
    if not base_url.endswith("/mcp"):
        base_url = base_url + "/mcp"
    payload = {"jsonrpc": "2.0", "id": 0, "method": "tools/list", "params": {}}
    headers = {"Content-Type": "application/json"}
    try:
        body = json.dumps(payload).encode("utf-8")
        req = urllib.request.Request(base_url, data=body, headers=headers, method="POST")
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = resp.read()
        parsed = json.loads(data)
        if "result" in parsed and "tools" in parsed["result"]:
            tools = []
            for t in parsed["result"]["tools"]:
                if not isinstance(t, dict):
                    continue
                tool_name = t.get("name", "")
                if not tool_name:
                    continue
                params = t.get("inputSchema", t.get("parameters", {"type": "object", "properties": {}}))
                desc = t.get("description", "")
                tools.append({"name": tool_name, "description": desc, "parameters": params})
            return tools
    except Exception:  # noqa: BLE001
        pass
    return []


def _mcp_invoke(server_name: str, tool_name: str, args: dict) -> str:
    """Invoke a tool on a registered MCP server via JSON-RPC. Returns the
    response body as a string (or a tagged error)."""
    reg = _mcp_load()
    entry = next((m for m in reg["mcps"] if m.get("name") == server_name), None)
    if entry is None:
        return f"[mcp error] unknown server {server_name!r}"
    base_url = entry["url"].rstrip("/")
    # Detect whether the URL already ends with /mcp or is a bare base URL.
    if not base_url.endswith("/mcp"):
        base_url = base_url + "/mcp"
    payload = {
        "jsonrpc": "2.0",
        "id": 1,
        "method": "tools/call",
        "params": {"name": tool_name, "arguments": args},
    }
    headers = {"Content-Type": "application/json"}
    headers.update(entry.get("headers") or {})
    try:
        body = json.dumps(payload).encode("utf-8")
        req = urllib.request.Request(base_url, data=body, headers=headers, method="POST")
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = resp.read()
        try:
            parsed = json.loads(data)
        except json.JSONDecodeError:
            return data.decode("utf-8", errors="replace")
        # JSON-RPC response: result is the tool output.
        if "result" in parsed:
            result = parsed["result"]
            if isinstance(result, str):
                return result
            return json.dumps(result, ensure_ascii=False, indent=2)
        if "error" in parsed:
            return f"[mcp error] {json.dumps(parsed['error'], ensure_ascii=False)}"
        return json.dumps(parsed, ensure_ascii=False, indent=2)
    except urllib.error.HTTPError as e:
        try:
            detail = e.read().decode("utf-8", errors="replace")[:400]
        except Exception:  # noqa: BLE001
            detail = ""
        return f"[mcp http {e.code}] {detail}"
    except Exception as e:  # noqa: BLE001
        return f"[mcp error] {type(e).__name__}: {e}"


def _mcp_dispatch_lookup(tool_name: str):
    """Return a callable that proxies to the right MCP, or None."""
    reg = _mcp_load()
    for m in reg["mcps"]:
        for t in (m.get("tools") or []):
            if isinstance(t, dict) and t.get("name") == tool_name:
                server = m["name"]
                def _proxy(_tool_name=tool_name, _server=server, **kwargs):
                    return _mcp_invoke(_server, _tool_name, kwargs)
                return _proxy
    return None


def _mcp_extra_tool_schemas() -> list[dict]:
    """Synthesize OpenAI-shape tool schemas for every registered MCP tool.
    Called lazily by `_filtered_tools` so newly-registered MCPs are visible
    on the next chat turn without a UI restart.
    """
    reg = _mcp_load()
    out: list[dict] = []
    for m in reg["mcps"]:
        for t in (m.get("tools") or []):
            if not isinstance(t, dict) or not t.get("name"):
                continue
            tool_name = t['name']
            params = t.get("parameters") or {"type": "object", "properties": {}}
            desc = t.get("description") or f"{t['name']} (custom MCP: {m['name']})"
            out.append({
                "type": "function",
                "function": {"name": tool_name, "description": desc[:300],
                              "parameters": params},
            })
    return out


# Register the user-facing mcp_* tools so chat agents can manage MCPs from the
# conversation if a UI isn't available.
def mcp_register_tool(name: str, url: str, headers_json: str = "{}",
                       tools_json: str = "[]") -> str:
    """Tool wrapper around mcp_register. headers_json and tools_json are
    JSON strings because tool args are typed."""
    try:
        headers = json.loads(headers_json) if isinstance(headers_json, str) else headers_json
        tools_arg = json.loads(tools_json) if isinstance(tools_json, str) else tools_json
    except json.JSONDecodeError as e:
        return f"[error] could not parse json args: {e}"
    try:
        entry = mcp_register(name, url, headers=headers, tools=tools_arg)
    except ValueError as e:
        return f"[error] {e}"
    return f"registered MCP {entry['name']!r} → {entry['url']} ({len(entry['tools'])} tools)"


def mcp_unregister_tool(name: str) -> str:
    if mcp_unregister(name):
        return f"unregistered MCP {name!r}"
    return f"[not found] no MCP named {name!r}"


DISPATCH["mcp_list"] = mcp_list
DISPATCH["mcp_register"] = mcp_register_tool
DISPATCH["mcp_unregister"] = mcp_unregister_tool


def dispatch(name: str, args: dict[str, Any]) -> str:
    if name in _disabled_tools():
        return f"[tool {name!r} is disabled in this session]"
    fn = DISPATCH.get(name)
    if fn is None:
        # Check the MCP namespace before giving up. Names like
        # `mcp_<server>__<tool>` route to the registered MCP server.
        proxy = _mcp_dispatch_lookup(name)
        if proxy is None:
            return f"[unknown tool] {name}"
        return str(proxy(**args))
    return str(fn(**args))


TOOLS: list[dict] = [
    {
        "type": "function",
        "function": {
            "name": "web_search",
            "description": (
                "Search the web (DuckDuckGo). Returns a numbered list of "
                "results with title, URL, and snippet. Use for up-to-date "
                "information. Built-in semantic-near-duplicate detection: "
                "if you re-issue the same intent with different keywords, "
                "the prior result is reused (cosine ≥ 0.97 or high lexical "
                "overlap), and repeated near-duplicates are refused. Use "
                "site/filetype filters BEFORE rephrasing to broaden recall.\n"
                "QUERY STYLE — keep it simple:\n"
                "- 3-6 keywords. Not a sentence, not a question.\n"
                "- AVOID combining `site:` + quoted multi-word phrases — "
                "that combination almost never matches and triggers a slow "
                "multi-engine fallback. Pick ONE: either `site:` OR a quoted "
                "phrase, not both.\n"
                "- For specific products/pages, search keywords + brand "
                "(e.g. `myprotein impact whey nutrition`), then web_fetch "
                "the most relevant URL from the results."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Search query — keywords, NOT a question.",
                    },
                    "max_results": {
                        "type": "integer",
                        "description": "Max results to return (default 5, max 25).",
                        "default": 5,
                    },
                    "site": {
                        "type": "string",
                        "description": "Restrict to a domain, e.g. 'arxiv.org', 'github.com'.",
                    },
                    "filetype": {
                        "type": "string",
                        "description": "Restrict to a file extension, e.g. 'pdf', 'csv'.",
                    },
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "web_fetch",
            "description": (
                "Fetch a webpage and return its readable text. Static HTTP "
                "first (~1s), Chromium fallback for JS-heavy SPAs. Long "
                "pages are auto-condensed: chunk-ranked by relevance to "
                "your task with table-detection + section-continuity "
                "boosts, so the right sections survive even on 200k+ docs.\n"
                "Set head_only=True if you only need the start (cheaper, "
                "lower noise).\n"
                "PREFER specialized tools when applicable:\n"
                "- SEC filings → use `sec_filings(ticker, form)` to get "
                "the right URL deterministically before fetching.\n"
                "- GitHub URLs → `github_repo` (clean structured data).\n"
                "- arXiv papers → `arxiv_fetch` for abstract/PDF.\n"
                "- DOIs → `doi_resolve` for citation metadata.\n"
                "- PDFs → `pdf_extract` (handles binary properly).\n"
                "Only fetch URLs you got from a prior `web_search` result, "
                "the user's message, or a preapproved docs site — fabricated "
                "product/page slugs will be refused."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "url": {"type": "string", "description": "Absolute URL to fetch."},
                    "force_browser": {
                        "type": "boolean",
                        "description": "Skip the fast static path; always use Chromium. Only set when you know the page is a JS-heavy SPA.",
                    },
                    "head_only": {
                        "type": "boolean",
                        "description": "Return only the head of the body (cheaper, used when you only want the abstract).",
                    },
                    "max_chars": {
                        "type": "integer",
                        "description": "LEAVE UNSET for reference docs (filings, annual reports, papers, long articles). The default (3000000) feeds the full stripped body into the condenser, which picks the most relevant chunks regardless of where they sit in the page. Passing a smaller value chops the body BEFORE the condenser runs and may discard the section you need. Only set when you specifically want less content (e.g. just the first paragraph).",
                    },
                    "mode": {
                        "type": "string",
                        "enum": ["numerical", "semantic"],
                        "description": (
                            "Retrieval profile for the condenser. Default 'semantic'. "
                            "Pick 'numerical' when the answer is stated verbatim in the source — a "
                            "specific value, count, name, date, or list item that you expect to find "
                            "as a literal substring of the fetched content. Uses lexical + structural "
                            "ranking only; tables and numeric content survive intact. "
                            "Pick 'semantic' when the answer is a synthesis, judgment, or inference "
                            "built from prose — derived across multiple sentences rather than stated "
                            "verbatim. Engages a semantic reranker that scores passages by whether "
                            "they actually answer the question, not just by lexical overlap."
                        ),
                    },
                },
                "required": ["url"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "find_in_url",
            "description": (
                "Fetch a URL and return only the lines matching `pattern`, "
                "each with a few surrounding context lines. Use this when "
                "web_fetch's condensed result dropped the specific table "
                "cell, footnote, or row you need. Operates on the FULL "
                "page content, not the condensed view — so you can locate "
                "data values regardless of where they sit in the document."
                "\n"
                "Examples:\n"
                "  find_in_url(<10-K url>, r\"Long-term debt\")\n"
                "  find_in_url(<10-K url>, r\"\\$\\d+,\\d{3}\")\n"
                "  find_in_url(<10-K url>, r\"convertible notes?\", context_lines=8)"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "url": {"type": "string", "description": "Absolute URL (must be a previously-seen URL or on a preapproved host)."},
                    "pattern": {"type": "string", "description": "Python regex (or literal text) to match."},
                    "context_lines": {
                        "type": "integer",
                        "description": "Lines of context before/after each match (default 3). Bump for table data.",
                    },
                    "max_matches": {
                        "type": "integer",
                        "description": "Cap on returned matches (default 20).",
                    },
                    "ignore_case": {
                        "type": "boolean",
                        "description": "Case-insensitive matching (default true).",
                    },
                },
                "required": ["url", "pattern"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "web_outline",
            "description": (
                "Return the heading hierarchy (h1-h4) of a webpage with a "
                "few words of body context per heading. ~30× cheaper than "
                "web_fetch. Use it to decide WHETHER to fetch a page or "
                "to identify the right section before fetching."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "url": {"type": "string", "description": "Absolute URL."},
                    "max_headings": {
                        "type": "integer",
                        "description": "Cap on returned headings (default 80).",
                    },
                },
                "required": ["url"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "arxiv_search",
            "description": (
                "Search arXiv by keyword via the official Atom API. Returns "
                "structured paper metadata (id, title, authors, category, "
                "abstract). Use this INSTEAD of web_search when you want "
                "actual arXiv papers — cleaner output, no scraping noise."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Keywords; arXiv full-text search.",
                    },
                    "max_results": {
                        "type": "integer",
                        "description": "Default 5, max 20.",
                    },
                    "sort_by": {
                        "type": "string",
                        "enum": ["relevance", "submittedDate", "lastUpdatedDate"],
                        "description": "Sort order (default 'relevance').",
                    },
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "arxiv_fetch",
            "description": (
                "Fetch one arXiv paper by id or URL. `what` selects the "
                "format: 'abstract' (~1KB metadata, fastest), 'html' "
                "(arXiv HTML5 rendering — best for sections), or 'pdf' "
                "(text extraction from PDF — fallback). Always start with "
                "'abstract'; only fetch html/pdf when you need the full body."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "id_or_url": {
                        "type": "string",
                        "description": "arXiv id (2401.12345) or any arxiv URL.",
                    },
                    "what": {
                        "type": "string",
                        "enum": ["abstract", "html", "pdf"],
                        "description": "Format to fetch (default 'abstract').",
                    },
                    "max_chars": {
                        "type": "integer",
                        "description": "Cap on returned text (default 60000).",
                    },
                },
                "required": ["id_or_url"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "pdf_extract",
            "description": (
                "Extract text from a PDF — local file path OR https URL. "
                "Use `pages` to limit to a range like '1-3' or '5,7,9-12'. "
                "For papers, fetching pages 1-3 typically gives "
                "title+abstract+intro, which is usually enough."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "path_or_url": {
                        "type": "string",
                        "description": "Local PDF path or https URL.",
                    },
                    "pages": {
                        "type": "string",
                        "description": "1-indexed page selection like '1-3' or '5,7,9-12'. Empty = all pages.",
                    },
                    "max_chars": {
                        "type": "integer",
                        "description": "Cap on returned text (default 80000).",
                    },
                },
                "required": ["path_or_url"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "github_repo",
            "description": (
                "Read a public GitHub repo via the API. `action`: "
                "'info' (metadata), 'list' (directory at `path`), 'read' "
                "(raw file at `path`), 'readme' (README at any common name). "
                "Use this INSTEAD of web_fetch on github.com URLs — direct "
                "API access, no HTML noise, supports branch/tag refs."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "repo": {
                        "type": "string",
                        "description": "owner/name OR a full github URL.",
                    },
                    "action": {
                        "type": "string",
                        "enum": ["info", "list", "read", "readme"],
                        "description": "Default 'info'.",
                    },
                    "path": {
                        "type": "string",
                        "description": "Required for 'read'; optional for 'list'.",
                    },
                    "ref": {
                        "type": "string",
                        "description": "Branch / tag / sha. Default = repo default branch.",
                    },
                    "max_chars": {
                        "type": "integer",
                        "description": "Cap on returned text (default 60000).",
                    },
                },
                "required": ["repo"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "doi_resolve",
            "description": (
                "Resolve a DOI to formatted citation metadata via the "
                "doi.org content negotiation API. Cleaner than fetching "
                "the publisher landing page — no paywalls / cookie banners."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "doi": {
                        "type": "string",
                        "description": "DOI or doi.org URL.",
                    },
                },
                "required": ["doi"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "sec_filings",
            "description": (
                "List recent SEC EDGAR filings for a US-listed company. "
                "Returns direct URLs to filings — far more reliable than "
                "web_search for finding the right 10-K, 10-Q, 8-K, DEF 14A, "
                "S-1, etc. for a specific period. Use this BEFORE web_search "
                "whenever you need an authoritative SEC document; then "
                "web_fetch the URL it returns. Saves search-budget round "
                "trips and avoids third-party scraper noise."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "ticker": {
                        "type": "string",
                        "description": "Stock ticker, e.g. NFLX, AAPL, TSLA.",
                    },
                    "form": {
                        "type": "string",
                        "description": ("Filing type. Common: 10-K (annual), "
                                        "10-Q (quarterly), 8-K (current), "
                                        "DEF 14A (proxy), S-1 (IPO). Default 10-K."),
                    },
                    "limit": {
                        "type": "integer",
                        "description": "How many recent filings to return (default 5).",
                    },
                    "year": {
                        "type": "string",
                        "description": ("Optional 4-digit year filter (e.g. "
                                        "'2024'). Filters by filing-date or "
                                        "report-period year."),
                    },
                },
                "required": ["ticker"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "csv_summary",
            "description": (
                "Quick stats for a CSV/TSV/JSONL file: shape, column types, "
                "first N rows, per-numeric-column min/max/mean/std. Cheaper "
                "than spinning up python_run + pandas for the common "
                "'what's in this file?' question."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Path to .csv/.tsv/.jsonl/.ndjson.",
                    },
                    "max_rows": {
                        "type": "integer",
                        "description": "Preview row count (default 20).",
                    },
                    "describe": {
                        "type": "boolean",
                        "description": "Include numeric describe (default true).",
                    },
                },
                "required": ["path"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "now",
            "description": (
                "Current date/time in any IANA timezone. Use whenever the "
                "user mentions 'today', 'now', 'this week' — the system "
                "prompt has a date but it gets stale across midnight in a "
                "long-running session, and timezones are easy to get wrong."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "tz": {
                        "type": "string",
                        "description": "IANA tz name (default 'UTC'). e.g. 'America/New_York', 'Europe/Berlin'.",
                    },
                    "fmt": {
                        "type": "string",
                        "enum": ["iso", "rfc", "date", "weekday", "epoch"],
                        "description": "Output format (default 'iso').",
                    },
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "read_file",
            "description": (
                "Read a slice of a file. Returns up to 2000 lines by default with "
                "1-indexed line numbers. When you know the rough location, use "
                "offset/limit to read just that section instead of the whole file. "
                "Files larger than 256KB must be read in slices."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Filesystem path. ~ is expanded."},
                    "offset": {"type": "integer", "description": "0-based starting line (default 0)."},
                    "limit": {"type": "integer", "description": "Max lines to return (default 2000)."},
                },
                "required": ["path"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "list_files",
            "description": (
                "List files under a path, optionally filtered by a glob pattern. "
                "Use this to discover the layout of a project before reading files. "
                "SCOPE: stay within the current working directory (the active "
                "project tree). Do NOT list the home directory itself, '/', "
                "'/Users', '/Library', etc. — those expose personal data, "
                "credentials, and app state. If the file you want isn't in "
                "the project, fetch from the web or open it by full path with "
                "read_file."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Directory to list (default: cwd). Must be inside the project tree — broad scopes like '~', '/', or '/Users' are off limits."},
                    "pattern": {
                        "type": "string",
                        "description": "Glob pattern, e.g. '**/*.py' or 'src/**/*.ts' (default: '**/*').",
                    },
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "grep",
            "description": (
                "Regex-search file contents under a path. Use this to locate code "
                "by symbol or keyword before reading whole files. Always prefer "
                "grep over reading entire files. VCS dirs (.git/.hg/etc) are "
                "excluded automatically. "
                "SCOPE: never grep the home directory itself, '/', '/Users', "
                "'/Library', '/System', '/etc', '/var', '/usr', or any other "
                "top-level system / user dir — those hold ssh keys, browser "
                "data, app state, and personal documents which must not be "
                "pulled into context. Default to the current working directory "
                "(the active project) or a named subfolder under it. If the "
                "data you need isn't in the project tree, you're using the "
                "wrong tool — use web_fetch / sec_filings / read_file on a "
                "specific path instead."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "pattern": {"type": "string", "description": "Regex pattern."},
                    "path": {"type": "string", "description": "Directory or file (default: cwd). Stay within the project tree — the home dir, '/', and system roots are off limits."},
                    "glob": {
                        "type": "string",
                        "description": "Restrict to files matching this glob (e.g. '*.py').",
                    },
                    "output_mode": {
                        "type": "string",
                        "enum": ["files_with_matches", "content", "count"],
                        "description": (
                            "files_with_matches → just file paths (cheapest, use first); "
                            "content → matching lines with file:line: prefix (default); "
                            "count → match counts per file."
                        ),
                    },
                    "head_limit": {
                        "type": "integer",
                        "description": "Cap output lines (default 250). Pass 0 for unlimited (avoid).",
                    },
                },
                "required": ["pattern"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "write_file",
            "description": (
                "Create or overwrite a file with the given content. Prefer edit_file for "
                "modifying existing files; use write_file for new files or full rewrites."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Filesystem path."},
                    "content": {"type": "string", "description": "Full file content."},
                },
                "required": ["path", "content"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "edit_file",
            "description": (
                "Surgically replace one exact substring in a file. old_string must appear "
                "exactly once — if not, the call fails and you should add surrounding context "
                "to make it unique. Preferred over write_file for in-place edits."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Filesystem path."},
                    "old_string": {"type": "string", "description": "Exact text to replace."},
                    "new_string": {"type": "string", "description": "Replacement text."},
                },
                "required": ["path", "old_string", "new_string"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "bash",
            "description": (
                "Run a shell command in the current working directory. Use for "
                "builds, tests, git, package managers, etc. Output captured and "
                "truncated; default timeout 60s.\n\n"
                "Avoid using bash to substitute for dedicated tools — it pollutes "
                "your context: read files with read_file (NOT cat/head/tail), "
                "search with grep (NOT grep/rg/ack), list dirs with list_files "
                "(NOT ls/find), edit with edit_file (NOT sed/awk)."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "command": {"type": "string", "description": "Shell command to execute."},
                    "timeout": {
                        "type": "integer",
                        "description": "Max seconds before the command is killed (default 60).",
                    },
                },
                "required": ["command"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "explore",
            "description": (
                "Spawn a read-only research subagent in an isolated context to "
                "answer a broad question, returning only its concise summary. "
                "Use this for open-ended questions — 'how does X work?', 'where "
                "is Y handled?', 'is the code overfit?', 'summarize this codebase' "
                "— so the search noise stays out of your context. The subagent "
                "has list_files, grep, read_file, web_search, web_fetch and "
                "cannot modify files. Prefer explore over a long sequence of "
                "your own grep/read calls when the question is exploratory."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "question": {
                        "type": "string",
                        "description": "Self-contained question for the subagent. Be specific.",
                    },
                    "max_steps": {
                        "type": "integer",
                        "description": (
                            "Subagent step budget (default 25). The subagent "
                            "self-prunes context at 40k tokens and gets a "
                            "halftime nudge to commit. Most questions resolve "
                            "in 5-10 steps; raise only for genuinely broad "
                            "research questions."
                        ),
                    },
                },
                "required": ["question"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "memory_save",
            "description": (
                "Save a piece of knowledge to persistent memory (across "
                "sessions). Use when you learn something durable: a "
                "decision, a non-obvious fact about a codebase, a user "
                "preference, a debugging breakthrough. The content is "
                "embedded once at save time so semantic search later is "
                "cheap. Pick a short descriptive `key` like "
                "'qwen-mtp-prefill-chunk-size' or 'user-prefers-tabs'. "
                "Updating an existing key overwrites it (use a fresh key "
                "if you want both)."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "key": {
                        "type": "string",
                        "description": "Short descriptive identifier.",
                    },
                    "content": {
                        "type": "string",
                        "description": "The knowledge to store. Be concrete (file:line, numbers, names).",
                    },
                    "tags": {
                        "type": "string",
                        "description": "Optional space- or comma-separated tags for filtering later (e.g. 'project:qwen perf').",
                    },
                },
                "required": ["key", "content"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "scratchpad",
            "description": (
                "In-session working notes that don't pollute long-term memory. "
                "Use for `what I'm about to check`, `what this fetch returned`, "
                "or `decisions made so far` — anything you'll re-read within "
                "this task. Cleared on /clear or new headless invocation. "
                "Choose memory_save instead only when the insight is durable "
                "across sessions (a fact, a pattern, a fix)."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "action": {
                        "type": "string",
                        "enum": ["append", "read", "clear", "list"],
                        "description": "append (default) | read | clear | list keys",
                    },
                    "content": {
                        "type": "string",
                        "description": "The note text (required for append).",
                    },
                    "key": {
                        "type": "string",
                        "description": "Optional named pad (default 'default'). Use '*' with read/clear for all pads.",
                    },
                },
                "required": ["action"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "ask_user",
            "description": (
                "Ask the user a clarifying question. ONLY when the request is "
                "ambiguous in a way that changes the artifact you produce — "
                "don't ask for information you can infer or look up. In "
                "headless/eval sessions there's no human to ask; the tool "
                "returns a `[no-user] proceed with best inference` marker "
                "and you must continue with your best guess (do NOT loop on "
                "asking)."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "question": {
                        "type": "string",
                        "description": "Short, specific question (1-2 sentences).",
                    },
                    "options": {
                        "type": "string",
                        "description": "Optional comma-separated suggested choices the UI can render as buttons.",
                    },
                },
                "required": ["question"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "graph_compose",
            "description": (
                "Design + save a fresh multi-agent graph from a natural-"
                "language description, optionally running it immediately. "
                "Use when none of the graphs from agent_graph_list() fit "
                "but the task decomposes cleanly into research → analyze → "
                "produce. Saves a new file under examples/<auto-name>_graph.py."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "description": {
                        "type": "string",
                        "description": "Paragraph describing the pipeline's purpose and shape (5-15 sentences).",
                    },
                    "run": {
                        "type": "boolean",
                        "description": "If true, immediately invoke the saved graph with `inputs`.",
                    },
                    "inputs": {
                        "type": "string",
                        "description": "JSON object string for the graph's initial inputs (only used when run=true).",
                    },
                },
                "required": ["description"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "memory_search",
            "description": (
                "Semantic search across all stored memories using vector "
                "similarity. Returns up to max_results entries that score "
                "above min_score (cosine, default 0.55). When NO memory is "
                "above the threshold, the response says so explicitly and "
                "lists the closest few as BELOW THRESHOLD — those are not "
                "real recall and you should not treat them as authoritative. "
                "Cosine ≥ 0.7 is a strong topical match; 0.55–0.7 is borderline; "
                "< 0.55 is usually unrelated."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Natural-language query. E.g. 'how to fix Metal OOM at long context'.",
                    },
                    "max_results": {
                        "type": "integer",
                        "description": "Top-K to return after filtering (default 5).",
                    },
                    "tag": {
                        "type": "string",
                        "description": "Optional substring filter on tags (e.g. 'project:qwen').",
                    },
                    "min_score": {
                        "type": "number",
                        "description": "Cosine threshold (default 0.55). Lower (e.g. 0.4) for fuzzy recall; raise (e.g. 0.7) when you only want strong matches.",
                    },
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "memory_get",
            "description": (
                "Fetch a specific memory entry by exact key. Use when you "
                "remember the key. For fuzzy/conceptual lookup use memory_search."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "key": {"type": "string", "description": "Exact key passed to memory_save."},
                },
                "required": ["key"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "memory_list",
            "description": (
                "List recently-updated memory entries (just the keys + "
                "timestamps, no embedding work). Useful to discover what "
                "knowledge already exists before saving duplicates."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "limit": {"type": "integer", "description": "Max entries (default 20)."},
                    "tag": {"type": "string", "description": "Optional tag substring filter."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "memory_delete",
            "description": "Delete a memory entry by exact key.",
            "parameters": {
                "type": "object",
                "properties": {
                    "key": {"type": "string", "description": "Exact key to remove."},
                },
                "required": ["key"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "todo_write",
            "description": (
                "Maintain a structured task list for the current session. "
                "Use proactively for any non-trivial multi-step task (3+ steps) — "
                "write the list once at the start, then update statuses as you go. "
                "Each todo: {content, activeForm, status} where status is "
                "'pending' | 'in_progress' | 'completed'. Exactly one task should be "
                "in_progress at a time. Mark complete IMMEDIATELY after finishing — "
                "don't batch completions. The list is persisted across turns so you "
                "can re-read it. Skip for trivial single-step requests."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "todos": {
                        "type": "array",
                        "description": "Full updated todo list (this REPLACES the previous list).",
                        "items": {
                            "type": "object",
                            "properties": {
                                "content": {"type": "string", "description": "Imperative task description (e.g. 'Run tests')."},
                                "activeForm": {"type": "string", "description": "Present-continuous form (e.g. 'Running tests')."},
                                "status": {"type": "string", "enum": ["pending", "in_progress", "completed"]},
                            },
                            "required": ["content", "status"],
                        },
                    }
                },
                "required": ["todos"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "enter_worktree",
            "description": (
                "Create an isolated git worktree (or sandbox copy if not in a git "
                "repo) and switch the working directory into it. Subsequent "
                "edit_file/write_file/bash operate on the copy, leaving the original "
                "tree completely untouched. Use this BEFORE any task that involves "
                "modifying files when you must not damage the originals — refactors, "
                "experiments, bug fixes you want to verify before committing, or "
                "tasks where the user has explicitly said the originals are off-limits. "
                "Always pair with exit_worktree to clean up."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "name": {
                        "type": "string",
                        "description": "Optional slug for the worktree branch/dir (auto-generated if omitted).",
                    },
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "exit_worktree",
            "description": (
                "Leave the current worktree, restore the original cwd. By default "
                "removes the worktree and its branch (clean slate). Pass keep=true "
                "to preserve the worktree files for inspection."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "keep": {
                        "type": "boolean",
                        "description": "If true, keep the worktree dir/branch instead of removing it.",
                    },
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "notebook_edit",
            "description": (
                "Edit a Jupyter notebook (.ipynb). Auto-creates the file when action "
                "is append/insert. Specify cell by cell_index (0-based) or cell_id. "
                "Use this for iterative data exploration where you want to keep cells "
                "as discrete units (load → analyze → plot → save) rather than one "
                "monolithic .py file. After edits, call notebook_run to execute."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Path to the .ipynb file."},
                    "source": {"type": "string", "description": "Cell source (Python code or markdown)."},
                    "cell_index": {"type": "integer", "description": "Target cell index (0-based, default -1 = none)."},
                    "cell_id": {"type": "string", "description": "Target cell id (alternative to index)."},
                    "action": {
                        "type": "string",
                        "enum": ["replace", "insert_after", "insert_before", "delete", "append"],
                        "description": "Edit operation. Default 'replace'.",
                    },
                    "cell_type": {
                        "type": "string",
                        "enum": ["code", "markdown"],
                        "description": "Type of new cell when inserting/appending (default 'code').",
                    },
                },
                "required": ["path"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "notebook_run",
            "description": (
                "Execute every cell of a Jupyter notebook in a fresh python3 kernel, "
                "save outputs back into the .ipynb, and return the textual outputs "
                "(stream, results, errors). Use after notebook_edit to verify cells "
                "execute successfully. Display outputs (matplotlib figures) are "
                "noted by mime type — actual images stay in the .ipynb."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Path to the .ipynb file."},
                    "timeout": {
                        "type": "integer",
                        "description": "Per-cell timeout in seconds (default 120).",
                    },
                },
                "required": ["path"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "python_run",
            "description": (
                "Execute Python code in a PERSISTENT kernel that lives for the entire "
                "agent session. Variables, imports, and open file handles persist "
                "across calls — so you can iteratively explore data, build an analysis "
                "incrementally, or debug interactively without restarting Python. "
                "Outputs (stdout, return values, errors with tracebacks) are streamed back. "
                "Prefer this over `bash python -c '...'` for any non-trivial computation: "
                "it's faster (no startup), keeps state, and gives cleaner errors. "
                "For matplotlib, use `plt.savefig('path.png')` — figure display is captured "
                "as a marker only."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "code": {"type": "string", "description": "Python code to execute. Multi-line is fine."},
                    "timeout": {
                        "type": "integer",
                        "description": "Max seconds before the cell is interrupted (kernel kept alive). Default 60.",
                    },
                },
                "required": ["code"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "python_reset",
            "description": (
                "Restart the persistent Python kernel, clearing all state. Use when "
                "you've gotten the kernel into a confused state (bad imports, large "
                "objects you can't free, etc.) and want a fresh start without "
                "exiting the agent."
            ),
            "parameters": {"type": "object", "properties": {}},
        },
    },
    {
        "type": "function",
        "function": {
            "name": "test_run",
            "description": (
                "Run pytest with structured output: counts of passed/failed/errored "
                "tests, plus short tracebacks for failures. Cleaner than `bash pytest` "
                "for understanding what broke. Pass a path (file or dir) and optionally "
                "a -k expression to select a subset."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "File or directory of tests (default: cwd)."},
                    "k": {"type": "string", "description": "pytest -k expression (e.g. 'test_concurrent and not slow')."},
                    "timeout": {
                        "type": "integer",
                        "description": "Max seconds for the entire test run (default 90).",
                    },
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "append_finding",
            "description": (
                "Append a `## heading` + content section to a markdown artifact "
                "(creates the file with a # title on first call). Use this WHILE "
                "investigating — commit each finding as you discover it instead of "
                "batching one big write_file at the end. Cheap to call, low friction, "
                "and breaks the 'read 12 files write 0' anti-pattern: by the time "
                "you've finished reading, your artifact is already mostly written. "
                "Idempotent on duplicate sections. Pair with write_file only when "
                "you need to overwrite the whole file."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Markdown file path."},
                    "heading": {"type": "string", "description": "Section heading (no '##' prefix; we add it)."},
                    "content": {"type": "string", "description": "Section body (markdown)."},
                    "create_with_title": {
                        "type": "string",
                        "description": "Optional `# Top-Level Title` for first-call. If omitted, derived from filename.",
                    },
                },
                "required": ["path", "heading", "content"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "write_file_verified",
            "description": (
                "Atomic write + Python verification. Writes `content` to `path`, "
                "then runs `verifier_code` in the persistent Python kernel. "
                "If the verifier raises ANY exception (AssertionError, NameError, "
                "etc.), the file is REVERTED and an error is returned. Use this "
                "whenever your artifact contains a numerical, algebraic, or "
                "structural claim that you can re-derive in Python — sequence values, "
                "formulas, identities, invariants. The verifier should be self-contained "
                "(imports + helper functions + assertions in one block). Catches "
                "wrong-value commits before they harden into final artifacts."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "File to write."},
                    "content": {"type": "string", "description": "File content."},
                    "verifier_code": {
                        "type": "string",
                        "description": (
                            "Self-contained Python that re-derives the claim and "
                            "asserts it. If it raises, the write is reverted. "
                            "Example: `def a(n): ...; assert [a(n) for n in range(1,11)] == [1,2,0,3,...]`."
                        ),
                    },
                    "timeout": {
                        "type": "integer",
                        "description": "Max seconds for verifier (default 30).",
                    },
                },
                "required": ["path", "content", "verifier_code"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "apply_patch",
            "description": (
                "Apply a unified diff to one or more files. Use INSTEAD of "
                "write_file when MODIFYING an existing file: emit only the "
                "changed lines as a diff, not the whole file. 5-20× faster "
                "on edits because you skip re-generating unchanged content. "
                "Format: standard unified diff with `--- a/path`, `+++ b/path`, "
                "`@@ -a,b +c,d @@`, and context/`-`/`+` lines. Use "
                "`--- /dev/null` for new files. Multiple file diffs in one "
                "patch are supported (concatenate the headers). All-or-nothing: "
                "any hunk failure aborts the whole patch. Always read_file "
                "first to confirm the exact context lines."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "patch": {
                        "type": "string",
                        "description": (
                            "Complete unified diff. Must include `--- a/path` "
                            "and `+++ b/path` headers and at least one `@@` "
                            "hunk. Context lines must match the existing file "
                            "exactly (whitespace included)."
                        ),
                    },
                },
                "required": ["patch"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "subagent_implement",
            "description": (
                "Run a write-capable subagent in an isolated context for a "
                "focused code-gen / edit task. Returns only the subagent's "
                "final summary. Use this for multi-step edits (read+edit+test "
                "loops) so the iteration noise stays out of YOUR context. "
                "Compounds with apply_patch (the subagent can use it). The "
                "subagent has full file/edit/python/bash/test tools but "
                "cannot spawn further subagents. Step budget: 20 by default."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "task": {
                        "type": "string",
                        "description": (
                            "Self-contained imperative description of the task. "
                            "Be specific about expected file paths and outcome. "
                            "E.g. 'Add a parse_csv(path) function to /tmp/foo.py "
                            "that returns list[dict]; add a test in test_foo.py'."
                        ),
                    },
                    "files": {
                        "type": "string",
                        "description": (
                            "Optional newline- or comma-separated list of "
                            "relevant file paths. Helps the subagent's first "
                            "reads land in the right places."
                        ),
                    },
                },
                "required": ["task"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "inspect_data",
            "description": (
                "Auto-summarize a data file by extension: CSV/TSV (shape + dtypes + "
                "head + numeric stats), JSON (shape + key sample), Parquet (schema "
                "+ rows), Notebook (cell outline), YAML (key types), Excel (sheet "
                "shapes), or any other file as plain text head. Use this BEFORE "
                "read_file when you've been handed a path and need to understand "
                "the data without pulling the whole file into context."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Path to a data file or directory.",
                    },
                    "max_chars": {
                        "type": "integer",
                        "description": "Cap on returned text for the plain-text fallback (default 4000).",
                    },
                },
                "required": ["path"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "make_table",
            "description": (
                "Render a deterministic Markdown table from headers + rows. "
                "Use this whenever you'd otherwise hand-write a table — the "
                "tool guarantees correct pipe count, separator row, and "
                "alignment, and pre-escapes pipes/newlines inside cells. "
                "Output is a string ready to drop into your final answer."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "headers": {
                        "type": "array", "items": {"type": "string"},
                        "description": "Column titles (strings).",
                    },
                    "rows": {
                        "type": "array",
                        "items": {"type": "array"},
                        "description": "List of rows; each row is a list of cells.",
                    },
                    "align": {
                        "type": "string",
                        "description": "Either 'left'|'right'|'center' for all columns, or a comma list like 'left,right,center' per column.",
                    },
                    "title": {
                        "type": "string",
                        "description": "Optional ## title rendered above the table.",
                    },
                    "numbered": {
                        "type": "boolean",
                        "description": "If true, prepend a '#' column with 1-based row numbers.",
                    },
                },
                "required": ["headers", "rows"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "mcp_list",
            "description": (
                "List currently registered custom MCP servers and their tools. "
                "Each MCP is exposed in the chat as `mcp_<server>__<tool>`. "
                "Call this when the user asks 'what MCPs do I have' or before "
                "registering a new one to avoid name collisions."
            ),
            "parameters": {"type": "object", "properties": {}},
        },
    },
    {
        "type": "function",
        "function": {
            "name": "mcp_register",
            "description": (
                "Register a custom MCP server in the persistent registry "
                "(~/.qwen/mcps.json). The server must respond to "
                "POST <url>/tools/<tool> with JSON args and return JSON. Tools "
                "become callable as `mcp_<name>__<tool>` on the next chat turn."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "name": {"type": "string", "description": "lowercase short name, e.g. 'weather'."},
                    "url": {"type": "string", "description": "Base URL, e.g. http://127.0.0.1:9100."},
                    "headers_json": {"type": "string", "description": "Optional JSON object string of extra headers (auth tokens, etc)."},
                    "tools_json": {"type": "string", "description": "JSON array of tool descriptors: [{\"name\": str, \"description\": str, \"parameters\": <JSON-schema>}]."},
                },
                "required": ["name", "url"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "mcp_unregister",
            "description": "Remove a registered MCP server by name.",
            "parameters": {
                "type": "object",
                "properties": {
                    "name": {"type": "string", "description": "Server name to remove."},
                },
                "required": ["name"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "agent_graph_list",
            "description": (
                "List the predefined multi-agent graphs available on this system. "
                "Each graph decomposes a complex task into specialized agents whose "
                "contexts stay compartmentalized. Returns a short table of graph "
                "names and node counts. Call this when the user asks for something "
                "broad (research a market, review a file, etc.) and you want to "
                "see if a tailored graph already exists."
            ),
            "parameters": {"type": "object", "properties": {}},
        },
    },
    {
        "type": "function",
        "function": {
            "name": "agent_graph_run",
            "description": (
                "Run a multi-agent graph and return its compact output summary. "
                "Each node is a specialized agent with a tailored system prompt and "
                "tool subset; outputs flow between them via AGFMT (token-efficient "
                "structured format). Use this for tasks that decompose naturally "
                "into research → analyze → produce steps and would otherwise blow "
                "your context window. Call agent_graph_list first to see options."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "graph": {
                        "type": "string",
                        "description": "Graph name (e.g. 'market_research') or absolute path to a *_graph.py file.",
                    },
                    "inputs": {
                        "type": "string",
                        "description": "JSON object string with the graph's initial inputs, e.g. '{\"topic\": \"US equities today\"}'.",
                    },
                    "max_parallel": {
                        "type": "integer",
                        "description": "Max nodes to run in parallel (default 4).",
                        "default": 4,
                    },
                },
                "required": ["graph", "inputs"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "done",
            "description": (
                "Signal that the task is complete — call this as your FINAL tool call. "
                "The harness reads this signal and gracefully closes the session — "
                "no more LLM rounds, no waiting for /exit. Pass a 1-2 sentence summary "
                "naming the deliverable and where it lives. SAFETY: this tool will REFUSE "
                "to mark complete if no artifact files were written this session. If you "
                "see 'no writes recorded' you haven't done the task yet — write the "
                "requested deliverable first, then call done() again."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "summary": {
                        "type": "string",
                        "description": "1-2 sentence completion summary (artifact path + brief description).",
                    },
                },
                "required": ["summary"],
            },
        },
    },
]


# Memoize the (env-disabled-set → filtered list) result. The chat path
# rebuilds this list on every chat request and the agent CLI rebuilds it
# on every tool turn — TOOLS itself is constant for the process lifetime,
# so the env-keyed cache is safe and saves a few microseconds per call.
# More importantly, callers that JSON-encode the result downstream now
# share the SAME list object, which lets json.dumps reuse its serializer
# state and lets upstream proxies recognize identical input bytes for
# their own cache lookups.
_FILTERED_TOOLS_MEMO: dict[frozenset, list[dict]] = {}


def _filtered_tools() -> list[dict]:
    """Return TOOLS (built-ins + registered MCP tools) minus any whose name is
    in QWEN_AGENT_TOOLS_DISABLE.

    MCP tool schemas are NOT cached because the registry can change at any
    moment (POST /api/mcps/register), and their cardinality is small.
    Built-ins are still cached via the disabled-set memo.
    """
    disabled = _disabled_tools()
    key = frozenset(disabled)
    cached = _FILTERED_TOOLS_MEMO.get(key)
    if cached is not None:
        base = cached
    else:
        if not disabled:
            base = TOOLS
        else:
            base = [t for t in TOOLS
                    if t.get("function", {}).get("name") not in disabled]
        _FILTERED_TOOLS_MEMO[key] = base
    mcp_extras = [t for t in _mcp_extra_tool_schemas()
                  if t.get("function", {}).get("name") not in disabled]
    if not mcp_extras:
        return base
    return list(base) + mcp_extras


# --------------------------------------------------------------------------
# embedder pre-warm — load BGE-small at module import optionally so the
# first memory_save / web_search call doesn't pay a 0.5-2s cold-start cost.
# Opt in with QWEN_PREWARM_EMBED=1 (default off so the agent CLI starts
# fast). The chat UI sets it explicitly at boot since it's long-running.

def prewarm_embedder() -> str | None:
    """Load the embedding model + tokenizer eagerly. Returns None on
    success or an error message. Safe to call multiple times — a second
    call is a no-op (the loader caches in `_embed_state`)."""
    try:
        _embed_load()
        return None
    except Exception as e:  # noqa: BLE001
        return f"{type(e).__name__}: {e}"
